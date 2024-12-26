
#!/usr/bin/env python

import os
import sys
import optparse
import random
from numpy import inf
import numpy as np
import matplotlib.pyplot as plt
from datetime import datetime

import sumolib #To query node/edge stuff about the network
import pickle #To save/load training data

import torch
from torch import nn
from torch.utils.data import Dataset, DataLoader
#import torch.multiprocessing
#torch.multiprocessing.set_sharing_strategy("file_system")

#import runnerQueueSplit
import runnerQueueSplit27 as runnerQueueSplit #KEEP THIS UP TO DATE!!! (If training from a network, not just IG)
import intersectionGeneratorBlocks15 as intersectionGenerator
from importlib import reload
from Net import Net

import openpyxl #For writing training data to .xlsx files
import time

#Dumps training data to an Excel file for human readability
def dumpTrainingData(trainingdata):
    timeout = 60
    starttime = time.time()
    print("Writing training data to spreadsheet")
    try:
        book = openpyxl.Workbook()
        sheets = dict()
        for light in trainingdata:
            sheets[light] = book.create_sheet(light, -1)
            sheets[light].cell(1, 1, "Input")
            row = 2
            for batch in trainingdata[light]:
                if time.time() - starttime > timeout:
                    print("Timed out while writing data to Excel")
                    assert(False) #This should trigger the try-catch and get us out of here
                for linenum in range(batch[0].size(0)):
                    col = 1
                    for entry in batch[0][linenum]:
                        sheets[light].cell(row, col, float(entry))
                        col += 1
                    if row == 2:
                        sheets[light].cell(1, col, "Expert Output")
                    sheets[light].cell(row, col, float(batch[1][linenum]))

                    col += 1
                    if row == 2:
                        sheets[light].cell(1, col, "NN Output")
                    if len(batch) > 2:
                        for outel in range(batch[2].size(3)):
                            sheets[light].cell(row, col, float(batch[2][linenum][0][0][outel]))
                            col += 1
                    row += 1
    except Exception as e:
        print(e)
        print("Error dumping training data to Excel, ignoring and continuing")
    finally:
        book.save("trainingdata/trainingdata_"+sys.argv[1]+".xlsx")

def main(filename):
    datafile = "trainingdata/trainingdata_"+sys.argv[1]+".pickle"
    print(datafile)

    with open(datafile, 'rb') as handle:
        trainingdata = pickle.load(handle)
    dumpTrainingData(trainingdata)

# this is the main entry point of this script
if __name__ == "__main__":
    main(sys.argv[1])