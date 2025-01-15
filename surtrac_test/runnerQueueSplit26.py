#!/usr/bin/env python
#Adapted from TraCI tutorial here: https://github.com/eclipse/sumo/blob/main/tests/complex/tutorial/traci_tls/runner.py

#QueueSplit5 added a first iteration of a Surtrac model to QueueSplit4
#New in QueueSplit6: When there's multiple light phases a lane can go in, don't double-create clusters
#New in QueueSplit7: Adding Surtrac to simulate-ahead
#New in QueueSplit8: Adding communication between intersections. Which then requires re-compacting clusters, etc. For now, don't take advantage of known routes
#New in QueueSplit9: Take advantage of known routes from vehicle routing (NOTE: route may change as vehicle approaches intersection...)
#New in QueueSplit10: Optimize for speed (calling Surtrac less often in routing, merging predicted clusters)
#New in QueueSplit11: Compute delay, not just average travel time, for cars. Also fixed a bug with Surtrac code DP (was removing sequences it shouldn't have)
#QueueSplit12: Multithread the Surtrac code (it's really slow otherwise). Also, use the full Surtrac schedule rather than assuming we'll update every timestep
#QueueSplit13: Surtrac now (correctly) no longer overwrites all the finish times of other lanes with the start time of the currently scheduled cluster (leads to problems when a long cluster, then compatible short cluster, get scheduled, as the next cluster can then start earlier than it should). VOI now gets split into all lanes on starting edge
#14: Anytime routing, better stats on timeouts and teleports, added mingap to all cluster durations, using a timestep that divides mingap and surtracFreq, opposing traffic blocks for mingap not just one timestep, combining clusters at lights into a single queue
#15: Surtrac now runs once, assuming all vehicles travel their same routes, and then gets looked up during routing simulations. Also adding a flag to disable predicting clusters with Surtrac. REAL-TIME PERFORMANCE!!!
#16: Be lazy in routing simulations - don't need to check a lane until the front vehicle gets close to the end. (Didn't seem to give much speedup though.) Reusing Surtrac between multiple timesteps (and backported this to 15). Various attempts at cleanup and speedup.
#17: Move computation of Surtrac's predicted outflows to the end, after we already know what the schedule should be, rather than building it into the algorithm and generating predicted outflows for schedules we won't end up using. Fixed a bug where initially splitting the VOI into all lanes added it at the modified start time, and possibly in the wrong place
#18: Add imitation learning version of Surtrac
#19: Going back to A* in SUMO since ghost cars are causing problems. Telling all ghost cars to turn left, and spawning left-turning ghost cars when the turn completes so we account for oncoming traffic
#20: Storing Surtrac results for reuse
#21: Libsumo (optional hopefully), multithreaded routing, save files under different names so we can run two sets of code without making one error
#22: Did routing before Surtrac for more realism (in case we're not simultaneous so Surtrac's out of date). Not much effect there.
#23: 21, but now using a detector model to reconstruct the traffic state at the start of each routing simulation
#24: Detector model stops tracking specific names of non-adopters
#25: New plan for lane changes - blindly sample which lane stuff ends up in
#26: Detector model for Surtrac in routing as well (since the goal is to approximate what the main simulation would be doing).

from __future__ import absolute_import
from __future__ import print_function

import torch
from torch import nn

import os
import sys
import optparse
import random
from numpy import inf
import numpy as np
import time
import matplotlib.pyplot as plt
import math
from copy import deepcopy, copy
from collections import Counter
from heapq import * #priorityqueue
import threading
import xml.etree.ElementTree as ET

# we need to import python modules from the $SUMO_HOME/tools directory
if 'SUMO_HOME' in os.environ:
    tools = os.path.join(os.environ['SUMO_HOME'], 'tools')
    sys.path.append(tools)
else:
    sys.exit("please declare environment variable 'SUMO_HOME'")

from sumolib import checkBinary

useLibsumo = True
if useLibsumo:
    import libsumo as traci
else:
    import traci  #To interface with SUMO simulations

import sumolib #To query node/edge stuff about the network
import pickle #To save/load traffic light states

from Net import Net
import openpyxl #For writing training data to .xlsx files

from multiprocessing import Process
import multiprocessing

try:
    multiprocessing.set_start_method("fork")
except:
    pass

manager = multiprocessing.Manager()
sumoconfig = None

pSmart = 1.0 #Adoption probability
useLastRNGState = False #To rerun the last simulation without changing the seed on the random number generator

clusterthresh = 5 #Time between cars before we split to separate clusters
mingap = 2.5 #Minimum allowed space between cars
timestep = 1 #Amount of time between updates. In practice, mingap rounds up to the nearest multiple of this #NOTE: I'm pretty sure this used to be the timestep length in routing simulations, but I've since just started using SUMO with the default timestep of 1. timestep clearly is still in the code, but I'm not sure what it does anymore
detectordist = 50 #How far before the end of a road the detectors that trigger reroutes are
simdetectordist = 0 #How far after the start of a road the detectors for reconstructing initial routing sim traffic state are. TODO I'm not actually using this when making detectors, I just assume they're at start of lane. But then they miss all the cars, so I'm just faking those detectors anyway

#Hyperparameters for multithreading
multithreadRouting = True #Do each routing simulation in a separate thread. Enable for speed, but can mess with profiling
if not useLibsumo:
    multithreadRouting = False
multithreadSurtrac = False #Compute each light's Surtrac schedule in a separate thread. Enable for speed, but can mess with profiling
reuseSurtrac = False #Does Surtrac computations in a separate thread, shared between all vehicles doing routing. Keep this true unless we need everything single-threaded (ex: for debugging), or if running with fixed timing plans (routingSurtracFreq is huge) to avoid doing this computation
debugMode = True #Enables some sanity checks and assert statements that are somewhat slow but helpful for debugging
simToSimStats = False
routingSimUsesSUMO = True #Only switch this if we go back to custom routing simulator or something
mainSurtracFreq = 1 #Recompute Surtrac schedules every this many seconds in the main simulation (technically a period not a frequency). Use something huge like 1e6 to disable Surtrac and default to fixed timing plans.
routingSurtracFreq = 1 #Recompute Surtrac schedules every this many seconds in the main simulation (technically a period not a frequency). Use something huge like 1e6 to disable Surtrac and default to fixed timing plans.
recomputeRoutingSurtracFreq = 1 #Maintain the previously-computed Surtrac schedules for all vehicles routing less than this many seconds in the main simulation. Set to 1 to only reuse results within the same timestep. Does nothing when reuseSurtrac is False.
disableSurtracPred = True #Speeds up code by having Surtrac no longer predict future clusters for neighboring intersections
predCutoffMain = 10 #Surtrac receives communications about clusters arriving this far into the future in the main simulation
predCutoffRouting = 10 #Surtrac receives communications about clusters arriving this far into the future in the routing simulations
predDiscount = 1 #Multiply predicted vehicle weights by this because we're not actually sure what they're doing. 0 to ignore predictions, 1 to treat them the same as normal cars.

testNNdefault = False #Uses NN over Dumbtrac for light control if both are true
noNNinMain = True
debugNNslowness = False #Prints context information whenever loadClusters is slow, and runs the NN 1% of the time ignoring other NN settings
testDumbtrac = False #If true, overrides Surtrac with Dumbtrac (FTP or actuated control) in simulations and training data (if appendTrainingData is also true)
FTP = True #If false, and testDumbtrac = True, runs actuated control instead of fixed timing plans. If true, runs fixed timing plans (should now be same as SUMO defaults)
resetTrainingData = False
appendTrainingData = False

detectorModel = True
detectorSurtrac = detectorModel
detectorRouting = detectorModel
detectorRoutingSurtrac = detectorModel #If false, uses omniscient Surtrac in routing regardless of detectorSurtrac. If true, defers to detectorSurtrac
adopterComms = True
adopterCommsSurtrac = adopterComms
adopterCommsRouting = adopterComms

testNNrolls = []
nVehicles = []

learnYellow = False #False to strictly enforce that yellow lights are always their minimum length (no scheduling clusters during yellow+turn arrow, and ML solution isn't used there)
learnMinMaxDurations = False #False to strictly enforce min/max duration limits (in particular, don't call ML, just do the right thing)

#Don't change parameters below here
#For testing durations to see if there's drift between fixed timing plans executed in main simulation and routing simulations.
simdurations = dict()
simdurationsUsed = False
realdurations = dict()

dumpIntersectionData = False
intersectionData = dict()
vehicleIntersectionData = dict()

max_edge_speed = 0.0 #Overwritten when we read the route file

lanes = []
edges = []
carsOnNetwork = []
oldids = dict()
isSmart = dict() #Store whether each vehicle does our routing or not
lightphasedata = dict()
lightlinks = dict()
prioritygreenlightlinks = dict()
lowprioritygreenlightlinks = dict()
prioritygreenlightlinksLE = dict()
lowprioritygreenlightlinksLE = dict()
lightlanes = dict()
lightoutlanes = dict()
notlightlanes = dict()
notlightoutlanes = dict()
lights = []
notLights = []
lightlinkconflicts = dict()
nLanes = dict()
speeds = dict()
fftimes = dict() #Free flow times for each edge/lane (dict contains both) and from each light (min over outlanes)
links = dict()
lengths = dict()
turndata = []
normprobs = dict()
timedata = dict()
surtracdata = dict()
lanephases = dict()
mainlastswitchtimes = dict()
currentRoutes = dict()
routeStats = dict()
hmetadict = dict()
delay3adjdict = dict()
lightphases = dict()
laneDict = dict()
sumoPredClusters = [] #This'll update when we call doSurtrac from sumo things
rerouterLanes = dict()
rerouterEdges = dict()
nonExitEdgeDetections = dict() #nonExitEdgeDetections[road] = array of sections of road, each with a detector in each lane at the start. Store (madeupname, lane, time) for all cars. If new vehicle in first road segment, look up where it came from and steal the oldest car from that lane segment, else the oldest car we can find. If non-first road segment, steal oldest from previous segment disregarding lane
nonExitLaneDetectors = dict() #nonExitLaneDetections[lane] = [(detectorname1, detectorpos1), ...], should be same length as nonExitEdgeDetections[road]
wasFull = dict() #We're now using this to store stats for lane transition probabilities (specifically, the times we saw vehicles)
wasFullWindow = 300
vehiclesOnNetwork = []
dontReroute = []
surtracDict = dict()
adopterinfo = dict()

#Predict traffic entering network
arrivals = dict()
maxarrivalwindow = 300 #Use negative number to not predict new incoming cars during routing
arrivals2 = dict()
maxarrivalwindow2 = 60 #Same as maxarrivalwindow if you just want the baseline arrival rate. Go smaller (~0.5-1 light cycles) if you want to predict initial off-network queues based on recent arrival rates being low
newcarcounter = 0

totalSurtracTime = 0
totalSurtracClusters = 0
totalSurtracRuns = 0

totalLoadTime = 0
totalLoadCars = 0
totalLoadRuns = 0

#Threading routing
toReroute = []
reroutedata = dict()
threads = dict()
killSurtracThread = True

nRoutingCalls = 0
nSuccessfulRoutingCalls = 0
routingTime = 0
routeVerifyData = []

AStarCutoff = inf

oldids = dict()
timedata = dict()

savename = "MAIN_" + str(time.time())
netfile = "UNKNOWN_FILENAME_OOPS"

#Neural net things
import torch
from torch import nn
LOG_FILE = 'imitate.log' # Logs will be appended every time the code is run.
MODEL_FILES = dict()
def log(*args, **kwargs):
    if LOG_FILE:
        with open(LOG_FILE, 'a+') as f:
            print(*args, file=f, **kwargs)
    print(*args, **kwargs)

agents = dict()
optimizers = dict()
trainingdata = dict() #Going to be a list of 2-elt tuples (in, out) = ([in1, in2, ...], out)

actions = [0, 1]
learning_rate = 0.0005
avgloss = 0
nlosses = 0
nLossesBeforeReset = 1000

ndumbtrac = 0
ndumbtracerr = 0

teleportdata = []

#Non-NN stuff
def mergePredictions(clusters, predClusters):
    mergedClusters = pickle.loads(pickle.dumps(clusters)) #Because pass-by-value stuff
    for lane in clusters:
        if lane in predClusters:
            mergedClusters[lane] += predClusters[lane] #Concatenate known clusters with predicted clusters
            consolidateClusters(mergedClusters[lane])
    return mergedClusters

def consolidateClusters(clusters):
    i = 0
    while i < len(clusters):
        j = i+1
        while j < len(clusters):
            #Check if clusters i and j should merge
            if clusters[i]["arrival"] <= clusters[j]["arrival"] and clusters[j]["arrival"] <= clusters[i]["departure"] + clusterthresh:
                #Merge j into i
                clusters[i]["departure"] = max(clusters[i]["departure"], clusters[j]["departure"])
                clusters[i]["weight"] += clusters[j]["weight"]
                clusters[i]["cars"] += clusters[j]["cars"] #Concatenate (I hope)
                clusters.pop(j)
            else:
                if clusters[j]["arrival"] <= clusters[i]["arrival"] and clusters[i]["arrival"] <= clusters[j]["departure"] + clusterthresh:
                    #Merge i into j
                    clusters[j]["departure"] = max(clusters[i]["departure"], clusters[j]["departure"])
                    clusters[j]["weight"] += clusters[i]["weight"]
                    clusters[j]["cars"] += clusters[i]["cars"] #Concatenate (I hope)
                    clusters[i] = clusters[j]
                    clusters.pop(j)
            j+=1
        i+=1

def dumbtrac(simtime, light, clusters, lightphases, lastswitchtimes):
    if FTP:
        return dumbtracFTP(simtime, light, clusters, lightphases, lastswitchtimes)
    else:
        return dumbtracActuated(simtime, light, clusters, lightphases, lastswitchtimes)

def dumbtracFTP(simtime, light, clusters, lightphases, lastswitchtimes):

    phase = lightphases[light]
    lastSwitch = lastswitchtimes[light]

    #assert(traci.trafficlight.getNextSwitch(light) - simtime == lightphasedata[light][phase].duration - (simtime-lastSwitch)) #Falso occasionally. Probably has to do with phase switches, but haven't checked in detail.
    return lightphasedata[light][phase].duration - (simtime-lastSwitch)

    # #For FTP
    # if "Y" in lightphasedata[light][phase].state or "y" in lightphasedata[light][phase].state:
    #     return surtracdata[light][phase]["minDur"] - (simtime-lastSwitch)
    # else:
    #     return 30 - (simtime-lastSwitch)

    

def dumbtracActuated(simtime, light, clusters, lightphases, lastswitchtimes):
    phase = lightphases[light]
    lastSwitch = lastswitchtimes[light]
    
    maxfreq = max(routingSurtracFreq, mainSurtracFreq, timestep, 1)

    if surtracdata[light][phase]["maxDur"]- maxfreq <= surtracdata[light][phase]["minDur"]:
        #Edge case where duration range is smaller than period between updates, in which case overruns are unavoidable
        if simtime - lastSwitch < surtracdata[light][phase]["minDur"]:
            phaselenprop = -1
        else:
            phaselenprop = 2
    else:
        phaselenprop = (simtime - lastSwitch - surtracdata[light][phase]["minDur"])/(surtracdata[light][phase]["maxDur"]- maxfreq - surtracdata[light][phase]["minDur"])

    #Satisfy phase length requirements
    if phaselenprop < 0:
        return 10 #If haven't reached min length, continue
    if phaselenprop >= 1:
        return -10 #If >= max length, switch

    out = -10
    for lane in surtracdata[light][phase]["lanes"]:
        if len(clusters[lane]) > 0 and clusters[lane][0]["arrival"] <= simtime+mingap:
            out = 10#max(out, 2.5*(clusters[lane][0]["weight"]+1)) #If anyone's waiting, continue
    return out #If no one's waiting, switch

def convertToNNInput(simtime, light, clusters, lightphases, lastswitchtimes):
    maxnlanes = 3 #Going to assume we have at most 3 lanes per road, and that the biggest number lane is left-turn only
    maxnroads = 4 #And assume 4-way intersections for now
    nqueued = np.zeros(maxnroads*maxnlanes)
    ntotal = np.zeros(maxnroads*maxnlanes)
    phase = lightphases[light]
    lastSwitch = lastswitchtimes[light]

    maxfreq = max(routingSurtracFreq, mainSurtracFreq, timestep, 1)

    if False: #surtracdata[light][phase]["maxDur"]- maxfreq <= surtracdata[light][phase]["minDur"]:
        #Edge case where duration range is smaller than period between updates, in which case overruns are unavoidable
        print("Warning: minDur and maxDur for light " + light + " are too close together, we might have rounding errors. Also training data is going to be hacky here.")
        if simtime - lastSwitch < surtracdata[light][phase]["minDur"]:
            phaselenprop = -1
        else:
            phaselenprop = 2
    else:
        phaselenprop = (simtime - lastSwitch - surtracdata[light][phase]["minDur"])/(surtracdata[light][phase]["maxDur"]- maxfreq - surtracdata[light][phase]["minDur"])

    #phaselenprop = (simtime - lastSwitch - surtracdata[light][phase]["minDur"])/surtracdata[light][phase]["maxDur"]
    #phaselenprop is negative if we're less than minDur, and greater than 1 if we're greater than maxDur

    prevRoad = None
    roadind = -1
    laneind = -1

    for lane in lightlanes[light]:
        temp = lane.split("_")
        road = temp[0] #Could have problems if road name has underscores, but ignoring for now...
        if nLanes[road] > maxnlanes:
            print("Warning: " + str(road) + " exceeds maxnlanes in convertToNNInput, ignoring some lanes")
        lanenum = int(temp[-1])
        if road != prevRoad or roadind < 0:
            roadind += 1
            laneind = -1
            prevRoad = road
        laneind += 1
        #Last lane on road assumed to be left-turn only and being inserted in last slot
        if laneind + 1 == nLanes[road] or laneind + 1 >= maxnlanes:
            laneind = maxnlanes - 1

        if len(clusters[lane]) > 0 and clusters[lane][0]["arrival"] <= simtime + mingap:
            nqueued[roadind*maxnlanes+laneind] = clusters[lane][0]["weight"]
        ntotaltemp = 0
        for clusterind in range(len(clusters[lane])):
            ntotaltemp += clusters[lane][clusterind]["weight"]
        ntotal[roadind*maxnlanes+laneind] = ntotaltemp

    #return torch.Tensor(np.array([np.concatenate(([phase], [phaselenprop]))]))
    return torch.Tensor(np.array([np.concatenate((nqueued, ntotal, [phase], [phaselenprop]))]))

def convertToNNInputSurtrac(simtime, light, clusters, lightphases, lastswitchtimes):
    maxnlanes = 3 #Going to assume we have at most 3 lanes per road, and that the biggest number lane is left-turn only
    maxnroads = 4 #And assume 4-way intersections for now
    maxnclusters = 5 #And assume at most 10 clusters per lane
    ndatapercluster = 3 #Arrival, departure, weight

    clusterdata = np.zeros(maxnroads*maxnlanes*maxnclusters*ndatapercluster)

    nqueued = np.zeros(maxnroads*maxnlanes)
    ntotal = np.zeros(maxnroads*maxnlanes)
    phase = lightphases[light]
    lastSwitch = lastswitchtimes[light]

    maxfreq = max(routingSurtracFreq, mainSurtracFreq, timestep, 1)

    if surtracdata[light][phase]["maxDur"]- maxfreq <= surtracdata[light][phase]["minDur"]:
        #Edge case where duration range is smaller than period between updates, in which case overruns are unavoidable
        if simtime - lastSwitch < surtracdata[light][phase]["minDur"]:
            phaselenprop = -1
        else:
            phaselenprop = 2
    else:
        phaselenprop = (simtime - lastSwitch - surtracdata[light][phase]["minDur"])/(surtracdata[light][phase]["maxDur"]- maxfreq - surtracdata[light][phase]["minDur"])

    #phaselenprop = (simtime - lastSwitch - surtracdata[light][phase]["minDur"])/surtracdata[light][phase]["maxDur"]
    #phaselenprop is negative if we're less than minDur, and greater than 1 if we're greater than maxDur

    prevRoad = None
    roadind = -1
    laneind = -1

    for lane in lightlanes[light]:
        temp = lane.split("_")
        road = temp[0] #Could have problems if road name has underscores, but ignoring for now...
        if nLanes[road] > maxnlanes:
            print("Warning: " + str(road) + " exceeds maxnlanes in convertToNNInput, ignoring some lanes")
        lanenum = int(temp[-1])
        if road != prevRoad or roadind < 0:
            roadind += 1
            laneind = -1
            prevRoad = road
        laneind += 1

        #Not sharing weights so I'll skip this
        #Last lane on road assumed to be left-turn only and being inserted in last slot
        # if laneind + 1 == nLanes[road] or laneind + 1 >= maxnlanes:
        #     laneind = maxnlanes - 1

        for clusterind in range(len(clusters[lane])):
            if clusterind > maxnclusters:
                print("Warning: Too many clusters on " + str(lane) + ", ignoring the last ones")
                break
            clusterdata[((roadind*maxnlanes+laneind)*maxnclusters+clusterind)*ndatapercluster : ((roadind*maxnlanes+laneind)*maxnclusters+clusterind+1)*ndatapercluster] = [clusters[lane][clusterind]["arrival"]-simtime, clusters[lane][clusterind]["departure"]-simtime, clusters[lane][clusterind]["weight"]]

    #return torch.Tensor(np.array([np.concatenate(([phase], [phaselenprop]))]))
    return torch.Tensor(np.array([np.concatenate((clusterdata, [phase], [phaselenprop]))]))

#@profile
def doSurtracThread(network, simtime, light, clusters, lightphases, lastswitchtimes, inRoutingSim, predictionCutoff, toSwitch, catpreds, bestschedules):
    global totalSurtracRuns
    global totalSurtracClusters
    global totalSurtracTime

    if inRoutingSim:
        freq = max(routingSurtracFreq, timestep)
        ttimestep = timestep
    else:
        freq = max(mainSurtracFreq, 1)
        ttimestep = 1
        
    i = lightphases[light]
    if not learnYellow and ("Y" in lightphasedata[light][i].state or "y" in lightphasedata[light][i].state):
        #Force yellow phases to be min duration regardless of what anything else says, and don't store it as training data
        if simtime - lastswitchtimes[light] >= surtracdata[light][i]["minDur"]:
            dur = 0
        else:
            dur = (surtracdata[light][i]["minDur"] - (simtime - lastswitchtimes[light]))//ttimestep*ttimestep
        #Replace first element with remaining duration, rather than destroying the entire schedule, in case of Surtrac or similar
        if light in bestschedules:
            temp = pickle.loads(pickle.dumps(bestschedules[light][7]))
        else:
            temp = [0]
        temp[0] = dur
        bestschedules[light] = [None, None, None, None, None, None, None, temp]
        return

    if not learnMinMaxDurations:
        #Force light to satisfy min/max duration requirements and don't store as training data
        if simtime - lastswitchtimes[light] < surtracdata[light][i]["minDur"]:
            dur = 1e6
            if light in bestschedules:
                temp = pickle.loads(pickle.dumps(bestschedules[light][7]))
            else:
                temp = [0]
            temp[0] = dur
            bestschedules[light] = [None, None, None, None, None, None, None, temp]
            return
        if simtime - lastswitchtimes[light] + freq > surtracdata[light][i]["maxDur"]:
            #TODO this is slightly sloppy if freq > ttimestep - if we're trying to change just before maxDur this'll assume we tried to change at it instead
            dur = (surtracdata[light][i]["maxDur"] - (simtime - lastswitchtimes[light]))//ttimestep*ttimestep
            if light in bestschedules:
                temp = pickle.loads(pickle.dumps(bestschedules[light][7]))
            else:
                temp = [0]
            temp[0] = dur
            bestschedules[light] = [None, None, None, None, None, None, None, temp]
            return

    if (testNN and (inRoutingSim or not noNNinMain)) or testDumbtrac: #If using NN and/or dumbtrac
        if (testNN and (inRoutingSim or not noNNinMain)): #If using NN
            if testDumbtrac: #And also dumbtrac
                nnin = convertToNNInputSurtrac(simtime, light, clusters, lightphases, lastswitchtimes)

                # nnin = convertToNNInput(simtime, light, clusters, lightphases, lastswitchtimes) #Obsolete - Surtrac architecture works for dumbtrac too!
            else: #NN but not dumbtrac
                nnin = convertToNNInputSurtrac(simtime, light, clusters, lightphases, lastswitchtimes)
            
            surtracStartTime = time.time()
            totalSurtracRuns += 1
        
            outputNN = agents[light](nnin) # Output from NN

            if debugMode:
                totalSurtracTime += time.time() - surtracStartTime

            if outputNN <= 0:
                actionNN = 1 #Switch
            else:
                actionNN = 0 #Stick

        if testDumbtrac and not (testNN and (inRoutingSim or not noNNinMain)): #Dumbtrac but not NN
            outputDumbtrac = dumbtrac(simtime, light, clusters, lightphases, lastswitchtimes)
            if outputDumbtrac <= 0: #Stick for <= 0 seconds
                actionDumbtrac = 1 #Switch
            else:
                actionDumbtrac = 0 #Stick
            actionNN = actionDumbtrac

        if actionNN == 0:
            dur = 1e6 #Something really big so we know the light won't change
        else:
            dur = 0
        testnnschedule = [None, None, None, None, None, None, None, [dur]] #Only thing the output needs is a schedule; returns either [0] for switch immediately or [1] for continue for at least another timestep
        assert(len(testnnschedule[7]) > 0)
        #return #Don't return early, might still need to append training data

    if (not (testNN and (inRoutingSim or not noNNinMain)) and not testDumbtrac) or (appendTrainingData and not testDumbtrac): #(No NN or append training data) and no dumbtrac - get the actual Surtrac result
        #print("Running surtrac, double-check that this is intended.")
        #We're actually running Surtrac

        surtracStartTime = time.time()
        totalSurtracRuns += 1

        sult = 3 #Startup loss time
        greedyDP = True

        #Figure out what an initial and complete schedule look like
        nPhases = len(surtracdata[light]) #Number of phases
        bestschedules[light] = [[]] #In case we terminate early or something??

        emptyStatus = dict()
        fullStatus = dict()
        nClusters = 0
        maxnClusters = 0

        #Does this vectorize somehow?
        for lane in lightlanes[light]:
            emptyStatus[lane] = 0
            fullStatus[lane] = len(clusters[lane])
            nClusters += fullStatus[lane]
            if maxnClusters < fullStatus[lane]:
                maxnClusters = fullStatus[lane]
        if debugMode:
            totalSurtracClusters += nClusters
        #If there's nothing to do, send back something we recognize as "no schedule"
        if nClusters == 0:
            bestschedules[light] = [[]]
            return

        #Stuff in the partial schedule tuple
        #0: list of indices of the clusters we've scheduled
        #1: schedule status (how many clusters from each lane we've scheduled)
        #2: current light phase
        #3: time when each direction will have finished its last scheduled cluster
        #4: time when all directions are finished with scheduled clusters ("total makespan" + starting time...)
        #5: total delay
        #6: last switch time
        #7: planned total durations of all phases
        #8: predicted outflows (as clusters - arrival, departure, list of cars, weights, etc.) Blank for now since I'll generate it at the end.
        #9: pre-predict data (cluster start times and compression factors) which I'll use to figure out predicted outflows once we've determined the best schedule

        emptyPreds = dict()
        for lane in lightoutlanes[light]:
            emptyPreds[lane] = []

        # emptyPrePreds = dict()
        # for lane in lightlanes[light]:
        #     emptyPrePreds[lane] = []
        lenlightlaneslight = len(lightlanes[light])
        assert(lenlightlaneslight > 0)
        emptyPrePreds = np.zeros((lenlightlaneslight, maxnClusters, 2))

        phase = lightphases[light]
        lastSwitch = lastswitchtimes[light]
        schedules = [([], emptyStatus, phase, [simtime]*len(surtracdata[light][phase]["lanes"]), simtime, 0, lastSwitch, [simtime - lastSwitch], [], emptyPrePreds)]

        for _ in range(nClusters): #Keep adding a cluster until #clusters added = #clusters to be added
            scheduleHashDict = dict()
            for schedule in schedules:
                for laneindex in range(lenlightlaneslight):
                    lane = lightlanes[light][laneindex]
    
                    if schedule[1][lane] == fullStatus[lane]:
                        continue
                    #Consider adding next cluster from surtracdata[light][i]["lanes"][j] to schedule
                    newScheduleStatus = copy(schedule[1]) #Shallow copy okay? Dict points to int, which is stored by value
                    newScheduleStatus[lane] += 1
                    assert(newScheduleStatus[lane] <= maxnClusters)
                    phase = schedule[2]

                    #Now loop over all phases where we can clear this cluster
                    try:
                        assert(len(lanephases[lane]) > 0)
                    except:
                        print(lane)
                        print("ERROR: Can't clear this lane ever?")
                        
                    for i in lanephases[lane]:
                        if not learnYellow and ("Y" in lightphasedata[light][i].state or "y" in lightphasedata[light][i].state):
                            continue
                        directionalMakespans = copy(schedule[3])

                        nLanes = len(surtracdata[light][i]["lanes"])
                        j = surtracdata[light][i]["lanes"].index(lane)

                        newDurations = copy(schedule[7]) #Shallow copy should be fine

                        clusterind = newScheduleStatus[lane]-1 #We're scheduling the Xth cluster; it has index X-1
                        ist = clusters[lane][clusterind]["arrival"] #Intended start time = cluster arrival time
                        dur = clusters[lane][clusterind]["departure"] - ist + mingap #+mingap because next cluster can't start until mingap after current cluster finishes
                        
                        mindur = max((clusters[lane][clusterind]["weight"] )*mingap, 0) #No -1 because fencepost problem; next cluster still needs 2.5s of gap afterwards
                        delay = schedule[5]

                        if dur < mindur:
                            #print("Warning, dur < mindur???")
                            dur = mindur

                        if phase == i:
                            pst = schedule[3][j]
                            newLastSwitch = schedule[6] #Last switch time doesn't change
                            ast = max(ist, pst)
                            newdur = max(dur - (ast-ist), mindur) #Try to compress cluster as it runs into an existing queue
                            currentDuration = max(ist, ast)+newdur-schedule[6] #Total duration of current light phase if we send this cluster without changing phase

                        if not phase == i or currentDuration > surtracdata[light][i]["maxDur"]: #We'll have to switch the light, possibly mid-cluster

                            if not phase == i:
                                #Have to switch light phases.
                                newFirstSwitch = max(schedule[6] + surtracdata[light][phase]["minDur"], schedule[4]-mingap, simtime) #Because I'm adding mingap after all clusters, but here the next cluster gets delayed. Except for first phase, which usually wants to switch 2.5s in the past if there's no clusters
                            else:
                                #This cluster is too long to fit entirely in the current phase
                                newFirstSwitch = schedule[6] + surtracdata[light][phase]["maxDur"] #Set current phase to max duration
                                #Figure out how long the remaining part of the cluster is
                                tSent = surtracdata[light][i]["maxDur"] - (max(ist, ast)-schedule[6]) #Time we'll have run this cluster for before the light switches
                                if tSent < 0: #Cluster might arrive after the light would have switched due to max duration (ist is big), which would have made tSent go negative
                                    tSent = 0
                                    try:
                                        assert(mindur >= 0)
                                        assert(dur >= 0)
                                    except AssertionError as e:
                                        print(mindur)
                                        print(dur)
                                        raise(e)

                                if mindur > 0 and dur > 0: #Having issues with negative weights, possibly related to cars contributing less than 1 to weight having left the edge
                                    #We've committed to sending this cluster in current phase, but current phase ends before cluster
                                    #So we're sending what we can, cycling through everything else, then sending the rest
                                    #Compute the delay on the stuff we sent through, then treat the rest as a new cluster and compute stuff then
                                    delay += tSent/dur*clusters[surtracdata[light][i]["lanes"][j]][clusterind]["weight"]*((ast-ist)-1/2*(dur-newdur) ) #Weight of stuff sent through, times amount the start time got delayed minus half the squishibility
                                    mindur *= 1-tSent/dur #Assuming uniform density, we've sent tSent/dur fraction of vehicles through, so 1-tSent/dur remain to be handled
                                else:
                                    print("Negative weight, what just happened?")
                                dur -= tSent

                            newLastSwitch = newFirstSwitch + surtracdata[light][(phase+1)%nPhases]["timeTo"][i] #Switch right after previous cluster finishes (why not when next cluster arrives minus sult? Maybe try both?)                        
                            pst = newLastSwitch + sult #Total makespan + switching time + startup loss time
                            #Technically this sult implementation isn't quite right, as a cluster might reach the light as the light turns green and not have to stop and restart
                            directionalMakespans = [pst]*nLanes #Other directions can't schedule a cluster before the light switches

                            newDurations[-1] = newFirstSwitch - schedule[6] #Previous phase lasted from when it started to when it switched
                            tempphase = (phase+1)%nPhases
                            while tempphase != i:
                                newDurations.append(surtracdata[light][i]["minDur"])
                                tempphase = (tempphase+1)%nPhases
                            newDurations.append(0) #Duration of new phase i. To be updated on future loops once we figure out when the cluster finishes
                            assert(newDurations != schedule[7]) #Confirm that shallow copy from before is fine

                        ast = max(ist, pst)
                        newdur = max(dur - (ast-ist), mindur) #Compress cluster once cars start stopping

                        newPrePredict = copy(schedule[9])#pickle.loads(pickle.dumps(schedule[9]))
                        # print(np.shape(newPrePredict))
                        # print(lightoutlanes[light])
                        # print(lane)
                        # print(laneindex)
                        # print(newScheduleStatus[lane]-1)
                        newPrePredict[laneindex][newScheduleStatus[lane]-1][0] = ast #-1 because zero-indexing; first cluster has newScheduleStatus[lane] = 1, but is stored at index 0
                        if dur <= mindur:
                            newPrePredict[laneindex][newScheduleStatus[lane]-1][1] = 0 #Squish factor = 0 (no squishing)
                        else:
                            newPrePredict[laneindex][newScheduleStatus[lane]-1][1] = (dur-newdur)/(dur-mindur) #Squish factor equals this thing
                            #If newdur = mindur, compression factor = 1, all gaps are 2.5 (=mindur)
                            #If newdur = dur, compression factor = 0, all gaps are original values
                            #Otherwise smoothly interpolate
                        
                        #Tell other clusters to also start no sooner than max(new ast, old directionalMakespan value) to preserve order
                        #That max is important, though; blind overwriting is wrong, as you could send a long cluster, then a short one, then change the light before the long one finishes
                        assert(len(directionalMakespans) == len(surtracdata[light][i]["lanes"]))
                        directionalMakespans[j] = ast+newdur+mingap

                        directionalMakespans = np.maximum(directionalMakespans, ast).tolist()
                        
                        delay += clusters[surtracdata[light][i]["lanes"][j]][clusterind]["weight"]*((ast-ist)-1/2*(dur-newdur) ) #Delay += #cars * (actual-desired). 1/2(dur-newdur) compensates for the cluster packing together as it waits (I assume uniform compression)
                        try:
                            assert(delay >= schedule[5] - 1e-10) #Make sure delay doesn't go negative somehow
                        except AssertionError as e:
                            print("Negative delay, printing lots of debug stuff")
                            #print(clusters)
                            print(light)
                            print(lane)
                            print(clusters[surtracdata[light][i]["lanes"][j]][clusterind])
                            print(ast)
                            print(ist)
                            print(dur)
                            print(newdur)
                            print((ast-ist)-1/2*(dur-newdur))
                            raise(e)

                        newMakespan = max(directionalMakespans)
                        currentDuration = newMakespan - newLastSwitch

                        newDurations[-1] = currentDuration 
                        #Stuff in the partial schedule tuple
                        #0: list of indices of the clusters we've scheduled
                        #1: schedule status (how many clusters from each lane we've scheduled)
                        #2: current light phase
                        #3: time when each direction will have finished its last scheduled cluster
                        #4: time when all directions are finished with scheduled clusters ("total makespan" + starting time...)
                        #5: total delay
                        #6: last switch time
                        #7: planned total durations of all phases
                        #8: predicted outflows (as clusters - arrival, departure, list of cars, weights, etc.) Blank for now since I'll generate it at the end.
                        #9: pre-predict data (cluster start times and compression factors) which I'll use to figure out predicted outflows once we've determined the best schedule

                        newschedule = (schedule[0]+[(i, j)], newScheduleStatus, i, directionalMakespans, newMakespan, delay, newLastSwitch, newDurations, [], newPrePredict)
                        
                        #DP on partial schedules
                        key = (tuple(newschedule[1].values()), newschedule[2]) #Key needs to be something immutable (like a tuple, not a list)

                        if not key in scheduleHashDict:
                            scheduleHashDict[key] = [newschedule]
                        else:
                            keep = True
                            testscheduleind = 0
                            while testscheduleind < len(scheduleHashDict[key]):
                                testschedule = scheduleHashDict[key][testscheduleind]

                                #These asserts should follow from how I set up scheduleHashDict
                                if debugMode:
                                    assert(newschedule[1] == testschedule[1])
                                    assert(newschedule[2] == testschedule[2])
                                
                                #NOTE: If we're going to go for truly optimal, we also need to check all makespans, plus the current phase duration
                                #OTOH, if people seem to think fast greedy approximations are good enough, I'm fine with that
                                if newschedule[5] >= testschedule[5] and (greedyDP or newschedule[4] >= testschedule[4]):
                                    #New schedule was dominated; remove it and don't continue comparing (old schedule beats anything new one would)
                                    keep = False
                                    break
                                if newschedule[5] <= testschedule[5] and (greedyDP or newschedule[4] <= testschedule[4]):
                                    #Old schedule was dominated; remove it
                                    scheduleHashDict[key].pop(testscheduleind)
                                    continue
                                #No dominance, keep going
                                testscheduleind += 1

                            if keep:
                                scheduleHashDict[key].append(newschedule)
                        if debugMode:
                            assert(len(scheduleHashDict[key]) > 0)

            schedules = sum(list(scheduleHashDict.values()), []) #Each key has a list of non-dominated partial schedules. list() turns the dict_values object into a list of those lists; sum() concatenates to one big list of partial schedules. (Each partial schedule is stored as a tuple)

        mindelay = np.inf
        bestschedule = [[]]
        for schedule in schedules:
            if schedule[5] < mindelay:
                mindelay = schedule[5]
                bestschedule = schedule

        if not bestschedule == [[]]:
            #We have our best schedule, now need to generate predicted outflows
            if disableSurtracPred:
                newPredClusters = emptyPreds
            else:
                newPredClusters = pickle.loads(pickle.dumps(emptyPreds)) #Deep copy needed if I'm going to merge clusters

                nextSendTimes = [] #Priority queue
                clusterNums = dict()
                carNums = dict()
                #for lane in lightlanes[light]:
                for laneind in range(lenlightlaneslight):
                    lane = lightlanes[light][laneind]
                    clusterNums[lane] = 0
                    carNums[lane] = 0
                    if len(clusters[lane]) > clusterNums[lane]: #In case there's no clusters on a given lane
                        #heappush(nextSendTimes, (bestschedule[9][lane][clusterNums[lane]][0], lane)) #Pre-predict for appropriate lane for first cluster, get departure time, stuff into a priority queue
                        heappush(nextSendTimes, (bestschedule[9][laneind][clusterNums[lane]][0], laneind)) #Pre-predict for appropriate lane for first cluster, get departure time, stuff into a priority queue

                while len(nextSendTimes) > 0:
                    (nextSendTime, laneind) = heappop(nextSendTimes)
                    lane = lightlanes[light][laneind]
                    if nextSendTime + fftimes[light] > simtime + predictionCutoff:
                        #fftimes[light] is the smallest fftime of any output lane
                        #So if we're here, there's no way we'll ever want to predict this or any later car
                        break

                    cartuple = clusters[lane][clusterNums[lane]]["cars"][carNums[lane]]
                    if not cartuple[0] in isSmart or isSmart[cartuple[0]]: #It's possible we call this from QueueSim, at which point we split the vehicle being routed and wouldn't recognize the new names. Anything else should get assigned to isSmart or not on creation
                        #Split on "|" and "_" to deal with splitty cars correctly
                        route = currentRoutes[cartuple[0].split("|")[0].split("_")[0]] #.split to deal with the possibility of splitty cars in QueueSim
                        edge = lane.split("_")[0]
                        if not edge in route:
                            #Not sure if or why this happens - maybe the route is changing and predictions aren't updating?
                            #Can definitely happen for a splitty car inside QueueSim
                            #Regardless, don't predict this car forward and hope for the best?
                            if not "|" in cartuple[0] and not "_" in cartuple[0]:
                                #Smart car is on an edge we didn't expect. Most likely it changed route between the previous and current Surtrac calls. Get rid of it for now; can we be cleverer?
                                # print(cartuple[0])
                                # print(route)
                                # print(edge)
                                # print("Warning, smart car on an edge that's not in its route. This shouldn't happen? Assuming a mispredict and removing")
                                continue
                            #TODO: else should predict it goes everywhere?
                            continue
                        edgeind = route.index(edge)
                        if edgeind+1 == len(route):
                            #At end of route, don't care
                            continue
                        nextedge = route[edgeind+1]
                        
                        if not nextedge in normprobs[lane]:
                            #Means normprobs[lane] would be 0; nobody turned from this lane to this edge in the initial data
                            #Might be happening if the car needs to make a last-minute lane change to stay on its route?
                            #TODO: Find a lane where it can continue with the route and go from there? Ignoring for now
                            #NEXT TODO: Apparently still a thing even with splitting the initial VOI to multiple lanes???
                            continue

                        for nextlaneind in range(nLanes[nextedge]):
                            nextlane = nextedge+"_"+str(nextlaneind)
                            arr = nextSendTime + fftimes[nextlane]
                            if arr > simtime + predictionCutoff:
                                #Don't add to prediction; it's too far in the future. And it'll be too far into the future for all other lanes on this edge too, so just stop
                                break

                            if not nextlane in turndata[lane] or turndata[lane][nextlane] == 0:
                                #Car has zero chance of going here, skip
                                continue

                            if len(newPredClusters[nextlane]) == 0 or arr > newPredClusters[nextlane][-1]["departure"] + clusterthresh:
                                #Add a new cluster
                                newPredCluster = dict()
                                newPredCluster["endpos"] = 0
                                newPredCluster["time"] = ast
                                newPredCluster["arrival"] = arr
                                newPredCluster["departure"] = arr
                                newPredCluster["cars"] = []
                                newPredCluster["weight"] = 0
                                newPredClusters[nextlane].append(newPredCluster)

                            modcartuple = (cartuple[0], arr, cartuple[2]*predDiscount*turndata[lane][nextlane] / normprobs[lane][nextedge], cartuple[3])
                            newPredClusters[nextlane][-1]["cars"].append(modcartuple)
                            newPredClusters[nextlane][-1]["weight"] += modcartuple[2]
                            newPredClusters[nextlane][-1]["departure"] = arr
                    else:
                        for nextlane in turndata[lane]:
                            #Copy-paste previous logic for creating a new cluster
                            arr = nextSendTime + fftimes[nextlane]
                            if arr > simtime + predictionCutoff:
                                #Don't add to prediction; it's too far in the future. Other lanes may differ though
                                continue

                            if not nextlane in turndata[lane] or turndata[lane][nextlane] == 0:
                                #Car has zero chance of going here, skip
                                continue

                            if len(newPredClusters[nextlane]) == 0 or arr > newPredClusters[nextlane][-1]["departure"] + clusterthresh:
                                #Add a new cluster
                                newPredCluster = dict()
                                newPredCluster["endpos"] = 0
                                newPredCluster["time"] = ast
                                newPredCluster["arrival"] = arr
                                newPredCluster["departure"] = arr
                                newPredCluster["cars"] = []
                                newPredCluster["weight"] = 0
                                newPredClusters[nextlane].append(newPredCluster)

                            modcartuple = (cartuple[0], arr, cartuple[2]*turndata[lane][nextlane], cartuple[3])
                            newPredClusters[nextlane][-1]["cars"].append(modcartuple)
                            newPredClusters[nextlane][-1]["weight"] += modcartuple[2]
                            newPredClusters[nextlane][-1]["departure"] = arr
                    
                    #Added car to predictions, now set up the next car
                    carNums[lane] += 1
                    while len(clusters[lane]) > clusterNums[lane] and len(clusters[lane][clusterNums[lane]]["cars"]) == carNums[lane]: #Should fire at most once, but use while just in case of empty clusters...
                        clusterNums[lane] += 1
                        carNums[lane] = 0
                    if len(clusters[lane]) == clusterNums[lane]:
                        #Nothing left on this lane, we're done here
                        #nextSendTimes.pop(lane)
                        continue
                    if carNums[lane] == 0:
                        heappush(nextSendTimes, (bestschedule[9][laneind][clusterNums[lane]][0], laneind)) #Time next cluster is scheduled to be sent
                    else:
                        #Account for cluster compression
                        prevSendTime = nextSendTime #When we sent the car above
                        rawSendTimeDelay = clusters[lane][clusterNums[lane]]["cars"][carNums[lane]][1] - clusters[lane][clusterNums[lane]]["cars"][carNums[lane]-1][1] #Time between next car and this car in the original cluster
                        compFac = bestschedule[9][laneind][clusterNums[lane]][1] #Compression factor in case cluster is waiting at a red light
                        sendTimeDelay = compFac*mingap + (1-compFac)*rawSendTimeDelay #Update time delay using compression factor
                        newSendTime = prevSendTime + sendTimeDelay #Compute time we'd send next car
                        heappush(nextSendTimes, (newSendTime, laneind))

            #Predicting should be done now
            #bestschedule[8] = newPredClusters #I'd store this, but tuples are immutable and we don't actually use it anywhere...
        
            catpreds.update(newPredClusters)
            if len(bestschedule[7]) > 0:
                bestschedule[7][0] -= (simtime - lastswitchtimes[light])
            bestschedules[light] = bestschedule
        else:
            print(light)
            print("No schedules anywhere? That shouldn't happen...")


            
        #nnin = convertToNNInput(simtime, light, clusters, lightphases, lastswitchtimes) #If stuff breaks, make sure none of this gets changed as we go. (Tested when I first wrote this.)
        # actionSurtrac = 0
        # if bestschedule[7][0] <= simtime - lastswitchtimes[light]:
        #     actionSurtrac = 1
        #NN should take in nnin, and try to return action, and do backprop accordingly
        #target = torch.tensor([bestschedule[7][0] - (simtime - lastswitchtimes[light])]) # Target from expert

        if debugMode:
            totalSurtracTime += time.time() - surtracStartTime
    
    if appendTrainingData:
        if testDumbtrac:
            outputDumbtrac = dumbtrac(simtime, light, clusters, lightphases, lastswitchtimes)
            target = torch.tensor([[outputDumbtrac-0.25]]) # Target from expert
            #nnin = convertToNNInput(simtime, light, clusters, lightphases, lastswitchtimes) #Obsolete - use Surtrac architecture anyway
            nnin = convertToNNInputSurtrac(simtime, light, clusters, lightphases, lastswitchtimes)
        else:
            target = torch.tensor([[bestschedule[7][0]-0.25]]) # Target from expert
            nnin = convertToNNInputSurtrac(simtime, light, clusters, lightphases, lastswitchtimes)

        if (testNN and (inRoutingSim or not noNNinMain)): #If NN
            trainingdata[light].append((nnin, target, torch.tensor([[outputNN]])))
        else:
            trainingdata[light].append((nnin, target)) #Record the training data, but obviously not what the NN did since we aren't using an NN
        
    
    if (testNN and (inRoutingSim or not noNNinMain)) or testDumbtrac:
        bestschedules[light] = testnnschedule
        

#@profile
def doSurtrac(network, simtime, realclusters=None, lightphases=None, lastswitchtimes=None, predClusters=None, inRoutingSim=True, nonExitEdgeDetections4 = nonExitEdgeDetections): #deepcopy breaks main Surtrac somehow?!
    global clustersCache
    global totalLoadRuns
    global totalLoadTime

    global testNN
    global testNNrolls
    if debugNNslowness:
        testNN = random.random() < 0.01
        testNNrolls.append(testNN)

    toSwitch = []
    if disableSurtracPred or not multithreadRouting:
        catpreds = dict()
    else:
        catpreds = manager.dict()
    remainingDuration = dict()
    bestschedules = dict()

    surtracThreads = dict()

    if realclusters == None or lightphases == None:
        #if clustersCache == None: #This scares me, let's not cache for now. Probably not the slow part here anyway
        totalLoadRuns += 1
        loadStart = time.time()
        if inRoutingSim and not detectorRoutingSurtrac: #Only use the detector model for Surtrac if we're not routing
            clustersCache = loadClusters(network, simtime, nonExitEdgeDetections4)
        else:
            if detectorSurtrac:
                clustersCache = loadClustersDetectors(network, simtime, nonExitEdgeDetections4) #This at least grabs the same vehicles as standard loadClusters, including ungrabbing them once they hit an exit road. Positions are probably slightly inaccurate, though, since this uses a detector model
            else:
                clustersCache = loadClusters(network, simtime, nonExitEdgeDetections4)
            
        runTime = time.time() - loadStart
        totalLoadTime += runTime
        if debugNNslowness and runTime > 1.5*totalLoadTime/totalLoadRuns:
            print(inRoutingSim)
            print(testNNrolls[-5:])
            print(nVehicles[-5:])

        (temprealclusters, templightphases) = pickle.loads(pickle.dumps(clustersCache))
        if realclusters == None:
            realclusters = temprealclusters
        if lightphases == None:
            lightphases = templightphases

    #predCutoff
    if inRoutingSim:
        predictionCutoff = predCutoffRouting #Routing
    else:
        predictionCutoff = predCutoffMain #Main simulation
    

    if not predClusters == None:
        clusters = mergePredictions(realclusters, predClusters)
    else:
        clusters = pickle.loads(pickle.dumps(realclusters))

    for light in lights:
        if multithreadSurtrac:
            surtracThreads[light] = threading.Thread(target=doSurtracThread, args=(network, simtime, light, clusters, lightphases, lastswitchtimes, inRoutingSim, predictionCutoff, toSwitch, catpreds, bestschedules))
            surtracThreads[light].start()
        else:
            doSurtracThread(network, simtime, light, clusters, lightphases, lastswitchtimes, inRoutingSim, predictionCutoff, toSwitch, catpreds, bestschedules)

    for light in lights:
        if multithreadSurtrac:
            surtracThreads[light].join()

        #bestschedules gets re-created each call to doSurtrac, so it's not some weird carryover thing
        #print(light)
        #print(bestschedules[light])
        bestschedule = bestschedules[light]
        if not bestschedule[0] == []: #Check for the case of Surtrac seeing no vehicles (which would default to default fixed timing plans)
            spentDuration = simtime - lastswitchtimes[light]
            remainingDuration[light] = pickle.loads(pickle.dumps(bestschedule[7]))

            if len(remainingDuration[light]) == 0:
                print('pretest - empty remainingDuration')
            if len(remainingDuration[light]) > 0:
                # if remainingDuration[light][0] <= 0:
                #     remainingDuration[light][0] = 0.01
                if remainingDuration[light][0] >= 0 and (not inRoutingSim or routingSimUsesSUMO):
                    #Update duration

                    if not(testDumbtrac and FTP):
                        traci.trafficlight.setPhaseDuration(light, remainingDuration[light][0]) #setPhaseDuration sets the remaining duration in the phase
                
                if remainingDuration[light][0] <= 0: #Light needs to change
                    pass
                    #Light needs to change
                    toSwitch.append(light)

                    curphase = lightphases[light]
                    nPhases = len(surtracdata[light]) #Number of phases

                    #If Surtrac tells a light to change, the phase duration should be within the allowed bounds
                    #Surtrac in routing (which uses larger timesteps) might exceed maxDur, but by less than the timestep
                    #TODO: Actually, might exceed but by no more than routing's surtracFreq - pipe surtracFreq into this function eventually?
                    # if not (simtime - lastswitchtimes[light] >= surtracdata[light][curphase]["minDur"] and simtime - lastswitchtimes[light] <= surtracdata[light][curphase]["maxDur"]+timestep):
                    #     print("Duration violation on light " + light + "; actual duration " + str(simtime - lastswitchtimes[light]))

                    if simtime - lastswitchtimes[light] < surtracdata[light][curphase]["minDur"]:
                        print("Duration violation on light " + light + "; actual duration " + str(simtime - lastswitchtimes[light]) + " but min duration " + str(surtracdata[light][curphase]["minDur"]))
                    if simtime - lastswitchtimes[light] > surtracdata[light][curphase]["maxDur"]+timestep:
                        print("Duration violation on light " + light + "; actual duration " + str(simtime - lastswitchtimes[light]) + " but max duration " + str(surtracdata[light][curphase]["maxDur"]))

                    lightphases[light] = (curphase+1)%nPhases #This would change the light if we're in routing sim
                    lastswitchtimes[light] = simtime

                    remainingDuration[light].pop(0)

                    if len(remainingDuration[light]) == 0:
                        remainingDuration[light] = [lightphasedata[light][(lightphases[light]+1)%len(lightphasedata[light])].duration]

                    if not inRoutingSim or routingSimUsesSUMO: #Actually change the light
                        #print("Surtrac switching light " + light)
                        traci.trafficlight.setPhase(light, (curphase+1)%nPhases) #Increment phase, duration defaults to default
                        if len(remainingDuration[light]) > 0:
                            #And set the new duration if possible
                            traci.trafficlight.setPhaseDuration(light, remainingDuration[light][0]) #Update duration if we know it
                            #pass
            else:
                print("Warning: Surtrac's giving back an empty schedule!")


    #Predict-ahead for everything else; assume no delays.
    if not disableSurtracPred:
        for light in notLights:
            for lane in notlightlanes[light]:
                if not lane in turndata:
                    continue
                edge = lane.split("_")[0]

                #This might not merge cars coming from different lanes in chronological order. Is this a problem?
                
                for clusterind in range(len(clusters[lane])): #TODO: This is going to be really bad at chronological ordering. Should probably just recompute clusters at the end
                    predLanes = []
                    for outlane in turndata[lane]: #turndata has psuedocounts so everything should be fine here?
                    
                        ist = clusters[lane][clusterind]["arrival"] #Intended start time = cluster arrival time
                        dur = clusters[lane][clusterind]["departure"] - ist
                        
                        arr = ist + fftimes[outlane]
                        if arr > simtime + predictionCutoff:
                            #Cluster is farther in the future than we want to predict; skip it
                            continue
                        newPredCluster = dict()
                        newPredCluster["endpos"] = 0
                        newPredCluster["time"] = ist
                        newPredCluster["arrival"] = arr
                        newPredCluster["departure"] = newPredCluster["arrival"] + dur
                        newPredCluster["cars"] = []
                        newPredCluster["weight"] = 0
                        if outlane in catpreds:
                            catpreds[outlane].append(newPredCluster)
                        else:
                            catpreds[outlane] = [newPredCluster]
                        predLanes.append(outlane) #Track which lanes' clusters are within the prediction cutoff
                        assert(len(catpreds[outlane]) >= 1)
                    #Add cars to new clusters
                    for cartuple in clusters[lane][clusterind]["cars"]:
                        #cartuple[0] is name of car; cartuple[1] is departure time; cartuple[2] is debug info
                        #assert(cartuple[0] in isSmart)
                        if not cartuple[0] in isSmart or isSmart[cartuple[0]]: #It's possible we call this from QueueSim, at which point we split the vehicle being routed and wouldn't recognize the new names. Anything else should get assigned to isSmart or not on creation
                            route = currentRoutes[cartuple[0].split("|")[0].split("_")[0]] #.split to deal with the possibility of splitty cars in QueueSim
                            if not edge in route:
                                #Not sure if or why this happens - maybe the route is changing and predictions aren't updating?
                                #Can definitely happen for a splitty car inside QueueSim
                                #Regardless, don't predict this car forward and hope for the best?
                                if not "|" in cartuple[0]:
                                    pass
                                    #TODO: This still happens sometimes, not sure why
                                    #print("Warning, smart car on an edge that's not in its route. Assuming a mispredict and removing")
                                #TODO: Else should predict it goes everywhere? Does this happen??
                                continue
                            edgeind = route.index(edge)
                            if edgeind+1 == len(route):
                                #At end of route, don't care
                                continue
                            nextedge = route[edgeind+1]

                            if not nextedge in normprobs[lane]:
                                #Might be happening if the car needs to make a last-minute lane change to stay on its route?
                                #TODO: Find a lane where it can continue with the route and go from there? Ignoring for now
                                #NEXT TODO: We're splitting the initial vehicle onto all starting lanes; this shouldn't be happening at all anymore. Verify?
                                #This seems fine, but similar code in doSurtracThread is still triggering. This is the code for prediction through non-lights, though, which probably just doesn't come up often enough for me to notice a problem. Probably still bad.
                                ##print("Warning, no data, having Surtrac prediction ignore this car instead of making something up")
                                #print(lane)
                                #print("normprob == 0, something's probably wrong??")
                                continue

                            for nextlaneind in range(nLanes[nextedge]):
                                nextlane = nextedge+"_"+str(nextlaneind)
                                modcartuple = (cartuple[0], cartuple[1]+fftimes[nextlane], cartuple[2]*turndata[lane][nextlane] / normprobs[lane][nextedge], cartuple[3])
                                if nextlane in predLanes:
                                    #Make sure we're predicting this cluster
                                    catpreds[nextlane][-1]["cars"].append(modcartuple)
                                    catpreds[nextlane][-1]["weight"] += modcartuple[2]

                        else:
                            for nextlane in predLanes:
                                modcartuple = (cartuple[0], cartuple[1]+fftimes[nextlane], cartuple[2]*turndata[lane][nextlane], cartuple[3])
                                catpreds[nextlane][-1]["cars"].append(modcartuple)
                                catpreds[nextlane][-1]["weight"] += modcartuple[2]

                    for outlane in predLanes:
                        if catpreds[outlane][-1]["weight"] == 0:
                            #Remove predicted clusters that are empty
                            catpreds[outlane].pop(-1)
                            continue

                        if len(catpreds[outlane]) >=2 and catpreds[outlane][-1]["arrival"] - catpreds[outlane][-2]["departure"] < clusterthresh:
                            #Merge this cluster with the previous one
                            #Pos and time don't do anything here
                            #Arrival doesn't change - previous cluster arrived first
                            catpreds[outlane][-2]["departure"] = max(catpreds[outlane][-2]["departure"], catpreds[outlane][-1]["departure"])
                            catpreds[outlane][-2]["cars"] += catpreds[outlane][-1]["cars"] # += concatenates
                            catpreds[outlane][-2]["weight"] += catpreds[outlane][-1]["weight"]
                            catpreds[outlane].pop(-1)

        #Confirm that weight of cluster = sum of weights of cars
        for lane in catpreds:
            for preind in range(len(catpreds[lane])):
                weightsum = 0
                for ind in range(len(catpreds[lane][preind]["cars"])):
                    weightsum += catpreds[lane][preind]["cars"][ind][2]
                assert(abs(weightsum - catpreds[lane][preind]["weight"]) < 1e-10)

    return (toSwitch, catpreds, remainingDuration)

#For computing free-flow time, which is used for computing delay. Stolen from my old A* code, where it was a heuristic function.
def backwardDijkstra(network, goal):
    goalcost = lengths[goal+"_0"]/network.getEdge(goal).getSpeed()
    gvals = dict()
    gvals[goal] = goalcost
    pq = []
    heappush(pq, (goalcost, goal)) #Cost to traverse from start of goal to end of goal is goalcost, not 0

    prevgval = goalcost
    while len(pq) > 0: #When the queue is empty, we're done
        #print(pq)
        stateToExpand = heappop(pq)
        #Sanity check: Things we pop off the heap should keep getting bigger
        assert(prevgval <= stateToExpand[0])
        prevgval = stateToExpand[0]
        #fval = stateToExpand[0]
        edge = stateToExpand[1]
        gval = gvals[edge]

        #Get predecessor IDs
        succs = []
        for succ in list(network.getEdge(edge).getIncoming()):
            succs.append(succ.getID())
        
        for succ in succs:
            c = lengths[edge+"_0"]/network.getEdge(edge).getSpeed()

            h = 0 #Heuristic not needed here - search is fast
            if succ in gvals and gvals[succ] <= gval+c:
                #Already saw this state, don't requeue
                continue

            #If we found a better way to reach succ, the old way is still in the queue; delete it
            if succ in gvals and gvals[succ] > gval+c:
                pq.remove((gvals[succ], succ))
                heapify(pq)

            #Otherwise it's new or we're now doing better, so requeue it
            gvals[succ] = gval+c
            heappush(pq, (gval+c+h, succ))
    return gvals

#@profile
def run(network, rerouters, pSmart, verbose = True):
    global sumoPredClusters
    global currentRoutes
    global hmetadict
    global delay3adjdict
    global actualStartDict
    global edgeDict
    global laneDict
    global clustersCache
    global teleportdata
    global adopterinfo
    #netfile is the filepath to the network file, so we can call sumolib to get successors
    #rerouters is the list of induction loops on edges with multiple successor edges
    
    startDict = dict()
    endDict = dict()
    delayDict = dict()
    delay2adjdict = dict()
    delay3adjdict = dict()
    edgeDict = dict()
    laneDict = dict()
    leftDict = dict()
    carsOnNetwork = []
    remainingDuration = dict()

    tstart = time.time()
    simtime = 0

    while traci.simulation.getMinExpectedNumber() > 0 and (not appendTrainingData or simtime < 5000):
        simtime += 1
        traci.simulationStep() #Tell the simulator to simulate the next time step

        if debugMode:
            assert(simtime == traci.simulation.getTime())
        clustersCache = None #Clear stored clusters list

        dontReroute = []

        #Decide whether new vehicles use our routing
        for vehicle in traci.simulation.getDepartedIDList():
            isSmart[vehicle] = random.random() < pSmart
            if isSmart[vehicle]:
                traci.vehicle.setColor(vehicle, [0, 255, 0, 255])
            else:
                traci.vehicle.setColor(vehicle, [255, 0, 0, 255])
            timedata[vehicle] = [simtime, -1, -1, 'unknown', 'unknown']
            currentRoutes[vehicle] = traci.vehicle.getRoute(vehicle)
            routeStats[vehicle] = dict()
            routeStats[vehicle]["nCalls"] = 0
            routeStats[vehicle]["nCallsFirst"] = 0
            routeStats[vehicle]["nCallsAfterFirst"] = 0
            routeStats[vehicle]["nSwaps"] = 0
            routeStats[vehicle]["nSwapsFirst"] = 0
            routeStats[vehicle]["nSwapsAfterFirst"] = 0
            routeStats[vehicle]["swapped"] = False
            routeStats[vehicle]["nTimeouts"] = 0
            routeStats[vehicle]["nTeleports"] = 0
            routeStats[vehicle]["distance"] = 0

            goaledge = currentRoutes[vehicle][-1]
            if not goaledge in hmetadict:
                hmetadict[goaledge] = backwardDijkstra(network, goaledge)
            delayDict[vehicle] = -hmetadict[goaledge][currentRoutes[vehicle][0]] #I'll add the actual travel time once the vehicle arrives
            laneDict[vehicle] = traci.vehicle.getLaneID(vehicle)
            edgeDict[vehicle] = traci.vehicle.getRoadID(vehicle)

            startDict[vehicle] = simtime
            leftDict[vehicle] = 0

            lane = laneDict[vehicle]
            if not lane in arrivals:
                arrivals[lane] = []
                arrivals2[lane] = []
            arrivals[lane].append(simtime) #Don't care who arrived, just when they arrived - this is for estimating future inflows if we turn that on in routing
            arrivals2[lane].append(simtime)

            #Manually pretend to be a detector at the start of the input lanes, since newly added vehicles have issues triggering those
            temp = lane.split("_")
            edge = temp[0]
            lanenum = int(temp[-1])
            if edge in nonExitEdgeDetections: #If it's an exit lane we hopefully don't care
                if isSmart[vehicle]:
                    nonExitEdgeDetections[edge][0].append((vehicle, lane, simtime))
                else:
                    nonExitEdgeDetections[edge][0].append((lane+"."+str(simtime), lane, simtime))

        #Check predicted vs. actual travel times
        for vehicle in traci.simulation.getArrivedIDList():
            if isSmart[vehicle]:
                timedata[vehicle][1] = simtime
                # print("Actual minus expected: %f" % ((timedata[vehicle][1]-timedata[vehicle][0]) - timedata[vehicle][2]))
                # print("Actual : %f" % (timedata[vehicle][1]-timedata[vehicle][0]))
                # print("Expected : %f" % (timedata[vehicle][2]))
                # print("Percent error : %f" % ( ((timedata[vehicle][1]-timedata[vehicle][0]) - timedata[vehicle][2]) / (timedata[vehicle][1]-timedata[vehicle][0]) * 100))
                # print("Route from " + timedata[vehicle][3] + " to " + timedata[vehicle][4])
            endDict[vehicle] = simtime

            #In case the target was a non-exit road, delete from old road detections
            if edgeDict[vehicle] in nonExitEdgeDetections:
                oldEdgeStuff = nonExitEdgeDetections[edgeDict[vehicle]][0] #Since we're only storing stuff in index 0 anyway
                if len(oldEdgeStuff) > 0:
                    oldEdgeStuff.pop(0) #Pop oldest from old road, don't care from which lane. Might not actually be the adopter in question
                else:
                    print("Warning: Ran out of cars to remove on edge " + edgeDict[vehicle] + "!!!!!!!!!!!!!!!!!")

                #Make sure we don't have a duplicate of this adopter on the last edge. If we do, make it a random car instead
                if isSmart[vehicle]:
                    for vehicletupleind in range(len(nonExitEdgeDetections[edgeDict[vehicle]][0])):
                        vehicletuple = nonExitEdgeDetections[edgeDict[vehicle]][0][vehicletupleind]
                        if vehicletuple[0] == vehicle:
                            nonExitEdgeDetections[edgeDict[vehicle]][0][vehicletupleind] = (edgeDict[vehicle]+".0oldsmartcar."+str(simtime), vehicletuple[1], vehicletuple[2]) #This seems to alias as intended

            edgeDict.pop(vehicle)
            laneDict.pop(vehicle)
            dontReroute.append(vehicle) #Vehicle has left network and does not need to be rerouted

        vehiclesOnNetwork = traci.vehicle.getIDList()
        carsOnNetwork.append(len(vehiclesOnNetwork)) #Store number of cars on network (for plotting)

        #Count left turns
        for id in laneDict:
            newlane = traci.vehicle.getLaneID(id)
            if len(newlane) == 0 or newlane[0] == ":":
                dontReroute.append(id) #Vehicle is mid-intersection or off network, don't try to reroute them
            if newlane != laneDict[id] and len(newlane) > 0 and  newlane[0] != ":":
                newloc = traci.vehicle.getRoadID(id)

                #Pretend to be detectors at the start of each road (need to know where we came from so we can steal from the correct previous lane)
                if newloc != edgeDict[id]: #Moved to a new road
                    #Delete from old road detections
                    if edgeDict[id] in nonExitEdgeDetections:
                        oldEdgeStuff = nonExitEdgeDetections[edgeDict[id]][0] #Since we're only storing stuff in index 0 anyway
                        if len(oldEdgeStuff) > 0:
                            oldEdgeStuff.pop(0) #Pop oldest from old road, don't care from which lane. Might not actually be the adopter in question
                        else:
                            print("Warning: Ran out of cars to remove on edge " + edgeDict[id] + "!!!!!!!!!!!!!!!!!")

                        #Make sure we don't have a duplicate of this adopter on the last edge. If we do, make it a random car instead
                        if isSmart[id]:
                            for vehicletupleind in range(len(nonExitEdgeDetections[edgeDict[id]][0])):
                                vehicletuple = nonExitEdgeDetections[edgeDict[id]][0][vehicletupleind]
                                if vehicletuple[0] == id:
                                    nonExitEdgeDetections[edgeDict[id]][0][vehicletupleind] = (edgeDict[id]+".0oldsmartcar."+str(simtime), vehicletuple[1], vehicletuple[2]) #This seems to alias as intended

                    #Add to new road detections
                    if newloc in nonExitEdgeDetections:
                        assert(newlane.split("_")[0] == newloc)
                        if isSmart[id]:
                            nonExitEdgeDetections[newloc][0].append((id, newlane, simtime))
                        else:
                            nonExitEdgeDetections[newloc][0].append((newlane+".0maindetect."+str(simtime), newlane, simtime))

                c0 = network.getEdge(edgeDict[id]).getFromNode().getCoord()
                c1 = network.getEdge(edgeDict[id]).getToNode().getCoord()
                theta0 = math.atan2(c1[1]-c0[1], c1[0]-c0[0])

                c2 = network.getEdge(newloc).getToNode().getCoord()
                theta1 = math.atan2(c2[1]-c1[1], c2[0]-c1[0])

                if (theta1-theta0+math.pi)%(2*math.pi)-math.pi > 0:
                    leftDict[id] += 1
                laneDict[id] = newlane
                edgeDict[id] = newloc
                #assert(laneDict[id] == traci.vehicle.getLaneID(id))
                try:
                    assert(laneDict[id].split("_")[0] in currentRoutes[id])
                except Exception as e:
                    print("Vehicle got off route somehow?")
                    #print(traci.getLabel()) #Should be main
                    print(id)
                    print(laneDict[id])
                    print(traci.vehicle.getLaneID(id))
                    print(currentRoutes[id])
                    print(traci.vehicle.getRoute(id))
                    #assert(traci.getLabel() == "main")
                    #raise(e)
                    laneDict[id] = traci.vehicle.getLaneID(id)
                    currentRoutes[id] = traci.vehicle.getRoute(id)
                    assert(laneDict[id].split("_")[0] in currentRoutes[id])
                #Remove vehicle from predictions, since the next intersection should actually see it now
                if not disableSurtracPred:
                    for predlane in sumoPredClusters:
                        for predcluster in sumoPredClusters[predlane]:

                            predcarind = 0
                            minarr = inf
                            maxarr = -inf
                            while predcarind < len(predcluster["cars"]):
                                predcartuple = predcluster["cars"][predcarind]
                                if predcartuple[0] == id:
                                    predcluster["cars"].pop(predcarind)
                                    predcluster["weight"] -= predcartuple[2]
                                else:
                                    predcarind += 1
                                    if predcartuple[1] < minarr:
                                        minarr = predcartuple[1]
                                    if predcartuple[1] > maxarr:
                                        maxarr = predcartuple[1]
                            if len(predcluster["cars"]) == 0:
                                sumoPredClusters[predlane].remove(predcluster)
                            else:
                                pass
                                #predcluster["arrival"] = minarr #predcluster["cars"][0][1]
                                #predcluster["departure"] = maxarr #predcluster["cars"][-1][1]

                            weightsum = 0
                            for predcarind in range(len(predcluster["cars"])):
                                weightsum += predcluster["cars"][predcarind][2]
                            assert(abs(weightsum - predcluster["weight"]) < 1e-10)

                #Store data to compute delay after first intersection
                if not id in delay2adjdict:
                    delay2adjdict[id] = simtime

                #Compute distance travelled if on last edge of route (since we can't do this once we leave the network)
                if newlane.split("_")[0] == currentRoutes[id][-1]:
                    routeStats[id]["distance"] = traci.vehicle.getDistance(id) + lengths[newlane]


        #Update vehicle position estimates using detector model
        #Except now it's just updating lane transition probability data
        for lane in nonExitLaneDetectors:
            edge = lane.split("_")[0]
            for ind in range(1, len(nonExitLaneDetectors[lane])): #Skip segment 0, our fake detector code looking for road changes covers that
                detector = nonExitLaneDetectors[lane][ind][0]
                ids = traci.inductionloop.getLastStepVehicleIDs(detector) #All vehicles to be rerouted
                if len(wasFull[detector]) > 0 and wasFull[detector][0] < simtime - wasFullWindow: #Remove old data
                    wasFull[detector].pop(0)
                if len(ids) == 0:
                    pass
                    # #Consider deleting old data
                    # if len(wasFull[detector]) > 0 and wasFull[detector][-1] < simtime - 5: #If we've not seen a vehicle for 5 seconds, start deleting vehicles that should've arrived by now
                    #     olddataind = 0
                    #     while olddataind < len(nonExitEdgeDetections[edge][0]):
                    #         (testname, testlane, testtime) = nonExitEdgeDetections[edge][0][olddataind]
                    #         if testtime > simtime - fftimes[edge] + 2: #TODO using +2 as a buffer in case travel times are slightly off somehow, test this
                    #             break #We're up to non-old data, stop looking for stuff to delete
                    #         if testname in isSmart and isSmart[testname]:
                    #             #Vehicle is an adopter. Not sure what's happening, but we can track it, so presumably it's accurate?
                    #             olddataind += 1
                    #             continue
                    #         if testlane != lane:
                    #             #Not our problem as it's in another lane, skip
                    #             olddataind += 1
                    #             continue
                    #         nonExitEdgeDetections[edge][0].pop(olddataind) #Else it's a non-adopter in this lane that we should've seen by now but haven't; delete. (TODO try relocating instead??)
                else:
                    if len(wasFull[detector]) == 0 or wasFull[detector][-1] < simtime - 1: #Make sure we don't have the same car just sitting on the detector forever
                        allSmart = True
                        for testid in ids:
                            if not testid in isSmart or not isSmart[testid]:
                                allSmart = False
                                break
                        if not allSmart:
                            wasFull[detector].append(simtime)
                        else:
                            #We want our stats to only count the non-adopters (assuming we can get that level of precision), as adopters might do different stuff
                            #Don't need to update last time in wasFull - if a non-smart shows up, we're okay with detecting it immediately
                            pass
                    else:
                        wasFull[detector][-1] = simtime #If someone has been sitting here forever, update the last timestamp to be now so we don't get duplicate readings

        surtracFreq = mainSurtracFreq #Period between updates in main SUMO sim
        if simtime%surtracFreq >= (simtime+1)%surtracFreq:
            temp = doSurtrac(network, simtime, None, None, mainlastswitchtimes, sumoPredClusters, False, deepcopy(nonExitEdgeDetections))
            #Don't bother storing toUpdate = temp[0], since doSurtrac has done that update already
            sumoPredClusters = temp[1]
            remainingDuration.update(temp[2])

        #Check for lights that switched phase (because previously-planned duration ran out, not because Surtrac etc. changed the plan); update custom data structures and current phase duration
        for light in lights:
            temp = traci.trafficlight.getPhase(light)
            if not(light in remainingDuration and len(remainingDuration[light]) > 0):
                #Only update remainingDuration if we have no schedule, in which case grab the actual remaining duration from SUMO
                remainingDuration[light] = [traci.trafficlight.getNextSwitch(light) - simtime]
            else:
                remainingDuration[light][0] -= 1
            if temp != lightphases[light]:
                mainlastswitchtimes[light] = simtime
                lightphases[light] = temp
                #Duration of previous phase was first element of remainingDuration, so pop that and read the next, assuming everything exists
                if light in remainingDuration and len(remainingDuration[light]) > 0:
                    #print(remainingDuration[light][0]) #Prints -1. Might be an off-by-one somewhere, but should be pretty close to accurate?
                    #NOTE: The light switches when remaining duration goes negative (in this case -1)
                    remainingDuration[light].pop(0)
                    if len(remainingDuration[light]) == 0:
                        remainingDuration[light] = [traci.trafficlight.getNextSwitch(light) - simtime]
                    else:
                        traci.trafficlight.setPhaseDuration(light, remainingDuration[light][0])
                else:
                    print("Unrecognized light " + light + ", this shouldn't happen")
        
        realdurations[simtime] = pickle.loads(pickle.dumps(remainingDuration))
        # if simtime in simdurations:
        #     print("DURRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRR")
        #     print(simdurations[simtime][lights[0]])
        #     print(realdurations[simtime][lights[0]])
        
        for car in traci.simulation.getStartingTeleportIDList():
            routeStats[car]["nTeleports"] += 1
            print("Warning: Car " + car + " teleported, time=" + str(simtime))
            isSmartVal = "Unknown"
            if car in isSmart:
                isSmartVal = isSmart[car]
            if car in laneDict and car in isSmart:
                teleportdata.append((car, simtime, laneDict[car], isSmartVal))
            else:
                try:
                    teleportdata.append((car, simtime, traci.vehicle.getLaneID(car)+"_TraCIlookup", isSmartVal))
                except:
                    print("TraCI lookup fail?")
                    teleportdata.append((car, simtime, "TraCI lookup failed, not sure what happened", isSmartVal))

        #Moving this to the bottom so we've already updated the vehicle locations (when we checked left turns)
        oldRemainingDuration = pickle.loads(pickle.dumps(remainingDuration))
        oldMainLastSwitchTimes = pickle.loads(pickle.dumps(mainlastswitchtimes))

        adopterinfo = dict()
        for vehicle in laneDict:
            try:
                if isSmart[vehicle]:
                    adopterinfo[vehicle] = [laneDict[vehicle], traci.vehicle.getLanePosition(vehicle), traci.vehicle.getSpeed(vehicle)] #lane, lanepos, speed
            except:
                print("get adopter info error")
                pass

        #testNonExitEdgeDetections = pickle.loads(pickle.dumps(nonExitEdgeDetections))
        #assert(testNonExitEdgeDetections == nonExitEdgeDetections)
        reroute(rerouters, network, simtime, remainingDuration, sumoPredClusters) #Reroute cars (including simulate-ahead cars)
        #assert(testNonExitEdgeDetections == nonExitEdgeDetections)

        assert(remainingDuration == oldRemainingDuration)
        assert(mainlastswitchtimes == oldMainLastSwitchTimes)

        #Grab data about vehicle behavior near intersections
        if dumpIntersectionData:
            nearInt = 20 #Look at data within this many meters of intersection
            for vehicle in traci.vehicle.getIDList():

                if traci.vehicle.getLanePosition(vehicle) > lengths[laneDict[vehicle]] - nearInt:
                    #Start recording
                    vehicleIntersectionData[vehicle] = [traci.vehicle.getLaneID(vehicle), traci.vehicle.getDistance(vehicle)-traci.vehicle.getLanePosition(vehicle)+(lengths[laneDict[vehicle]] - nearInt), [], [], [], [], []] #startlane, startdist, X, Y, dist, speed, acceleration
                if vehicle in vehicleIntersectionData and vehicleIntersectionData[vehicle] != None:
                    lane = traci.vehicle.getLaneID(vehicle)
                    #Store vehicle data
                    xy = traci.vehicle.getPosition(vehicle)
                    vehicleIntersectionData[vehicle][2].append(xy[0])
                    vehicleIntersectionData[vehicle][3].append(xy[1])
                    vehicleIntersectionData[vehicle][4].append(traci.vehicle.getDistance(vehicle)-vehicleIntersectionData[vehicle][1])
                    vehicleIntersectionData[vehicle][5].append(traci.vehicle.getSpeed(vehicle))
                    vehicleIntersectionData[vehicle][6].append(traci.vehicle.getAcceleration(vehicle))
                    

                    if len(lane) > 0 and lane[0] != ":":
                        if not lane in intersectionData:
                            intersectionData[lane] = dict()
                        if lane != vehicleIntersectionData[vehicle][0] and traci.vehicle.getLanePosition(vehicle) > nearInt:
                            #We're done, dump data
                            if not lane in intersectionData[vehicleIntersectionData[vehicle][0]]:
                                intersectionData[vehicleIntersectionData[vehicle][0]][lane] = []
                            intersectionData[vehicleIntersectionData[vehicle][0]][lane].append(vehicleIntersectionData[vehicle][2:])
                            vehicleIntersectionData[vehicle] = None

        #Plot and print stats
        if simtime%100 == 0 or not traci.simulation.getMinExpectedNumber() > 0:
            #After we're done simulating... 
            plt.figure()
            plt.plot(carsOnNetwork)
            plt.xlabel("Time (s)")
            plt.ylabel("Cars on Network")
            plt.title("Congestion, Adoption Prob=" + str(pSmart))
            #plt.show() #NOTE: Blocks code execution until you close the plot
            plt.savefig("Plots/Congestion, AP=" + str(pSmart)+".png")
            plt.close()


            #Stats
            avgTime = 0
            avgLefts = 0
            bestTime = inf
            worstTime = 0

            avgTimeSmart = 0
            avgLeftsSmart = 0
            bestTimeSmart = inf
            worstTimeSmart = 0
            avgTimeNot = 0
            avgLeftsNot = 0
            bestTimeNot = inf
            worstTimeNot = 0

            totalcalls = 0
            totalcallsafterfirst = 0
            totalcallsfirst = 0
            totalswaps = 0
            totalswapsafterfirst = 0
            totalswapsfirst = 0
            nswapped = 0

            avgTime2 = 0
            avgTimeSmart2 = 0
            avgTimeNot2 = 0

            avgTime3 = 0
            avgTimeSmart3 = 0
            avgTimeNot3 = 0

            avgTime0 = 0
            avgTimeSmart0 = 0
            avgTimeNot0 = 0

            nCars = 0
            nSmart = 0
            ntimeouts = 0
            nsmartteleports = 0
            nnotsmartteleports = 0
            nteleports = 0

            avgerror = 0
            avgabserror = 0
            avgpcterror = 0
            avgabspcterror = 0

            totaldistance = 0
            totaldistanceSmart = 0
            totaldistanceNot = 0

            for id in endDict:
                if actualStartDict[id] >= 600 and actualStartDict[id] < 3000:
                    nCars += 1
                    if isSmart[id]:
                        nSmart += 1

            for id in endDict:
                #Only look at steady state - ignore first and last 10 minutes of cars
                if actualStartDict[id] < 600 or actualStartDict[id] >= 3000:
                    continue

                ntimeouts += routeStats[id]["nTimeouts"]
                nteleports += routeStats[id]["nTeleports"]
                if isSmart[id]:
                    nsmartteleports += routeStats[id]["nTeleports"]
                else:
                    nnotsmartteleports += routeStats[id]["nTeleports"]

                ttemp = (endDict[id] - startDict[id])+delayDict[id]
                avgTime += ttemp/nCars
                avgLefts += leftDict[id]/nCars
                if ttemp > worstTime:
                    worstTime = ttemp
                if ttemp < bestTime:
                    bestTime = ttemp

                if ttemp < 0:
                    print("Negative ttemp (=delay)?")
                    print(id)

                if isSmart[id]:
                    avgTimeSmart += ttemp/nSmart
                    avgLeftsSmart += leftDict[id]/nSmart
                    if ttemp > worstTimeSmart:
                        worstTimeSmart = ttemp
                    if ttemp < bestTimeSmart:
                        bestTimeSmart = ttemp
                else:
                    avgTimeNot += ttemp/(nCars-nSmart)
                    avgLeftsNot += leftDict[id]/(nCars-nSmart)
                    if ttemp > worstTimeNot:
                        worstTimeNot = ttemp
                    if ttemp < bestTimeNot:
                        bestTimeNot = ttemp

                #Delay2 computation (start clock after first intersection)
                if not id in delay2adjdict:
                    delay2adjdict[id] = startDict[id]
                ttemp2 = (endDict[id] - delay2adjdict[id])+delayDict[id]
                avgTime2 += ttemp2/nCars
                if isSmart[id]:
                    avgTimeSmart2 += ttemp2/nSmart
                else:
                    avgTimeNot2 += ttemp2/(nCars-nSmart)

                #Delay3 computation (start clock after first routing call)
                if not id in delay3adjdict:
                    delay3adjdict[id] = startDict[id]
                ttemp3 = (endDict[id] - delay3adjdict[id])+delayDict[id]
                avgTime3 += ttemp3/nCars
                if isSmart[id]:
                    avgTimeSmart3 += ttemp3/nSmart
                else:
                    avgTimeNot3 += ttemp3/(nCars-nSmart)

                #Delay0 computation (start clock at intended entrance time)
                ttemp0 = (endDict[id] - actualStartDict[id])+delayDict[id]
                avgTime0 += ttemp0/nCars
                if isSmart[id]:
                    avgTimeSmart0 += ttemp0/nSmart
                else:
                    avgTimeNot0 += ttemp0/(nCars-nSmart)

                totalcalls += routeStats[id]["nCalls"]
                totalcallsafterfirst += routeStats[id]["nCallsAfterFirst"]
                totalcallsfirst += routeStats[id]["nCallsFirst"]
                totalswaps += routeStats[id]["nSwaps"]
                totalswapsafterfirst += routeStats[id]["nSwapsAfterFirst"]
                totalswapsfirst += routeStats[id]["nSwapsFirst"]
                totaldistance += routeStats[id]["distance"]
                if isSmart[id]:
                    totaldistanceSmart += routeStats[id]["distance"]
                else:
                    totaldistanceNot += routeStats[id]["distance"]
                #Check which routes don't run into routing decisions at all
                # if isSmart[id] and routeStats[id]["nSwapsFirst"] == 0:
                #     print(currentRoutes[id])
                if routeStats[id]["swapped"] == True:
                    nswapped += 1

                try:
                    if isSmart[id]:
                        avgerror += ((timedata[id][1]-timedata[id][0]) - timedata[id][2])/nSmart
                        avgabserror += abs((timedata[id][1]-timedata[id][0]) - timedata[id][2])/nSmart
                        avgpcterror += ((timedata[id][1]-timedata[id][0]) - timedata[id][2])/(timedata[id][1]-timedata[id][0])/nSmart*100
                        avgabspcterror += abs((timedata[id][1]-timedata[id][0]) - timedata[id][2])/(timedata[id][1]-timedata[id][0])/nSmart*100
                except:
                    print("Missing timedata for vehicle " + id)
                    print(isSmart[id])
                    print(timedata[id])

            if verbose or not traci.simulation.getMinExpectedNumber() > 0 or (appendTrainingData and simtime == 5000):
                print(pSmart)
                print("\nCurrent simulation time: %f" % simtime)
                print("Total run time: %f" % (time.time() - tstart))
                print("Number of vehicles in network: %f" % traci.vehicle.getIDCount())
                print("Total cars that left the network: %f" % len(endDict))
                print("Average delay: %f" % avgTime)
                print("Average delay0: %f" % avgTime0)
                print("Best delay: %f" % bestTime)
                print("Worst delay: %f" % worstTime)
                print("Average number of lefts: %f" % avgLefts)
                if nCars > 0:
                    print("Average number of calls to routing: %f" % (totalcalls/nCars))
                    if totalcalls > 0:
                        print("Proportion of timeouts in routing: %f" % (ntimeouts/totalcalls))
                    print("Average number of route changes: %f" % (totalswaps/nCars))
                    print("Average number of route changes after first routing decision: %f" % (totalswapsafterfirst/nCars))
                    print("Proportion of cars that changed route at least once: %f" % (nswapped/nCars))
                    print("Average number of teleports: %f" % (nteleports/nCars))
                    print("Average distance travelled: %f" % (totaldistance/nCars))
                print("Among adopters:")
                print("Average delay: %f" % avgTimeSmart)
                print("Best delay: %f" % bestTimeSmart)
                print("Worst delay: %f" % worstTimeSmart)
                print("Average number of lefts: %f" % avgLeftsSmart)
                if nSmart > 0:
                    print("Average error (actual minus expected) in predicted travel time: %f" % (avgerror))
                    print("Average absolute error in predicted travel time: %f" % (avgabserror))
                    print("Average percent error in predicted travel time: %f" % (avgpcterror))
                    print("Average absolute percent error in predicted travel time: %f" % (avgabspcterror))

                    if len(routeVerifyData) > 0:
                        avgVerifyError = 0
                        avgAbsVerifyError = 0
                        avgPctVerifyError = 0
                        avgPctAbsVerifyError = 0
                        nVerifyPts = 0
                        for verifytuple in routeVerifyData:
                            if verifytuple[0] < 0 or verifytuple[1] < 0: #If routing times out, these return -1
                                continue
                            avgVerifyError += (verifytuple[0]-verifytuple[1])
                            avgAbsVerifyError += abs((verifytuple[0]-verifytuple[1]))
                            avgPctVerifyError += (verifytuple[0]-verifytuple[1])/verifytuple[0]*100
                            avgPctAbsVerifyError += abs((verifytuple[0]-verifytuple[1]))/verifytuple[0]*100
                            nVerifyPts += 1
                        avgVerifyError /= nVerifyPts
                        avgAbsVerifyError /= nVerifyPts
                        avgPctVerifyError /= nVerifyPts
                        avgPctAbsVerifyError /= nVerifyPts
                        print("Average sim-to-sim error (actual minus expected) in predicted travel time: %f" % (avgVerifyError))
                        print("Average absolute sim-to-sim error (actual minus expected) in predicted travel time: %f" % (avgAbsVerifyError))
                        print("Average percent sim-to-sim error (actual minus expected) in predicted travel time: %f" % (avgPctVerifyError))
                        print("Average absolute percent sim-to-sim error (actual minus expected) in predicted travel time: %f" % (avgPctAbsVerifyError))
                            
                    print("Proportion of cars that changed route at least once: %f" % (nswapped/nSmart))
                    print("Average number of teleports: %f" % (nsmartteleports/nSmart))
                    print("Average distance travelled: %f" % (totaldistanceSmart/nSmart))
                    print("Average number of calls to routing: %f" % (totalcalls/nSmart)) #NOTE: This only counts calls done by cars that exited
                    print("Average number of route changes: %f" % (totalswaps/nSmart))
                    print("Average number of route changes after first routing decision: %f" % (totalswapsafterfirst/nSmart))

                if totalcalls > 0:
                    print("Proportion of timeouts in routing: %f" % (ntimeouts/totalcalls))
                    print("Proportion of routing decisions leading to a route change: %f" % (totalswaps/totalcalls))
                    if totalcallsfirst > 0:
                        print("Proportion of first routing decisions leading to a route change: %f" % (totalswapsfirst/totalcallsfirst))
                    else:
                        print("WARNING: Some routing calls, but no first routing calls; something's wrong with the stats!")
                    if totalcallsafterfirst > 0:
                        print("Proportion of routing decisions after first leading to a route change: %f" % (totalswapsafterfirst/totalcallsafterfirst))

                print("Among non-adopters:")
                print("Average delay: %f" % avgTimeNot)
                print("Best delay: %f" % bestTimeNot)
                print("Worst delay: %f" % worstTimeNot)
                print("Average number of lefts: %f" % avgLeftsNot)
                if nCars - nSmart > 0:
                    print("Average number of teleports: %f" % (nnotsmartteleports/(nCars-nSmart)))
                    print("Average distance travelled: %f" % (totaldistanceNot/(nCars-nSmart)))

                print("Routing calls (including cars not yet exited): " + str(nRoutingCalls))
                print("Total routing time: " + str(routingTime))
                if nRoutingCalls > 0:
                    print("Average time per call: " + str(routingTime/nRoutingCalls))
                    print("Proportion of successful routing calls: " + str(nSuccessfulRoutingCalls/nRoutingCalls))
                if debugMode:
                    print("Surtrac calls (one for each light): " + str(totalSurtracRuns))
                    if totalSurtracRuns > 0:
                        print("Average time per Surtrac call: " + str(totalSurtracTime/totalSurtracRuns))
                        print("Average number of clusters per Surtrac call: " + str(totalSurtracClusters/totalSurtracRuns))
                    print("loadClusters calls: " + str(totalLoadRuns))
                    if totalLoadRuns > 0:
                        print("Average time per loadClusters call: " + str(totalLoadTime/totalLoadRuns))
                        print("Average cars per loadClusters call: " + str(totalLoadCars/totalLoadRuns))
                print("\n")

                for lane in arrivals:
                    while len(arrivals[lane]) > 0 and arrivals[lane][0] < simtime - maxarrivalwindow:
                        arrivals[lane] = arrivals[lane][1:]

                for lane in arrivals2:
                    while len(arrivals2[lane]) > 0 and arrivals2[lane][0] < simtime - maxarrivalwindow2:
                        arrivals2[lane] = arrivals2[lane][1:]

    #Dump intersection data to Excel
    if dumpIntersectionData:
        dumpIntersectionDataFun(intersectionData, network)

    try:
        os.remove("savestates/teststate_"+savename+".xml") #Delete the savestates file so we don't have random garbage building up over time
    except FileNotFoundError:
        print("Warning: Trying to clean up savestates file, but no file found. This is weird - did you comment out routing or something? Ignoring for now.")
        pass

    return [avgTime, avgTimeSmart, avgTimeNot, avgTime2, avgTimeSmart2, avgTimeNot2, avgTime3, avgTimeSmart3, avgTimeNot3, avgTime0, avgTimeSmart0, avgTimeNot0, time.time()-tstart, nteleports, teleportdata]  

def dumpIntersectionDataFun(intersectionData, network):
    print("Writing intersection data to spreadsheet")
    for lane in intersectionData:
        for nextlane in intersectionData[lane]:
            try:
                book = openpyxl.Workbook()
                sheets = dict()
                labels = ["X", "Y", "Distance", "Speed", "Acceleration"]
                for ind in range(len(labels)):
                    sheets[ind] = book.create_sheet(labels[ind], -1)
                    for row in range(len(intersectionData[lane][nextlane])):
                        for col in range(len(intersectionData[lane][nextlane][row][ind])):
                            sheets[ind].cell(row+1, col+1, intersectionData[lane][nextlane][row][ind][col]) #+1 because Excel 1-indexes
            except Exception as e:
                print(e)
                print("Error dumping training data to Excel, ignoring and continuing")
            finally:
                book.save("intersectiondata/"+lane+"|"+nextlane+".xlsx")

    thetarows = dict()
    thetabooks = dict()
    thetasheets = dict()
    try:
        for lane in intersectionData:
            for nextlane in intersectionData[lane]:
                dtheta = getDTheta(lane.split("_")[0], nextlane.split("_")[0], network) * 180/math.pi
            
                if not dtheta in thetabooks:
                    book = openpyxl.Workbook()
                    thetabooks[dtheta] = book
                    thetarows[dtheta] = 0
                    thetasheets[dtheta] = dict()
                    labels = ["X", "Y", "Distance", "Speed", "Acceleration"]
                    for ind in range(len(labels)):
                        thetasheets[dtheta][ind] = book.create_sheet(labels[ind], -1)
                 
                for row in range(len(intersectionData[lane][nextlane])):
                    for ind in range(len(labels)):
                        for col in range(len(intersectionData[lane][nextlane][row][ind])):
                            thetasheets[dtheta][ind].cell(thetarows[dtheta]+1, col+1, intersectionData[lane][nextlane][row][ind][col]) #+1 because Excel 1-indexes
                    thetarows[dtheta] += 1
    except Exception as e:
        print(e)
        print("Error dumping training data to Excel, ignoring and continuing")
    finally:
        for dtheta in thetabooks:
            thetabooks[dtheta].save("intersectiondata/theta"+str(math.floor(dtheta))+".xlsx")

#@profile
def loadClusters(net, simtime, VOI=None):
    global totalLoadCars
    global nVehicles
    #Load locations of cars and current traffic light states into custom data structures
    #If given, VOI is the vehicle triggering the routing call that triggered this, and needs to be unaffected when we add noise
    #TODO: We're caching the loaded clusters, which means we'll need to be better about not adding noise to any vehicles that could potentially be routed
    clusters = dict()
    nVehicles.append(0)

    #Cluster data structures
    for edge in edges:
        if edge[0] == ":": #Edge is an internal edge, so we're inside an intersection

            #Add those vehicles to the nVehicles total, but that's it
            for lanenum in range(nLanes[edge]):
                lane = edge + "_" + str(lanenum)
                temp = traci.lane.getLastStepVehicleIDs(lane)
                nVehicles[-1] += len(temp)
            continue

        for lanenum in range(nLanes[edge]):
            lane = edge + "_" + str(lanenum)
            clusters[lane] = []
            temp = traci.lane.getLastStepVehicleIDs(lane)
            totalLoadCars += len(temp)
            #if lane in nonExitEdgeDetections: #Needed this to confirm that nVehicles from this matched nVehicles from the detectors version
            nVehicles[-1] += len(temp)
            for vehicle in reversed(temp): #Reversed so we go from end of edge to start of edge - first clusters to leave are listed first
                #Process vehicle into cluster somehow
                #If nearby cluster, add to cluster in sorted order (could probably process in sorted order)
                lanepos = traci.vehicle.getLanePosition(vehicle)
                if len(clusters[lane]) > 0 and abs(clusters[lane][-1]["time"] - simtime) < clusterthresh and abs(clusters[lane][-1]["endpos"] - lanepos)/speeds[edge] < clusterthresh:
                    #Add to cluster. pos and time track newest added vehicle to see if the next vehicle merges
                    #Departure time (=time to fully clear cluster) increases, arrival doesn't
                    clusters[lane][-1]["endpos"] = lanepos
                    clusters[lane][-1]["time"] = simtime
                    clusters[lane][-1]["departure"] = simtime + (lengths[lane]-clusters[lane][-1]["endpos"])/speeds[edge]
                    clusters[lane][-1]["cars"].append((vehicle, clusters[lane][-1]["departure"], 1, "Load append"))
                    clusters[lane][-1]["weight"] = len(clusters[lane][-1]["cars"])
                else:
                    #Else make a new cluster
                    newcluster = dict()
                    newcluster["startpos"] = lanepos
                    newcluster["endpos"] = lanepos
                    newcluster["time"] = simtime
                    newcluster["arrival"] = simtime + (lengths[edge+"_0"]-newcluster["endpos"])/speeds[edge]
                    newcluster["departure"] = newcluster["arrival"]
                    newcluster["cars"] = [(vehicle, newcluster["departure"], 1, "Load new")]
                    newcluster["weight"] = len(newcluster["cars"])
                    clusters[lane].append(newcluster)
                assert(clusters[lane][-1]["departure"] > simtime - 1e-10)
                assert(clusters[lane][-1]["departure"] > clusters[lane][-1]["arrival"] - 1e-10)

    
    #Traffic light info
    lightphases = dict()
    for light in lights:
        lightphases[light] = traci.trafficlight.getPhase(light)

    #Add noise
    #clusters = addNoise(clusters, VOI, 0.9, 2) #To simulate detector stuff
    return (clusters, lightphases)

#This is currently only used for Surtrac; should be another function that handles starting routing sims
def loadClustersDetectors(net, simtime, nonExitEdgeDetections3, VOI=None):
    global totalLoadCars
    global nVehicles
    #Load locations of cars and current traffic light states into custom data structures
    #If given, VOI is the vehicle triggering the routing call that triggered this, and needs to be unaffected when we add noise
    #TODO: We're caching the loaded clusters, which means we'll need to be better about not adding noise to any vehicles that could potentially be routed
    clusters = dict()
    nVehicles.append(0)

    #Cluster data structures
    totallanedata = dict()
    for edge in nonExitEdgeDetections3: #Assuming exit lanes don't matter since they shouldn't have traffic - this saves us from extra exit detectors at their ends
        totallanedata[edge] = 0
        for laneind in range(nLanes[edge]):
            lane = edge + "_" + str(laneind)
            clusters[lane] = []
            totallanedata[edge] += len(wasFull[nonExitLaneDetectors[lane][1][0]]) + 1 #[lane][1][0] because 1 is the index of the exit detector and 0 is the index of its name. +1 as a permanent psuedocount on all detectors, mostly in case we have no data whatsoever
    for edge in nonExitEdgeDetections3: #Assuming exit lanes don't matter since they shouldn't have traffic - this saves us from extra exit detectors at their ends
        temp = nonExitEdgeDetections3[edge]
        totalLoadCars += len(temp)
        nVehicles[-1] += len(temp)
        for roadsectionind in reversed(range(len(temp))): #First road section (=closest to start) is listed first, and we want to go backwards from the end, so reverse this
            roadsectiondata = temp[roadsectionind]
            for (vehicle, detlane, detecttime) in roadsectiondata: #Earliest time (=farthest along road) is listed first, don't reverse this
                #Sample a lane randomly
                if not vehicle in isSmart or not isSmart[vehicle] or not adopterCommsSurtrac:
                    r = random.random()
                    for laneind in range(nLanes[edge]):
                        lane = edge + "_" + str(laneind)
                        r -= (len(wasFull[nonExitLaneDetectors[lane][1][0]]) + 1)/totallanedata[edge]
                        if r < 0:
                            break #lane is now equal to a lane sampled from the lane change probabilities data from wasFull

                    assert(detlane.split("_")[0] == edge)
                    assert(lane.split("_")[0] == edge)

                    #Process vehicle into cluster somehow
                    #If nearby cluster, add to cluster in sorted order (could probably process in sorted order)
                    startOfSegment = nonExitLaneDetectors[lane][roadsectionind][1]
                    endOfSegment = lengths[lane]
                    if roadsectionind < len(temp)-1:
                        endOfSegment = nonExitLaneDetectors[lane][roadsectionind+1][1]
                    lanepos = min(endOfSegment, speeds[edge] * (simtime - detecttime+0.5)+startOfSegment) #+0.5 because we crossed the detector, then made somewhere between 0 and 1 seconds worth of forward movement; estimate it as 0.5

                    #lanepos = traci.vehicle.getLanePosition(vehicle)
                else:
                    #It's an adopter; we assume it shares its data
                    try:
                        lane = traci.vehicle.getLaneID(vehicle)
                        lanepos = traci.vehicle.getLanePosition(vehicle)
                    except:
                        print("Failing to look up adopter data")
                        print(simtime)
                        print(vehicle)
                        print(laneDict[vehicle])

                        #Not sure what happened; pretend it's a non-adopter?
                        r = random.random()
                        for laneind in range(nLanes[edge]):
                            lane = edge + "_" + str(laneind)
                            r -= (len(wasFull[nonExitLaneDetectors[lane][1][0]]) + 1)/totallanedata[edge]
                            if r < 0:
                                break #lane is now equal to a lane sampled from the lane change probabilities data from wasFull

                        assert(detlane.split("_")[0] == edge)
                        assert(lane.split("_")[0] == edge)

                        #Process vehicle into cluster somehow
                        #If nearby cluster, add to cluster in sorted order (could probably process in sorted order)
                        startOfSegment = nonExitLaneDetectors[lane][roadsectionind][1]
                        endOfSegment = lengths[lane]
                        if roadsectionind < len(temp)-1:
                            endOfSegment = nonExitLaneDetectors[lane][roadsectionind+1][1]
                        lanepos = min(endOfSegment, speeds[edge] * (simtime - detecttime+0.5)+startOfSegment) #+0.5 because we crossed the detector, then made somewhere between 0 and 1 seconds worth of forward movement; estimate it as 0.5


                if not lane.split("_")[0] in nonExitEdgeDetections3:
                    continue #Car in intersection, not sure what to do so I'll ignore it (which I'm pretty sure is what I did with omniscient Surtrac)
                    #Or it's on an exit lane, but we don't care

                if len(clusters[lane]) > 0 and abs(clusters[lane][-1]["time"] - simtime) < clusterthresh and abs(clusters[lane][-1]["endpos"] - lanepos)/speeds[edge] < clusterthresh:
                    #Add to cluster. pos and time track newest added vehicle to see if the next vehicle merges
                    #Departure time (=time to fully clear cluster) increases, arrival doesn't
                    clusters[lane][-1]["endpos"] = lanepos
                    clusters[lane][-1]["time"] = simtime
                    clusters[lane][-1]["departure"] = simtime + (lengths[lane]-clusters[lane][-1]["endpos"])/speeds[edge]
                    clusters[lane][-1]["cars"].append((vehicle, clusters[lane][-1]["departure"], 1, "Load append"))
                    clusters[lane][-1]["weight"] = len(clusters[lane][-1]["cars"])
                else:
                    #Else make a new cluster
                    newcluster = dict()
                    newcluster["startpos"] = lanepos
                    newcluster["endpos"] = lanepos
                    newcluster["time"] = simtime
                    newcluster["arrival"] = simtime + (lengths[edge+"_0"]-newcluster["endpos"])/speeds[edge]
                    newcluster["departure"] = newcluster["arrival"]
                    newcluster["cars"] = [(vehicle, newcluster["departure"], 1, "Load new")]
                    newcluster["weight"] = len(newcluster["cars"])
                    clusters[lane].append(newcluster)
                #assert(clusters[lane][-1]["departure"] > simtime - 1e-10)
                # print(clusters[lane][-1]["departure"])
                # print(clusters[lane][-1]["arrival"])
                # print(clusters[lane][-1]["departure"] - clusters[lane][-1]["arrival"])
                # print(clusters[lane][-1])
                #assert(clusters[lane][-1]["departure"] > clusters[lane][-1]["arrival"] - 1e-10)
    
    #Traffic light info
    lightphases = dict()
    for light in lights:
        lightphases[light] = traci.trafficlight.getPhase(light)

    #Add noise
    #clusters = addNoise(clusters, VOI, 0.9, 2) #To simulate detector stuff

    return (clusters, lightphases)

# def addNoise(clusters, VOI, detectprob, timeerr):
#     #Randomly delete cars with probability noiseprob
#     #Randomly clone non-deleted cars to make up for it

#     if reuseSurtrac:
#         print("Warning: We might be randomly deleting a vehicle to be routed later. TODO fix this...")

#     #Cluster data structures
#     for edge in edges:
#         if edge[0] == ":":
#             #Skip internal edges (=edges for the inside of each intersection)
#             continue
#         for lanenum in range(nLanes[edge]):
#             lane = edge + "_" + str(lanenum)
#             for clusternum in range(len(clusters[lane])):
#                 cluster = clusters[lane][clusternum]
#                 noisycluster = pickle.loads(pickle.dumps(cluster))
#                 noisycluster["cars"] = []
#                 for car in cluster["cars"]:
#                     if car[0] == VOI:
#                         #Don't perturb the VOI
#                         noisycluster["cars"].append(car)
#                         continue
#                     noisecar = (car[0], car[1] + random.random()*timeerr*2-timeerr, car[2], "noisetest") #Because tuples are immutable...
#                     if random.random() < 1-detectprob:
#                         #Don't add car to noisycluster
#                         noisycluster["weight"] -= noisecar[2]
#                         continue
#                     noisycluster["cars"].append(noisecar)
#                     #Duplicate car, potentially multiple times, to estimate deleted traffic
#                     while random.random() < 1-detectprob:
#                         #Duplicate this car
#                         noisycluster["cars"].append(noisecar)
#                         noisycluster["weight"] += noisecar[2]
#                 clusters[lane][clusternum] = noisycluster
#     return clusters

def LAISB(a, b):
    #Line a intersects segment b
    #a is a tuple of points on the line; b is endpoints of segment
    #Negative if they intersect, positive if they don't, zero if an endpoint touches
    va = [ a[1][0] - a[0][0], a[1][1]-a[0][1] ]
    vb0 = [ b[0][0] - a[0][0], b[0][1]-a[0][1] ]
    vb1 = [ b[1][0] - a[0][0], b[1][1]-a[0][1] ]
    return np.cross(vb0, va) * np.cross(vb1, va)

def isIntersecting(a, b):
    #a and b are tuples of endpoints of line segments
    iab = LAISB(a, b)
    iba = LAISB(b, a)
    if iab > 0 or iba > 0:
        return False
    if iab == 0 and iba == 0:
        #Colinear, yuck
        return (min(a[0][0], a[1][0]) <= max(b[0][0], b[1][0]) and min(b[0][0], b[1][0]) <= max(a[0][0], a[1][0]) and
        min(a[0][1], a[1][1]) <= max(b[0][1], b[1][1]) and min(b[0][1], b[1][1]) <= max(a[0][1], a[1][1]))
    #Each segment either touches or crosses the other line, so we're good
    return True

def get_options():
    optParser = optparse.OptionParser()
    optParser.add_option("--nogui", action="store_true",
                         default=False, help="run the commandline version of sumo")
    options, args = optParser.parse_args()
    return options

#Generates induction loops on all the edges
def generate_additionalfile(sumoconfig, networkfile):
    #Create a third instance of a simulator so I can query the network

    if useLibsumo:
        traci.start([checkBinary('sumo'), "-c", sumoconfig,
                                    "--start", "--no-step-log", "true",
                                    "--xml-validation", "never", "--quit-on-end"])
    else:
        try:
            traci.start([checkBinary('sumo'), "-c", sumoconfig,
                                    "--start", "--no-step-log", "true",
                                    "--xml-validation", "never", "--quit-on-end"], label="setup")
        except:
            #Worried about re-calling this without old setup instance being removed
            #traci.switch("setup")
            pass

    net = sumolib.net.readNet(networkfile)
    rerouters = dict()
    global max_edge_speed

    #Copying this from run() so I can use these in here too. Annoyingly, lengths needs a SUMO simulation to compute, and the SUMO sim needs to know about the additional file, so ordering these is annoying
    #Edges have speeds, but lanes have lengths, so it's a little annoying to get fftimes...
    edges = traci.edge.getIDList()
    lanes = traci.lane.getIDList()

    for lane in lanes:
        if not lane[0] == ":":
            links[lane] = traci.lane.getLinks(lane)
            lengths[lane] = traci.lane.getLength(lane)

    for edge in edges:
        nLanes[edge] = traci.edge.getLaneNumber(edge)
        if not edge[0] == ":":
            speeds[edge] = net.getEdge(edge).getSpeed()
            fftimes[edge] = lengths[edge+"_0"]/speeds[edge]

    with open("additional_autogen.xml", "w") as additional:
        print("""<additional>""", file=additional)
        if not useLibsumo:
            print('    <edgeData id="%s" file="%s" period="%i"/>' % (savename, "edgedata/"+savename+".xml", 1e6), file=additional)
            print('    <laneData id="%s" file="%s" period="%i"/>' % (savename, "lanedata/"+savename+".xml", 1e6), file=additional)
        for edge in traci.edge.getIDList():
            if edge[0] == ":":
                #Skip internal edges (=edges for the inside of each intersection)
                continue

            if (net.getEdge(edge).getSpeed() > max_edge_speed):
                max_edge_speed = net.getEdge(edge).getSpeed()

            for lanenum in range(traci.edge.getLaneNumber(edge)):
                lane = edge+"_"+str(lanenum)
                print('    <inductionLoop id="IL_%s" freq="1" file="outputAuto.xml" lane="%s" pos="-%i" friendlyPos="true" />' \
                      % (lane, lane, detectordist), file=additional)
                if len(net.getEdge(edge).getOutgoing()) > 1:
                    rerouters["IL_"+lane] = lane
                    rerouterLanes["IL_"+lane] = lane
                    rerouterEdges["IL_"+lane] = edge

                if len(net.getEdge(edge).getOutgoing()) > 0:
                    nonExitEdgeDetections[edge] = []
                    nonExitLaneDetectors[lane] = []
                    for dist in [0, lengths[lane]-2]: #Add to this if we need more detectors, remember to update it both here and below in additionalrouting_autogen
                        name = "ILd_" + lane + "_" + str(dist)
                        print('    <inductionLoop id="%s" freq="1" file="outputAuto.xml" lane="%s" pos="%i" friendlyPos="true" />' \
                        % (name, lane, dist), file=additional)
                        nonExitEdgeDetections[edge].append([]) #This is inefficient since we keep clearing and rebuilding nonExitEdgeDetections[edge] once per lane, but shouldn't be that terrible
                        nonExitLaneDetectors[lane].append((name, dist))
                        wasFull[name] = []
        print("</additional>", file=additional)

    with open("additionalrouting_autogen.xml", "w") as additional:
        print("""<additional>""", file=additional)
        for edge in traci.edge.getIDList():
            if edge[0] == ":":
                #Skip internal edges (=edges for the inside of each intersection)
                continue

            for lanenum in range(traci.edge.getLaneNumber(edge)):
                lane = edge+"_"+str(lanenum)
                print('    <inductionLoop id="IL_%s" freq="1" file="outputAuto.xml" lane="%s" pos="-%i" friendlyPos="true" />' \
                      % (lane, lane, detectordist), file=additional)
                if len(net.getEdge(edge).getOutgoing()) > 0:
                    for dist in [0, lengths[lane]-2]: #Add to this if we need more detectors, remember to update it both here and above in additional_autogen
                        name = "ILd_" + lane + "_" + str(dist)
                        print('    <inductionLoop id="%s" freq="1" file="outputAuto.xml" lane="%s" pos="%i" friendlyPos="true" />' \
                        % (name, lane, dist), file=additional)
        print("</additional>", file=additional)
    
    return rerouters

def sampleRouteFromTurnData(startlane, turndata):
    lane = startlane
    route = [lane.split("_")[0]]
    while lane in turndata:
        r = random.random()

        oops = True
        for nextlane in turndata[lane]:
            r -= turndata[lane][nextlane]
            if r <= 0:
                if nextlane.split("_")[0] == lane.split("_")[0]:
                    print("Warning: Sampling is infinite looping, stopping early")
                    return route

                #Check if lane connects to nextlane
                for nextlinktuple in links[lane]:
                    tempnextedge = nextlinktuple[0].split("_")[0]
                    if nextlane.split("_")[0] == tempnextedge:
                        oops = False
                        break
                if oops:
                    print("sampleRouteFromTurnData found an invalid connection??? Trying again...")
                    print(lane + " -> " + nextlane)
                    break
                # else:
                #     print("Yay, did a thing!!!")
                #     print(lane + " -> " + nextlane)
                lane = nextlane
                break
        if not oops:
            route.append(lane.split("_")[0])
        #print(route)
    return route

def readSumoCfg(sumocfg):
    netfile = ""
    roufile = ""
    with open(sumocfg, "r") as cfgfile:
        lines = cfgfile.readlines()
        for line in lines:
            if "net-file" in line:
                data = line.split('"')
                netfile = data[1]
            if "route-files" in line: #This is scary - probably breaks if there's many of them
                data = line.split('"')
                roufile = data[1]
    return (netfile, roufile)

def main(sumoconfigin, pSmart, verbose = True, useLastRNGState = False, appendTrainingDataIn = False):
    global lowprioritygreenlightlinks
    global prioritygreenlightlinks
    global edges
    global lanes
    global turndata
    global actualStartDict
    global trainingdata
    global testNN
    global appendTrainingData
    global noNNinMain
    global sumoconfig
    global netfile

    sumoconfig = sumoconfigin
    #options = get_options()

    if useLastRNGState:
        with open("lastRNGstate.pickle", 'rb') as handle:
            rngstate = pickle.load(handle)
            random.setstate(rngstate)
    else:
        rngstate = random.getstate()
        with open("lastRNGstate.pickle", 'wb') as handle:
            pickle.dump(rngstate, handle, protocol=pickle.HIGHEST_PROTOCOL)
    
    appendTrainingData = appendTrainingDataIn
    if appendTrainingDataIn:
        #We're training, so overwrite whatever else we're doing
        noNNinMain = False
        debugNNslowness = False

    # this script has been called from the command line. It will start sumo as a
    # server, then connect and run

    #sumoBinary = checkBinary('sumo')
    sumoBinary = checkBinary('sumo-gui')
    #NOTE: Script name is zeroth arg

    (netfile, routefile) = readSumoCfg(sumoconfig)

    network = sumolib.net.readNet(netfile)
    net = network
    rerouters = generate_additionalfile(sumoconfig, netfile)


    # this is the normal way of using traci. sumo is started as a
    # subprocess and then the python script connects and runs
    if useLibsumo:
        traci.load(["-c", sumoconfig,
                                "--additional-files", "additional_autogen.xml",
                                "--no-step-log", "true",
                                "--log", "LOGFILE", "--xml-validation", "never", "--start", "--quit-on-end"])
    else:
        try:
            traci.start([sumoBinary, "-c", sumoconfig,
                                    "--additional-files", "additional_autogen.xml",
                                    "--no-step-log", "true",
                                    "--log", "LOGFILE", "--xml-validation", "never", "--start", "--quit-on-end"], label="main")
            #Second simulator for running tests. No GUI
            #traci.start([sumoBinary, "-c", sumoconfig, #GUI in case we need to debug
            traci.start([checkBinary('sumo'), "-c", sumoconfig, #No GUI
                                    "--additional-files", "additionalrouting_autogen.xml",
                                    "--start", "--no-step-log", "true",
                                    "--xml-validation", "never", "--quit-on-end",
                                    "--step-length", "1"], label="test")
            dontBreakEverything()
        except:
            #Worried about re-calling this without old main instance being removed
            if not useLibsumo:
                traci.switch("main")
            traci.load([ "-c", sumoconfig,
                                    "--additional-files", "additional_autogen.xml",
                                    "--no-step-log", "true",
                                    #"--time-to-teleport", "-1",
                                    "--log", "LOGFILE", "--xml-validation", "never", "--start", "--quit-on-end"])
            if not useLibsumo:
                traci.switch("test")
            traci.load([ "-c", sumoconfig,
                                    "--additional-files", "additionalrouting_autogen.xml",
                                    "--no-step-log", "true",
                                    #"--time-to-teleport", "-1",
                                    "--log", "LOGFILE", "--xml-validation", "never", "--start", "--quit-on-end"])
            dontBreakEverything()

    

    for node in sumolib.xml.parse(netfile, ['junction']):
        if node.type == "traffic_light":
            lights.append(node.id)
        else:
            notLights.append(node.id)
            notlightlanes[node.id] = []
            notlightoutlanes[node.id] = []

    for lane in sumolib.xml.parse(netfile, ['lane']):
        edge = lane.id.split("_")[0]
        if len(edge) == 0 or edge[0] == ":":
            continue
        toNode = network.getEdge(edge).getToNode()
        if toNode.getType() != "traffic_light":
            notlightlanes[toNode.getID()].append(lane.id)
        fromNode = network.getEdge(edge).getFromNode()
        if fromNode.getType() != "traffic_light":
            notlightoutlanes[fromNode.getID()].append(lane.id)

    #Grab stuff once at the start to avoid slow calls to traci in the routing
    edges = traci.edge.getIDList()
    lanes = traci.lane.getIDList()

    #Edges have speeds, but lanes have lengths, so it's a little annoying to get fftimes...
    for lane in lanes:
        if not lane[0] == ":":
            links[lane] = traci.lane.getLinks(lane)
            lengths[lane] = traci.lane.getLength(lane)

    for edge in edges:
        nLanes[edge] = traci.edge.getLaneNumber(edge)
        if not edge[0] == ":":
            speeds[edge] = network.getEdge(edge).getSpeed()
            fftimes[edge] = lengths[edge+"_0"]/speeds[edge]

    for lane in lanes:
        if not lane[0] == ":":
            fftimes[lane] = fftimes[lane.split("_")[0]]

    lowprioritygreenlightlinks = dict()
    prioritygreenlightlinks = dict()

    for light in lights:
        lightlinkconflicts[light] = dict()
        lightphasedata[light] = traci.trafficlight.getAllProgramLogics(light)[0].phases
        lightlinks[light] = traci.trafficlight.getControlledLinks(light)
        lightphases[light] = traci.trafficlight.getPhase(light)
        fftimes[light] = np.inf

        lightlanes[light] = []
        lightoutlanes[light] = []

        linklistlist = lightlinks[light]
        for linklist in linklistlist:

            for linktuple in linklist:
                inlane = linktuple[0]
                if not inlane in lightlanes[light]:
                    lightlanes[light].append(inlane)
                outlane = linktuple[1]
                if not outlane in lightoutlanes[light]:
                    lightoutlanes[light].append(outlane)
                    if fftimes[light] > fftimes[outlane]:
                        fftimes[light] = fftimes[outlane]

                lightlinkconflicts[light][linktuple] = dict()
                for linklist2 in linklistlist:
                    for linktuple2 in linklist2:
                        lightlinkconflicts[light][linktuple][linktuple2] = isIntersecting( (network.getLane(linktuple[0]).getShape()[1], (net.getLane(linktuple[1]).getShape()[0])), 
                        (net.getLane(linktuple2[0]).getShape()[1], (network.getLane(linktuple2[1]).getShape()[0])) )

    #Surtrac data
    for light in lights:
        surtracdata[light] = []
        mainlastswitchtimes[light] = 0
        lowprioritygreenlightlinks[light] = []
        prioritygreenlightlinks[light] = []
        lowprioritygreenlightlinksLE[light] = []
        prioritygreenlightlinksLE[light] = []

        n = len(lightphasedata[light])
        for i in range(n):
            surtracdata[light].append(dict())
            surtracdata[light][i]["minDur"] = 7#5#lightphasedata[light][i].minDur #Min duration of yellow
            surtracdata[light][i]["maxDur"] = 120#lightphasedata[light][i].maxDur #Max duration of everything

            if "G" in lightphasedata[light][i].state or "g" in lightphasedata[light][i].state:
                surtracdata[light][i]["minDur"] = 5 #1#3.5#5#lightphasedata[light][i].minDur #Min duration of green
            # if "y" in lightphasedata[light][i].state:
            #     surtracdata[light][i]["minDur"] = 7 #1#3.5#5#lightphasedata[light][i].minDur #Min duration of yellow
            # if "Y" in lightphasedata[light][i].state or "y" in lightphasedata[light][i].state:
            #     surtracdata[light][i]["minDur"] = 2#5#lightphasedata[light][i].minDur #There is no all-red phase, keep this long
            
            #Force yellow to be the min length - not doing this since it messes with the training data (triggers the min-max <= maxfreq condition where we'll break stuff with discretization)
            # if "Y" in lightphasedata[light][i].state or "y" in lightphasedata[light][i].state:
            #     surtracdata[light][i]["maxDur"] = surtracdata[light][i]["minDur"] #Force this to be the correct length. Don't think this matters though...

            surtracdata[light][i]["lanes"] = []
            lightstate = lightphasedata[light][i].state
            lowprioritygreenlightlinks[light].append([])
            prioritygreenlightlinks[light].append([])
            lowprioritygreenlightlinksLE[light].append(dict())
            prioritygreenlightlinksLE[light].append(dict())
            
            linklistlist = lightlinks[light]
            for linklistind in range(len(linklistlist)):
                linkstate = lightstate[linklistind]
                for linktuple in linklistlist[linklistind]:
                    if not linktuple[0] in lanephases:
                        lanephases[linktuple[0]] = []

                    if linkstate == "G" and not linktuple[0] in surtracdata[light][i]["lanes"]:
                        surtracdata[light][i]["lanes"].append(linktuple[0]) #[0][x]; x=0 is from, x=1 is to, x=2 is via
                        
                        lanephases[linktuple[0]].append(i)

                linklist = linklistlist[linklistind]
                for link in linklist:
                    if linkstate == "G":
                        prioritygreenlightlinks[light][i].append(link)
                        #Make sure lane->edge dictionary knows about this lane->edge pair
                        if not link[0] in prioritygreenlightlinksLE[light][i]:
                            prioritygreenlightlinksLE[light][i][link[0]] = dict()
                        if not link[1].split("_")[0] in prioritygreenlightlinksLE[light][i][link[0]]:
                            prioritygreenlightlinksLE[light][i][link[0]][link[1].split("_")[0]] = []
                        #Add to lane->edge dictionary
                        prioritygreenlightlinksLE[light][i][link[0]][link[1].split("_")[0]].append(link)
                    if linkstate == "g":
                        lowprioritygreenlightlinks[light][i].append(link)
                        #Make sure lane->edge dictionary knows about this lane->edge pair
                        if not link[0] in lowprioritygreenlightlinksLE[light][i]:
                            lowprioritygreenlightlinksLE[light][i][link[0]] = dict()
                        if not link[1].split("_")[0] in lowprioritygreenlightlinksLE[light][i][link[0]]:
                            lowprioritygreenlightlinksLE[light][i][link[0]][link[1].split("_")[0]] = []
                        #Add to lane->edge dictionary
                        lowprioritygreenlightlinksLE[light][i][link[0]][link[1].split("_")[0]].append(link)
                
            #Remove lanes if there's any direction that gets a non-green light ("g" is fine, single-lane left turns are just sad)
            for linklistind in range(len(linklistlist)):
                linkstate = lightstate[linklistind]
                for linktuple in linklistlist[linklistind]:

                    if not (linkstate == "G" or linkstate == "g") and linktuple[0] in surtracdata[light][i]["lanes"]: #NOTE: I'm being sloppy and assuming one-element lists of tuples, but I've yet to see a multi-element list here
                        surtracdata[light][i]["lanes"].remove(linktuple[0])
                        lanephases[linktuple[0]].remove(i)
                
        for i in range(n):
            #Compute min transition time between the start of any two phases
            surtracdata[light][i]["timeTo"] = [0]*n
            for joffset in range(1, n):
                j = (i + joffset) % n
                jprev = (j-1) % n
                surtracdata[light][i]["timeTo"][j] = surtracdata[light][i]["timeTo"][jprev] + surtracdata[light][jprev]["minDur"]

    if pSmart < 1 or True:
        with open("Lturndata_"+routefile.split(".")[0]+".pickle", 'rb') as handle:
            turndata = pickle.load(handle) #This is lane-to-lane normalized turndata in the format turndata[lane][nextlane]
            #When predicting adopters ahead, we know their next edge, but not the specific lane on that edge. Calculate probability of that edge so we can normalize
            for lane in turndata:
                normprobs[lane] = dict()
                for nextlane in turndata[lane]:
                    nextedge = nextlane.split("_")[0]
                    if not nextedge in normprobs[lane]:
                        normprobs[lane][nextedge] = 0
                    normprobs[lane][nextedge] += turndata[lane][nextlane]

    #Parse route file to get intended departure times (to account for delayed SUMO insertions due to lack of space)
    #Based on: https://www.geeksforgeeks.org/xml-parsing-python/
    # create element tree object
    tree = ET.parse(routefile)

    # get root element
    root = tree.getroot()

    actualStartDict = dict()
    # iterate news items
    for item in root.findall('./vehicle'):
        actualStartDict[item.attrib["id"]] = float(item.attrib["depart"])
    for item in root.findall('./trip'):
        actualStartDict[item.attrib["id"]] = float(item.attrib["depart"])

    #Do NN setup
    testNN = testNNdefault
    print("testNN="+str(testNN))
    for light in lights:
        if resetTrainingData:
            trainingdata[light] = []

        if testNNdefault:
            if testDumbtrac:
                # agents[light] = Net(26, 1, 32)
                # #agents[light] = Net(2, 1, 32)
                # if FTP:
                agents[light] = Net(182, 1, 64)
            else:
                agents[light] = Net(182, 1, 64)
            optimizers[light] = torch.optim.Adam(agents[light].parameters(), lr=learning_rate)
            MODEL_FILES[light] = 'models/imitate_'+light+'.model' # Once your program successfully trains a network, this file will be written
            print("Checking if there's a learned model. Currently testNN="+str(testNN))
            try:
                agents[light].load(MODEL_FILES[light])
            except FileNotFoundError:
                print("Model doesn't exist - turning off testNN")
                testNN = False
    if not resetTrainingData and appendTrainingData:
        print("LOADING TRAINING DATA, this could take a while")
        try:
            with open("trainingdata/trainingdata_" + sys.argv[1] + ".pickle", 'rb') as handle:
                trainingdata = pickle.load(handle)
        except FileNotFoundError:
            print("Training data not found, starting fresh")
            for light in lights:
                trainingdata[light] = []
    
    outdata = run(network, rerouters, pSmart, verbose)
    
    if appendTrainingData:
        print("Saving training data")
        with open("trainingdata/trainingdata_" + sys.argv[1] + ".pickle", 'wb') as handle:
            pickle.dump(trainingdata, handle, protocol=pickle.HIGHEST_PROTOCOL)
    #traci.close()

    p = pSmart
    newdata = outdata
    try:
        with open("delaydata/delaydata_" + sys.argv[1] + ".pickle", 'rb') as handle:
                data = pickle.load(handle)
    except:
        #If no data found, start fresh
        data = dict()
    if not p in data:
        data[p] = dict()
    for l in ["All", "Adopters", "Non-Adopters", "All2", "Adopters2", "Non-Adopters2", "All3", "Adopters3", "Non-Adopters3", "All0", "Adopters0", "Non-Adopters0", "Runtime", "NTeleports", "TeleportData", "RNGStates"]:
        if not l in data[p]:
            data[p][l] = []

    data[p]["All"].append(newdata[0])
    data[p]["Adopters"].append(newdata[1])
    data[p]["Non-Adopters"].append(newdata[2])
    data[p]["All2"].append(newdata[3])
    data[p]["Adopters2"].append(newdata[4])
    data[p]["Non-Adopters2"].append(newdata[5])
    data[p]["All3"].append(newdata[6])
    data[p]["Adopters3"].append(newdata[7])
    data[p]["Non-Adopters3"].append(newdata[8])
    data[p]["All0"].append(newdata[9])
    data[p]["Adopters0"].append(newdata[10])
    data[p]["Non-Adopters0"].append(newdata[11])
    data[p]["Runtime"].append(newdata[12])
    data[p]["NTeleports"].append(newdata[13])
    data[p]["TeleportData"].append(newdata[14])
    
    data[p]["RNGStates"].append(rngstate)
    with open("delaydata/delaydata_" + sys.argv[1] + ".pickle", 'wb') as handle:
        pickle.dump(data, handle, protocol=pickle.HIGHEST_PROTOCOL)

    return [outdata, rngstate]

#Tell all the detectors to reroute the cars they've seen
#@profile
def reroute(rerouters, network, simtime, remainingDuration, sumoPredClusters=[]):
    global delay3adjdict
    #Clear any stored Surtrac stuff
    global surtracDict
    surtracDict = dict()

    routingthreads = dict()
    routingresults = manager.dict()

    saveStateInfo(savename, remainingDuration, mainlastswitchtimes, sumoPredClusters, lightphases)

    for detector in rerouters:
        ids = traci.inductionloop.getLastStepVehicleIDs(detector) #All vehicles to be rerouted
        if len(ids) == 0:
            #No cars to route, we're done here
            continue #return

        # getRoadID: Returns the edge id the vehicle was last on. Should be the same for all vehicles because it's the same detector
        lane = traci.vehicle.getLaneID(ids[0])
        # print(lane)
        # print(rerouters[detector])
        #assert lane == rerouters[detector] #Verify that we have the correct lane. Except it's bad - sometimes we get lane changes at exactly the wrong time!
        #lane = rerouters[detector] #Because apparently lane changes on top of the detector happen???

        for vehicle in ids:
            #lane = traci.vehicle.getLaneID(vehicle) #Because apparently lane changes on top of the detector happen??? Hopefully there aren't too many vehicles in each timestep - probably only one

            if detector in oldids and vehicle in oldids[detector]:
                #print("Duplicate car " + vehicle + " at detector " + detector)
                continue

            #For delay3, get time at first routing decision point
            if vehicle not in delay3adjdict:
                delay3adjdict[vehicle] = simtime

            #Decide whether we route this vehicle
            if not vehicle in isSmart:
                print("Oops, don't know " + vehicle)
                isSmart[vehicle] = random.random() < pSmart
            if isSmart[vehicle]:
                try:
                    
                    routingresults[vehicle] = manager.list([None, None])

                    if multithreadRouting:
                        #print("Starting vehicle routing thread")
                        routingthreads[vehicle] = Process(target=rerouteSUMOGC, args=(vehicle, lane, remainingDuration, mainlastswitchtimes, sumoPredClusters, lightphases, simtime, network, routingresults))
                        routingthreads[vehicle].start()
                    else:
                        rerouteSUMOGC(vehicle, lane, remainingDuration, mainlastswitchtimes, sumoPredClusters, lightphases, simtime, network, routingresults)
                    
                        if not useLibsumo:
                            assert(traci.getLabel() == "main")
                        else:
                            #(remainingDuration, mainlastswitchtimes, sumoPredClusters, lightphases) = loadStateInfo("MAIN", simtime)
                            loadStateInfo(savename, simtime, network)
                    #assert(traci.getLabel() == "main")

                except traci.exceptions.TraCIException as e:
                    #TODO ghost cars disappearing sometimes? Possibly related to intersection behavior (speed of new spawn?)
                    print("Error in rerouteSUMOGC?? Ignoring")
                    print(e)
                    raise(e)
                    if not useLibsumo:
                        traci.switch("main")

        oldids[detector] = ids

    for vehicle in routingresults:
        if multithreadRouting:
            routingthreads[vehicle].join()
        [newroute, esttime] = routingresults[vehicle]

        routeStats[vehicle]["nCalls"] += 1
        if timedata[vehicle][2] == -1:
            routeStats[vehicle]["nCallsFirst"] += 1
        else:
            routeStats[vehicle]["nCallsAfterFirst"] += 1 #Not necessarily nCalls-1; want to account for vehicles that never got routed
        try:
            if not tuple(newroute) == currentRoutes[vehicle] and not newroute == currentRoutes[vehicle][-len(newroute):]:
                routeStats[vehicle]["nSwaps"] += 1
                routeStats[vehicle]["swapped"] = True
                if timedata[vehicle][2] == -1:
                    routeStats[vehicle]["nSwapsFirst"] += 1
                else:
                    routeStats[vehicle]["nSwapsAfterFirst"] += 1
        except:
            print("Failed route compare")
            print(currentRoutes[vehicle])
            print(newroute)

        timedata[vehicle][0] = simtime #Time prediction was made
        #timedata[vehicle][1] is going to be actual time at goal
        timedata[vehicle][2] = esttime #Predicted time until goal
        timedata[vehicle][3] = currentRoutes[vehicle][0]
        timedata[vehicle][4] = currentRoutes[vehicle][-1]
                
        if not newroute == None:
            try:
                pass
                traci.vehicle.setRoute(vehicle, newroute)
                currentRoutes[vehicle] = newroute
            except traci.exceptions.TraCIException as e:
                print("Couldn't update route, not sure what happened, ignoring")
                print(e)
        else:
            print("newroute == None, likely a problem in routing")        


def backwardDijkstraAStar(network, goal):
    gvals = dict()
    gvals[goal] = 0
    pq = []
    heappush(pq, (0, goal))

    while len(pq) > 0: #When the queue is empty, we're done
        #print(pq)
        stateToExpand = heappop(pq)
        #fval = stateToExpand[0]
        edge = stateToExpand[1]
        gval = gvals[edge]

        #Get predecessor IDs
        succs = []
        for succ in list(network.getEdge(edge).getIncoming()):
            succs.append(succ.getID())
        
        for succ in succs:
            c = traci.lane.getLength(edge+"_0")/network.getEdge(edge).getSpeed()

            h = 0 #Heuristic not needed here - search is fast
            if succ in gvals and gvals[succ] <= gval+c:
                #Already saw this state, don't requeue
                continue

            #Otherwise it's new or we're now doing better, so requeue it
            gvals[succ] = gval+c
            heappush(pq, (gval+c+h, succ))
    return gvals

def getDTheta(startedge, nextedge, network):
    #Preprocess in case we accidentally pass in lanes instead of edges
    startedge = startedge.split("_")[0]
    nextedge = nextedge.split("_")[0]

    c0 = network.getEdge(startedge).getFromNode().getCoord()
    c1 = network.getEdge(startedge).getToNode().getCoord()
    theta0 = math.atan2(c1[1]-c0[1], c1[0]-c0[0])

    c2 = network.getEdge(nextedge).getToNode().getCoord()
    theta1 = math.atan2(c2[1]-c1[1], c2[0]-c1[0])

    dtheta = (theta1-theta0+math.pi)%(2*math.pi)-math.pi
    return dtheta

def getLeftEdge(startlane, network):
    #Returns the leftmost edge with a connection from startlane
    startedge = startlane.split("_")[0]
    leftedge = None
    maxleft = -np.inf
    for nextlinktuple in links[startlane]:
        #Add to all lanes on that edge
        nextedge = nextlinktuple[0].split("_")[0]

        dtheta = getDTheta(startedge, nextedge, network)
        if dtheta > maxleft:
            leftedge = nextedge
            maxleft = dtheta
    return leftedge

def prepGhostCars(VOIs, id, ghostcarlanes, network, spawnLeft, ghostcardata, simtime):
    
    if not spawnLeft:
        defaultghostcarpos = 5
    else:
        defaultghostcarpos = 5
    ghostcarpos = defaultghostcarpos

    lane = VOIs[id][0]
    oldpos = VOIs[id][1]
    oldspeed = VOIs[id][2]
    oldroute = VOIs[id][3]

    for nextlinktuple in links[lane]:
        #Add to all lanes on that edge
        nextedge = nextlinktuple[0].split("_")[0]
        if nextedge == VOIs[id][4] and not spawnLeft:
            #Don't spawn left-turn copies yet!
            continue
        if nextedge != VOIs[id][4] and spawnLeft:
            #Only spawn left-turn copies
            continue
        for nextlanenum in range(nLanes[nextedge]):
            nextlane = nextedge+"_"+str(nextlanenum)
            if not nextlane in ghostcarlanes:
                
                newghostcar = None
                
                if newghostcar == None:
                    newghostcar = id+"|"+nextlane

                    if spawnLeft:
                        #Whatever the leftmost road is (could conceivably be straight at a 3-way intersection)
                        leftdelay = 0
                        if not simtime + leftdelay in ghostcardata:
                            ghostcardata[simtime+leftdelay] = []
                        newspeed = min(13, oldspeed)
                        ghostcardata[simtime+leftdelay].append([newghostcar, nextlane, newspeed, ghostcarpos, oldroute])
                    elif abs(getDTheta(lane.split("_")[0], nextlane.split("_")[0], network)) < 0.1: #Straight
                        straightdelay = 3
                        if not simtime + straightdelay in ghostcardata:
                            ghostcardata[simtime+straightdelay] = []
                        newspeed = min(13, oldspeed)
                        ghostcardata[simtime+straightdelay].append([newghostcar, nextlane, newspeed, ghostcarpos, oldroute])
                    else: #Right turn or weird intersection
                        rightdelay = 5
                        if not simtime + rightdelay in ghostcardata:
                            ghostcardata[simtime+rightdelay] = []
                        newspeed = min(10, oldspeed)
                        ghostcardata[simtime+rightdelay].append([newghostcar, nextlane, newspeed, ghostcarpos, oldroute])

def spawnGhostCars(ghostcardata, ghostcarlanes, simtime, network, VOIs, laneDict2, edgeDict2, nonExitEdgeDetections):
    carcardist = 15 #TODO: Don't hard-code this in two different places in code!!! (Actually, maybe fine, other one might need a gap on both sides?)
    replaceExistingCar = False #TODO would turning this on help??
    touchNothing = False #WARNING: Setting this to True triggers the missing VOI sanity check, likely due to VOIs not actually being inserted on time. That said, I don't actually understand why, given it should still add to the VOIs list, unless it starts teleporting to its goal immediately or something weird

    if not simtime in ghostcardata:
        return
    for gcl in ghostcardata[simtime]:
        [newghostcar, nextlane, newspeed, ghostcarpos, oldroute] = gcl
        if not nextlane in ghostcarlanes:
            replacedCar = False
            ghostcarlanes.append(nextlane)

            if not touchNothing:
                #Actually add the new ghost car
                for tempveh in traci.lane.getLastStepVehicleIDs(nextlane):
                    lanepos = traci.vehicle.getLanePosition(tempveh)
                    if lanepos <= ghostcarpos+carcardist:
                        if replaceExistingCar:
                            #Convert the last car on the lane within carcardist of the start, if such a car exists
                            newghostcar = tempveh
                            oldspeed = traci.vehicle.getSpeed(newghostcar)
                            ghostcarpos = traci.vehicle.getLanePosition(newghostcar)
                            traci.vehicle.setRoute(newghostcar, [nextedge])
                            replacedCar = True
                            break
                        else:
                            traci.vehicle.remove(tempveh) #Only need to remove one; if there was space for it, there's space for the ghost car
                            if tempveh in edgeDict2: #Libsumo errors here sometimes with detectorModel off
                                laneDict2.pop(tempveh)
                                edgeDict2.pop(tempveh)
                            if nextlane.split("_")[0] in nonExitEdgeDetections and len(nonExitEdgeDetections[nextlane.split("_")[0]][0]) > 0:
                                nonExitEdgeDetections[nextlane.split("_")[0]][0].pop(-1) #Ghost cars spawn at edge start, so eat the newest detection, then remake it
                                #If we deleted the wrong car, rename the "correct" version of the thing we just deleted
                                for vehicletupleind in range(len(nonExitEdgeDetections[nextlane.split("_")[0]][0])):
                                    vehicletuple = nonExitEdgeDetections[nextlane.split("_")[0]][0][vehicletupleind]
                                    if vehicletuple[0] == tempveh:
                                        nonExitEdgeDetections[nextlane.split("_")[0]][0][vehicletupleind] = (nextlane.split("_")[0]+".0oldsmartcarghostcarreplace."+str(simtime), vehicletuple[1], vehicletuple[2]) #This seems to alias as intended
                            break
                    if lanepos > ghostcarpos+carcardist:
                        #Passed the insertion point, no need to keep checking cars
                        break

            nextedge = nextlane.split("_")[0]
            if not nextlane in traci.route.getIDList():
                traci.route.add(nextlane, [nextedge])
            if not replacedCar:
                traci.vehicle.add(newghostcar, nextlane, departLane=int(nextlane.split("_")[-1]), departSpeed="max")
                #traci.vehicle.add(newghostcar, nextlane, departLane=int(nextlane.split("_")[-1]), departSpeed=max(0, min(newspeed, 0.99*network.getEdge(nextedge).getSpeed())))
                #There should be a departPos argument, but somehow it takes a string? And probably tries to insert at or behind the pos, making VOIs disappear if no space if I don't explicitly call moveTo
                if not touchNothing:
                    traci.vehicle.moveTo(newghostcar, nextlane, ghostcarpos)
                laneDict2[newghostcar] = nextlane
                edgeDict2[newghostcar] = nextlane.split("_")[0]
                if nextlane.split("_")[0] in nonExitEdgeDetections:
                    nonExitEdgeDetections[nextlane.split("_")[0]][0].append((newghostcar+"."+str(simtime), nextlane, simtime))
            traci.vehicle.setSpeedFactor(newghostcar, 1)
            traci.vehicle.setColor(newghostcar, [0, 255, 255, 255])
            leftedge = getLeftEdge(nextlane, network)
            traci.vehicle.setRoute(newghostcar, [nextedge])
            if not leftedge == None:
                traci.vehicle.setRoute(newghostcar, [nextedge, leftedge])
            VOIs[newghostcar] = [nextlane, newspeed, ghostcarpos, oldroute+[nextedge], leftedge, True]

#@profile
def rerouteSUMOGC(startvehicle, startlane, remainingDurationIn, mainlastswitchtimes, sumoPredClusters, lightphases, simtime, network, reroutedata):
    global nRoutingCalls
    global nSuccessfulRoutingCalls
    global routingTime
    global surtracDict
    global nonExitEdgeDetections

    remainingDuration = pickle.loads(pickle.dumps(remainingDurationIn)) #This is apparently important, not sure why. It's especially weird given the next time we see remainingDuration is as the output of a loadClusters call

    nRoutingCalls += 1
    vehicle = startvehicle
    routestartwctime = time.time() #For timeouts and maybe stats
    timeout = 60

    ghostcardata = dict()

    startedge = startlane.split("_")[0]
    VOIs = dict()
    #VOIs[vehicle] stores current lane, current speed, current position, route to now, left turn edge (if any), and whether we still need to spawn non-left copies
    VOIs[vehicle] = [startlane, traci.vehicle.getSpeed(vehicle), traci.vehicle.getLanePosition(vehicle), [startedge], getLeftEdge(startlane, network), True]
    ghostcarlanes = []

    #New plan: We're doing ghost cars in SUMO
    #For starting split to all lanes, we'll convert the first car in the way, or make a new car if none exist
    #When a VOI reaches the end of a road, we'll insert new ghost cars on all outgoing lanes
    #Unless it reached the end of the goal road, in which case great, we're done
    #Hopefully those new inserts take priority over standard cars and it all works?
    #assert(traci.getLabel() == "main")

    #Get goal
    startroute = traci.vehicle.getRoute(vehicle)
    startind = startroute.index(startedge)
    startroute = startroute[startind:]
    goaledge = startroute[-1]

    if startedge == goaledge:
        #No rerouting needed, we're basically done
        ##traci.switch("main") #Not needed, didn't even switch to test yet
        routingTime += time.time() - routestartwctime
        #assert(traci.getLabel() == "main")
        reroutedata[startvehicle] = [startroute, -1]
        return reroutedata[startvehicle]

    if useLibsumo:
        pass
        #Apparently the new thread just comes with a copy of the old simulation - don't need to do this at all?
        # traci.start([checkBinary('sumo'), "-c", sumoconfig,
        #                         "--additional-files", "additional_autogen.xml",
        #                         "--log", "LOGFILE", "--xml-validation", "never", "--start", "--quit-on-end"])
        (remainingDuration, lastSwitchTimes, sumoPredClusters, testSUMOlightphases, edgeDict3, laneDict3) = loadStateInfo(savename, simtime, network)
    else:
        saveStateInfo(savename, remainingDuration, mainlastswitchtimes, sumoPredClusters, lightphases) #Saves the traffic state and traffic light timings #TODO pretty sure I do this at the start of reroute - at some point, make sure nothing breaks if I comment this
        traci.switch("test")

        (remainingDuration, lastSwitchTimes, sumoPredClusters, testSUMOlightphases, edgeDict3, laneDict3) = loadStateInfo(savename, simtime, network)
    
    #assert(traci.vehicle.getRoadID(vehicle) == startedge) #This fails with loadStateInfoDetectors; apparently getRoadID returns an empty string. Does adding vehicles not register until the next timestep? (But loading the save the normal way does?)

    #Load old detector readings
    nonExitEdgeDetections2 = deepcopy(nonExitEdgeDetections) #So we hopefully don't break anything in main sim
    
    #Tell the VOI to drive to the end of edge
    try:
        traci.vehicle.setColor(vehicle, [0, 255, 255, 255]) #Make the ghost car white for debug purposes
        #Tell ghost car to turn left
        leftedge = getLeftEdge(startlane, network)
        traci.vehicle.setRoute(vehicle, [startedge])
        if not leftedge == None:
            traci.vehicle.setRoute(vehicle, [startedge, leftedge])
        ghostcarlanes.append(startlane)
    except Exception as e:
        print("Something went wrong - couldn't set vehicle route.")
        print(e)
        if not useLibsumo:
            traci.switch("main")
        routingTime += time.time() - routestartwctime
        reroutedata[startvehicle] = [startroute, -1]
        return reroutedata[startvehicle] #Which we'll parse as "oops, stop"

    for lanenum in range(nLanes[startedge]):
        if lanenum == int(startlane.split("_")[-1]):
            #Skip the starting lane
            continue
        nextlane = startedge+"_"+str(lanenum)
        newghostcar = None
        #We're now going to add a new ghost car to this lane, replacing any existing car in the way
        carcardist = 25 #TODO check this to make sure we're not being overly careful here
        replaceExistingCar = False
        for tempveh in reversed(traci.lane.getLastStepVehicleIDs(nextlane)): #Reversed so we go from end of edge to start of edge - first clusters to leave are listed first
            #Process vehicle into cluster somehow
            #If nearby cluster, add to cluster in sorted order (could probably process in sorted order)
            lanepos = traci.vehicle.getLanePosition(tempveh)
            if lanepos <= VOIs[startvehicle][2] + 0.5*carcardist and lanepos > VOIs[startvehicle][2] - 0.5*carcardist:
                #Convert the first car in a window centered on the first vehicle of size carcardist, if such a car exists
                if replaceExistingCar:
                    newghostcar = tempveh
                    newspeed = traci.vehicle.getSpeed(newghostcar)
                    break
                else:
                    traci.vehicle.remove(tempveh)
            if lanepos <= VOIs[startvehicle][2] - carcardist:
                #Passed the insertion point, no need to keep checking cars
                break
        #If no existing car can be replaced, make a new one
        if newghostcar == None:
            newghostcar = vehicle+"newghostcar"+nextlane #Hopefully this name isn't taken...
            newspeed = VOIs[startvehicle][1]
            if not nextlane in traci.route.getIDList():
                traci.route.add(nextlane, [nextlane.split("_")[0]])
            traci.vehicle.add(newghostcar, nextlane, departLane=lanenum, departSpeed="max")#min(newspeed, 0.99*network.getEdge(startedge).getSpeed()))
            edgeDict3[newghostcar] = nextlane.split("_")[0]
            laneDict3[newghostcar] = nextlane
            #Not going to insert a detector reading for these, hopefully it's fine
            traci.vehicle.moveTo(newghostcar, nextlane, VOIs[startvehicle][2])
            #traci.vehicle.setSpeed(newghostcar, newspeed)
            
        #Regardless of whether we had to make a new ghost car or could convert an existing one, make it look like a ghost car
        traci.vehicle.setColor(newghostcar, [0, 255, 255, 255]) #Make the ghost car white for debug purposes
        #Tell new ghost car to turn left
        leftedge = getLeftEdge(nextlane, network)
        traci.vehicle.setRoute(newghostcar, [startedge])
        if not leftedge == None:
            traci.vehicle.setRoute(newghostcar, [startedge, leftedge])
        VOIs[newghostcar] = [nextlane, newspeed, traci.vehicle.getLanePosition(newghostcar), [startedge], leftedge, True]
        ghostcarlanes.append(nextlane)
    
    starttime = simtime

    #Generate excess queue that we think should be off the network, based on recent arrival rate being smaller than average
    for nextlane in arrivals:
        nextedge = nextlane.split("_")[0]
        if len(arrivals[nextlane]) == 0:
            #No recent arrivals, nothing to add
            continue
        avgarrivalrate = len(arrivals[nextlane])/min(starttime, maxarrivalwindow)
        recentarrivalrate = len(arrivals2[nextlane])/min(starttime, maxarrivalwindow2)

        if recentarrivalrate < avgarrivalrate:
            nToAdd = math.floor((avgarrivalrate - recentarrivalrate)*maxarrivalwindow2) #Don't need fancy min(starttime, ...) here; if we're near start of sim, these should be equal so this won't trigger
            for i in range(nToAdd):

                newvehicle = nextlane+"precar"+str(i)
                #Need a temp route so we don't immediately error, though we'll overwrite the route immediately afterwards anyway
                nextedge = nextlane.split("_")[0]
                if not nextlane in traci.route.getIDList():
                    traci.route.add(nextlane, [nextedge])
                traci.vehicle.add(newvehicle, nextlane) #No fancy args here, should be fine

                temproute = sampleRouteFromTurnData(nextlane, turndata)
                try:
                    traci.vehicle.setRoute(newvehicle, temproute)
                except Exception as e:
                    print("Failing to set initial route of predicted vehicle while creating predicted queue")
                    print(temproute)
                    print(newvehicle)
                    raise(e)

                #Pretend to be a detector at the start of the input lane
                #Hopefully these spawn off the network and we'll catch them as they enter
                laneDict3[newvehicle] = "off network"
                edgeDict3[newvehicle] = "off network"

    #START ROUTING SIM MAIN LOOP
    #Run simulation, track time to completion
    while(True):
        #Timeout if things have gone wrong somehow
        if time.time()-routestartwctime > timeout:
            print("Routing timeout: Edge " + startedge + ", time: " + str(starttime))
            routeStats[startvehicle]["nTimeouts"] += 1
            
            if not useLibsumo:
                traci.switch("main")
            routingTime += time.time() - routestartwctime
            reroutedata[startvehicle] = [startroute, -1]
            return reroutedata[startvehicle]

        traci.simulationStep()
        simtime+=1
        #print(VOIs)

        #Remove all future vehicles since we aren't supposed to know about them
        for id in traci.simulation.getDepartedIDList():
            if not id in edgeDict3:#isSmart: #We don't know if they're smart since they haven't shown up in main sim yet and aren't supposed to be here
                #NEXT TODO: isSmart check is super bad and deleting the vehicles I started with!
                try:
                    traci.vehicle.remove(id)
                    #These vehicles then apparently get transferred to arrived list, and then we get errors for trying to delete them from edgeDict3, so let's make temp entries to fix that
                    # edgeDict3[id] = "Removed on entry"
                    # laneDict3[id] = "Removed on entry"
                    # edgeDict3.pop(id)
                    # laneDict3.pop(id)
                except Exception as e:
                    print("things going wrong when removing vehicles")
                    print(id)
                    print(e)

        #Remove arrived vehicles from dicts so we don't try to update them with a detector model
        for vehicle in traci.simulation.getArrivedIDList():
            if vehicle in edgeDict3: #Without this check, we occasionally get errors without detector model, either with known vehicles trying to enter that we immediately remove, or apparently other stuff with libsumo???
                #In case the target was a non-exit edge, make sure to remove these from detector readings
                if edgeDict3[vehicle] in nonExitEdgeDetections2:
                    #Delete from old edge detector readings
                    oldEdgeStuff = nonExitEdgeDetections2[edgeDict3[vehicle]][0] #Since we're only storing stuff in index 0 anyway
                    if len(oldEdgeStuff) > 0:
                        oldEdgeStuff.pop(0) #Pop oldest from old road, don't care from which lane
                    else:
                        print("Warning: Ran out of cars to remove on edge " + edgeDict3[vehicle] + "!!!!!!!!!!!!!!!!!")

                    #If we deleted the wrong car, rename the "correct" version of the thing we just deleted
                    if edgeDict3[vehicle] in nonExitEdgeDetections2:
                        for vehicletupleind in range(len(nonExitEdgeDetections2[edgeDict3[vehicle]][0])):
                            vehicletuple = nonExitEdgeDetections2[edgeDict3[vehicle]][0][vehicletupleind]
                            if vehicletuple[0] == vehicle:
                                nonExitEdgeDetections2[edgeDict3[vehicle]][0][vehicletupleind] = (edgeDict3[vehicle]+".0oldsmartcar."+str(simtime), vehicletuple[1], vehicletuple[2]) #This seems to alias as intended
                
                edgeDict3.pop(vehicle)
                laneDict3.pop(vehicle)

        #Update detections for existing cars
        for id in laneDict3:
            try:
                newlane = traci.vehicle.getLaneID(id)
            except Exception as e:
                print("Test getLaneID fail")
                print(id)
                print(laneDict3[id])
                print(VOIs)
                raise(e)
            if newlane != laneDict3[id] and len(newlane) > 0 and  newlane[0] != ":":
                newloc = traci.vehicle.getRoadID(id)

                #Pretend to be detectors at the start of each road (need to know where we came from so we can steal from the correct previous lane)
                if newloc != edgeDict3[id]: #Moved to a new road
                    if edgeDict3[id] in nonExitEdgeDetections2:
                        #Delete from old edge detector readings
                        oldEdgeStuff = nonExitEdgeDetections2[edgeDict3[id]][0] #Since we're only storing stuff in index 0 anyway
                        if len(oldEdgeStuff) > 0:
                            oldEdgeStuff.pop(0) #Pop oldest from old road, don't care from which lane
                            #TODO what if this deletes an adopter? Actual adopter shows up later and fixes itself?
                        else:
                            print("Warning: Ran out of cars to remove on edge " + edgeDict3[id] + "!!!!!!!!!!!!!!!!!")

                        #If we didn't delete the "correct" vehicle, rename the non-deleted copy to some generic non-adopter
                        if edgeDict3[id] in nonExitEdgeDetections2:
                            for vehicletupleind in range(len(nonExitEdgeDetections2[edgeDict3[id]][0])):
                                vehicletuple = nonExitEdgeDetections2[edgeDict3[id]][0][vehicletupleind]
                                if vehicletuple[0] == id:
                                    nonExitEdgeDetections2[edgeDict3[id]][0][vehicletupleind] = (edgeDict3[id]+".0oldsmartcar."+str(simtime), vehicletuple[1], vehicletuple[2]) #This seems to alias as intended
                
                    #Add to new edge detector readings
                    if newloc in nonExitEdgeDetections2:
                        assert(newlane.split("_")[0] == newloc)
                        if id in VOIs or (id in isSmart and isSmart[id]):
                            nonExitEdgeDetections2[newloc][0].append((id, newlane, simtime))
                        else:
                            nonExitEdgeDetections2[newloc][0].append((newlane+".0routingdetect."+str(simtime), newlane, simtime))

                laneDict3[id] = newlane
                edgeDict3[id] = newloc
                #TODO main run() function clears Surtrac predictions at this point; I'll need to do that if I get this thing to share routes

        #Add new cars
        for nextlane in arrivals:
            nextedge = nextlane.split("_")[0]
            if len(arrivals[nextlane]) == 0:
                #No recent arrivals, nothing to add
                continue
            timeperarrival = min(starttime, maxarrivalwindow)/len(arrivals[nextlane])
            if timeperarrival <= timestep or simtime%timeperarrival >= (simtime+timestep)%timeperarrival:
                #Add a car
                newvehicle = nextlane+"predcar"+str(simtime)
                #Need a temp route so we don't immediately error, though we'll overwrite the route immediately afterwards anyway
                nextedge = nextlane.split("_")[0]
                if not nextlane in traci.route.getIDList():
                    traci.route.add(nextlane, [nextedge])
                traci.vehicle.add(newvehicle, nextlane) #No fancy args here, should be fine
                isSmart[newvehicle] = False

                temproute = sampleRouteFromTurnData(nextlane, turndata)
                try:
                    traci.vehicle.setRoute(newvehicle, temproute)
                except Exception as e:
                    print("Failing to set initial route of predicted vehicle")
                    print(temproute)
                    print(newvehicle)
                    raise(e)

                #Update location dicts and detector readings
                laneDict3[newvehicle] = nextlane
                edgeDict3[newvehicle] = nextedge
                if nextedge in nonExitEdgeDetections2:
                    nonExitEdgeDetections2[nextedge][0].append((nextlane+".0routingnewcar."+str(simtime), nextlane, simtime))

        #Sanity check for all VOIs deleted. (NOTE: This assumes VOIs list works...)
        if len(VOIs) == 0 and max(ghostcardata.keys()) < simtime:
            print("AAAAAAAAAAAH! All VOIs disappeared! No clue why! AAAAAAAAAAAAH!")
            print(ghostcarlanes)
            print(startvehicle)
            print(startedge)
            print(goaledge)
            OHNOADOPTERSWENTPOOF #Should throw an error
            assert(False) #In case it doesn't somehow

        #Check if VOIs got teleported. If so, problem, abort
        for id in traci.simulation.getStartingTeleportIDList():
            if id in VOIs:
                print("VOI started to teleport, simulation results unreliable, giving up")
                if not useLibsumo:
                    traci.switch("main")
                routingTime += time.time() - routestartwctime
                reroutedata[startvehicle] = [startroute, -1]
                return reroutedata[startvehicle]

        temp = traci.simulation.getArrivedIDList()
        toDelete = []
        tempVOIs = copy(VOIs) #Else Python gets mad about our dictionary changing size when we add new ghost cars
        for id in tempVOIs:
            if not id in temp:
                #VOIs[id] stores current lane, current speed, current position, route to now, left turn edge (if any), and whether we still need to spawn non-left copies
                if traci.vehicle.getLaneID(id) != VOIs[id][0]:
                    #If we've successfully started turning left out of the goal edge, we're actually done and should note that
                    if VOIs[id][0].split("_")[0] == goaledge:
                        if not useLibsumo:
                            traci.switch("main")
                        nSuccessfulRoutingCalls += 1
                        routingTime += time.time() - routestartwctime
                        reroutedata[startvehicle] = [VOIs[id][3], simtime - starttime]
                        return reroutedata[startvehicle]

                    #If we still need to spawn non-left copies (presumably we're in the intersection), do that
                    if VOIs[id][5]:
                        VOIs[id][5] = False #Don't repeatedly spawn these - ghostcarlanes should sort this, though
                        prepGhostCars(VOIs, id, ghostcarlanes, network, False, ghostcardata, simtime)

                    #If we made it out of the intersection and successfully turned left, delete yourself and spawn replacements
                    #VOIs[id] stores current lane, current speed, current position, route to now, left turn edge (if any), and whether we still need to spawn non-left copies
                    if traci.vehicle.getRoadID(id) == VOIs[id][4]:
                        toDelete.append(id)
                        traci.vehicle.remove(id) #This VOI has successfully turned left. Remove it before spawning its replacements (in all lanes)
                        
                        #Remove it from detections
                        #Consider doing the delete-and-rename thing from before? But we're in a routing sim, we can assume perfect information. So this might be better
                        #Pretty sure this isn't redundant - we moved from old edge to new edge during standard car stuff, but now need to delete from new edge
                        if edgeDict3[id] in nonExitEdgeDetections2:
                            vehicletupleind = 0
                            oldEdgeStuff = nonExitEdgeDetections2[edgeDict3[id]][0] #Since we're only storing stuff in index 0 anyway
                            while vehicletupleind < len(oldEdgeStuff):
                                if oldEdgeStuff[vehicletupleind][0] == id:
                                    oldEdgeStuff.pop(vehicletupleind)
                                else:
                                    vehicletupleind += 1
                        laneDict3.pop(id)
                        edgeDict3.pop(id)

                        prepGhostCars(VOIs, id, ghostcarlanes, network, True, ghostcardata, simtime)    

                else:
                    VOIs[id][1] = traci.vehicle.getLanePosition(id)
                    VOIs[id][2] = traci.vehicle.getSpeed(id)

        #TODO: We handle removing arrived vehicles from edgeDict3 and laneDict3 above, should maybe group this with that. But make sure we don't break stuff first
        for id in temp: #temp is the getArrivedList, so anything that left the network
            if id in VOIs:
                #If we've successfully exited the goal edge, we're done
                if VOIs[id][0].split("_")[0] == goaledge:
                    if not useLibsumo:
                        traci.switch("main")
                    nSuccessfulRoutingCalls += 1
                    routingTime += time.time() - routestartwctime
                    reroutedata[startvehicle] = [VOIs[id][3], simtime - starttime]
                    return reroutedata[startvehicle]

                toDelete.append(id)
                #TODO: Why are we prepping new ghost cars? I'd think there's just nothing to spawn? (Might've been from back when we said the leftmost road had to actually be on the left)
                #prepGhostCars(VOIs, id, ghostcarlanes, network, False, ghostcardata, simtime) #NOTE: Since the vehicle left the network, there must be no left edge, so no need to spawn left turn cars
        
        for id in toDelete:
            VOIs.pop(id)

        spawnGhostCars(ghostcardata, ghostcarlanes, simtime, network, VOIs, laneDict3, edgeDict3, nonExitEdgeDetections2)

        #Light logic for Surtrac, etc.

        surtracFreq = routingSurtracFreq #Period between updates
        if simtime%surtracFreq >= (simtime+1)%surtracFreq:
            
            updateLights = True
            if not(reuseSurtrac and simtime in surtracDict): #Overwrite unless we want to reuse
                updateLights = False
                surtracDict[simtime] = doSurtrac(network, simtime, None, testSUMOlightphases, lastSwitchTimes, sumoPredClusters, True, nonExitEdgeDetections2)
            # else:
            #     print("Reusing Surtrac, yay!")

            temp = pickle.loads(pickle.dumps(surtracDict[simtime]))

            #Don't bother storing toUpdate = temp[0], since doSurtrac has done that update already
            sumoPredClusters = temp[1]
            remainingDuration.update(temp[2])

            if updateLights:
                #We got to reuse stuff, but we still need to update the lights
                toUpdate = temp[0]
                for light in toUpdate:
                    curphase = lightphases[light]
                    nPhases = len(surtracdata[light]) #Number of phases

                    traci.trafficlight.setPhase(light, (curphase+1)%nPhases) #Increment phase, duration defaults to default
                for light in lights:
                    if len(remainingDuration[light]) > 0:
                        #And set the new duration if possible
                        traci.trafficlight.setPhaseDuration(light, remainingDuration[light][0]) #Update duration if we know it

        #Check for lights that switched phase (because previously-planned duration ran out, not because Surtrac etc. changed the plan); update custom data structures and current phase duration
        for light in lights:
            temp = traci.trafficlight.getPhase(light)
            if not(light in remainingDuration and len(remainingDuration[light]) > 0):
                #Only update remainingDuration if we have no schedule, in which case grab the actual remaining duration from SUMO
                remainingDuration[light] = [traci.trafficlight.getNextSwitch(light) - simtime]
            else:
                remainingDuration[light][0] -= 1
            if temp != testSUMOlightphases[light]:
                lastSwitchTimes[light] = simtime
                testSUMOlightphases[light] = temp
                #Duration of previous phase was first element of remainingDuration, so pop that and read the next, assuming everything exists
                if light in remainingDuration and len(remainingDuration[light]) > 0:
                    #print(remainingDuration[light][0]) #Prints -1. Might be an off-by-one somewhere, but should be pretty close to accurate?
                    #NOTE: The light switches when remaining duration goes negative (in this case -1)
                    remainingDuration[light].pop(0)
                    if len(remainingDuration[light]) == 0:
                        remainingDuration[light] = [traci.trafficlight.getNextSwitch(light) - simtime]
                    else:
                        traci.trafficlight.setPhaseDuration(light, remainingDuration[light][0])
                else:
                    print("Unrecognized light " + light + ", this shouldn't happen")


# Gets successor edges of a given edge in a given network
# Parameters:
#   edge: an edge ID string
#   network: the network object from sumolib.net.readNet(netfile)
# Returns:
#   successors: a list of edge IDs for the successor edges (outgoing edges from the next intersection)
def getSuccessors(edge, network):
    ids = []
    for succ in list(network.getEdge(edge).getOutgoing()):
        ids.append(succ.getID())
    return ids

def saveStateInfo(edge, remainingDuration, lastSwitchTimes, sumoPredClusters, lightphases):
    #Copy state from main sim to test sim
    traci.simulation.saveState("savestates/teststate_"+edge+".xml")
    #saveState apparently doesn't save traffic light states despite what the docs say
    #So save all the traffic light states and copy them over
    lightStates = dict()
    for light in traci.trafficlight.getIDList():
        lightStates[light] = [traci.trafficlight.getPhase(light), traci.trafficlight.getPhaseDuration(light)]
        #Why do the built-in functions have such terrible names?!
        lightStates[light][1] = traci.trafficlight.getNextSwitch(light) - traci.simulation.getTime()
    #Save lightStates to a file
    with open("savestates/lightstate_"+edge+".pickle", 'wb') as handle:
        pickle.dump(lightStates, handle, protocol=pickle.HIGHEST_PROTOCOL)
    with open("savestates/lightstate_"+edge+"_remainingDuration.pickle", 'wb') as handle:
        pickle.dump(remainingDuration, handle, protocol=pickle.HIGHEST_PROTOCOL)
    with open("savestates/lightstate_"+edge+"_lastSwitchTimes.pickle", 'wb') as handle:
        pickle.dump(lastSwitchTimes, handle, protocol=pickle.HIGHEST_PROTOCOL)
    with open("savestates/lightstate_"+edge+"_sumoPredClusters.pickle", 'wb') as handle:
        pickle.dump(sumoPredClusters, handle, protocol=pickle.HIGHEST_PROTOCOL)
    with open("savestates/lightstate_"+edge+"_lightphases.pickle", 'wb') as handle:
        pickle.dump(lightphases, handle, protocol=pickle.HIGHEST_PROTOCOL)

#prevedge is just used as part of the filename - can pass in a constant string so we overwrite, or something like a timestamp to support multiple instances of the code running at once
def loadStateInfo(prevedge, simtime, network): #simtime is just so I can pass it into loadStateInfoDetectors...
    if detectorRouting:
        return loadStateInfoDetectors(prevedge, simtime, network)

    #Load traffic state
    traci.simulation.loadState("savestates/teststate_"+prevedge+".xml")
    #Load light state
    with open("savestates/lightstate_"+prevedge+".pickle", 'rb') as handle:
        lightStates = pickle.load(handle)

    #Randomize non-adopter routes
    for lane in lanes:
        if len(lane) == 0 or lane[0] == ":":
            continue
        for vehicle in traci.lane.getLastStepVehicleIDs(lane):
            if not vehicle in isSmart or isSmart[vehicle] == False:
                traci.vehicle.setRoute(vehicle, sampleRouteFromTurnData(lane, turndata))

    #Copy traffic light timings
    for light in traci.trafficlight.getIDList():
        traci.trafficlight.setPhase(light, lightStates[light][0])
        traci.trafficlight.setPhaseDuration(light, lightStates[light][1])
        #print(lightStates[light][1])
    with open("savestates/lightstate_"+prevedge+"_remainingDuration.pickle", 'rb') as handle:
        remainingDuration = pickle.load(handle)
    with open("savestates/lightstate_"+prevedge+"_lastSwitchTimes.pickle", 'rb') as handle:
        lastSwitchTimes = pickle.load(handle)
    with open("savestates/lightstate_"+prevedge+"_sumoPredClusters.pickle", 'rb') as handle:
        sumoPredClusters = pickle.load(handle)
    with open("savestates/lightstate_"+prevedge+"_lightphases.pickle", 'rb') as handle:
        lightphases = pickle.load(handle)
    return (remainingDuration, lastSwitchTimes, sumoPredClusters, lightphases, deepcopy(edgeDict), deepcopy(laneDict))

#prevedge is just used as part of the filename - can pass in a constant string so we overwrite, or something like a timestamp to support multiple instances of the code running at once
def loadStateInfoDetectors(prevedge, simtime, network):
    global netfile

    newEdgeDict = dict()
    newLaneDict = dict()
    
    #Purge all vehicles
    traci.load(["-n", netfile,
                                "--additional-files", "additional_autogen.xml", "--no-step-log", "true",
                                "--log", "LOGFILE", "--xml-validation", "never", "--start", "--quit-on-end"])

    totallanedata = dict()
    for edge in nonExitEdgeDetections: #Assuming exit lanes don't matter since they shouldn't have traffic - this saves us from extra exit detectors at their ends
        totallanedata[edge] = 0
        for laneind in range(nLanes[edge]):
            lane = edge + "_" + str(laneind)
            totallanedata[edge] += len(wasFull[nonExitLaneDetectors[lane][1][0]]) + 1 #[lane][1][0] because 1 is the index of the exit detector and 0 is the index of its name. +1 as a permanent psuedocount on all detectors, mostly in case we have no data whatsoever

    #Read vehicles according to detector model
    for superedge in nonExitEdgeDetections: #Assuming exit lanes don't matter since they shouldn't have traffic - this saves us from extra exit detectors at their ends
        temp = nonExitEdgeDetections[superedge][0]
        
        for vehicletuple in temp: #Not reversed since the whole road is just one giant block now - oldest cars are furthest along and listed first
            (vehicle, templane, detecttime) = vehicletuple
            edge = templane.split("_")[0]

            if vehicle in isSmart and isSmart[vehicle] and adopterCommsRouting:
                if not vehicle in adopterinfo or superedge != adopterinfo[vehicle][0].split("_")[0]:
                    #This isn't actually the adopter's current location, replace the name with something else and continue as if it's a non-adopter
                    print("Warning: Adopter " + vehicle + " detections think it's in the wrong spot, replacing this adopter with a non-adopter")
                    vehicle = vehicle + ".notadopter." + superedge + "." + str(simtime) #Need to replace the name so we don't try to grab an invalid lane number or something later.
                    assert(vehicletuple[0] != vehicle)
                    pass #We're actually tracking adopter positions, don't change the name or anything

            #Sample a lane randomly
            if not vehicle in isSmart or not isSmart[vehicle] or not adopterCommsRouting:
                r = random.random()
                for laneind in range(nLanes[edge]):
                    lane = edge + "_" + str(laneind)
                    r -= (len(wasFull[nonExitLaneDetectors[lane][1][0]]) + 1)/totallanedata[edge]
                    if r < 0:
                        break #lane is now equal to a lane sampled from the lane change probabilities data from wasFull


                #Process vehicle into cluster somehow
                #If nearby cluster, add to cluster in sorted order (could probably process in sorted order)
                lanepos = min(lengths[lane], speeds[lane.split("_")[0]] * (simtime - detecttime + 0.5)+simdetectordist) #+0.5 because we crossed the detector, then made somewhere between 0 and 1 seconds worth of forward movement; estimate it as 0.5
                #lanepos = traci.vehicle.getLanePosition(vehicle)

                #Because apparently traci.vehicle.add needs a route stored in TraCI with a name, not just a list of edges. Why?!
                newroute = sampleRouteFromTurnData(lane, turndata)

            else:
                try:
                    lane = adopterinfo[vehicle][0]
                    if not lane.split("_")[0] == superedge:
                        print("Error: Adopter is apparently on the wrong road, but this should've been fixed earlier")
                    lanepos = adopterinfo[vehicle][1] #Use actual position - apparently bad when combined with non-adopters
                    #lanepos = min(lengths[lane], speeds[lane.split("_")[0]] * (simtime - detecttime + 0.5)+simdetectordist) #+0.5 because we crossed the detector, then made somewhere between 0 and 1 seconds worth of forward movement; estimate it as 0.5
                    newroute = routeFromHere(vehicle)#currentRoutes[vehicle]
                except:
                    print("Error when getting adopter info? Skipping and hoping for the best")
                    continue #Off network, or error when grabbing adopter info or something else weird?
                if not lane.split("_")[0] in nonExitEdgeDetections:
                    continue #In intersection or exit road

            if not vehicle+"."+str(simtime) in traci.route.getIDList():
                traci.route.add(vehicle+"."+str(simtime), newroute) #Using vehicle.time as the name of the new route, to maximize confusion for future me! (Also so we're guaranteed a unique name for the route)

            try:
                traci.vehicle.add(vehicle, vehicle+"."+str(simtime), departLane=lane.split("_")[-1], departPos=lanepos, departSpeed="max")
                newEdgeDict[vehicle] = lane.split("_")[0]
                newLaneDict[vehicle] = lane
            except Exception as e:
                print("Warning: Invalid departLane for no apparent reason?")
                if not vehicle in isSmart or not isSmart[vehicle] or not adopterCommsRouting:
                    print("Not using adopter comms on this vehicle")
                else:
                    print(adopterinfo[vehicle])
                print(e)
                try:
                    traci.vehicle.add(vehicle, vehicle+"."+str(simtime), departPos=lanepos, departSpeed="max")
                    newEdgeDict[vehicle] = lane.split("_")[0]
                    newLaneDict[vehicle] = lane #Might not be perfect but should be close
                except Exception as e:
                    print("Error: Duplicate vehicle?")

            if vehicle in isSmart and isSmart[vehicle]:

                startroute = currentRoutes[vehicle]
                startedge = edgeDict[vehicle]
                startind = startroute.index(startedge)
                startroute = startroute[startind:]
                traci.vehicle.setRoute(vehicle, startroute)

    #Load light state
    with open("savestates/lightstate_"+prevedge+".pickle", 'rb') as handle:
        lightStates = pickle.load(handle)

    #Randomize non-adopter routes
    #TODO thought I did this when inserting them?
    for lane in lanes:
        if len(lane) == 0 or lane[0] == ":":
            continue
        for vehicle in traci.lane.getLastStepVehicleIDs(lane):
            if not vehicle in isSmart or isSmart[vehicle] == False:
                traci.vehicle.setRoute(vehicle, sampleRouteFromTurnData(lane, turndata))

    #Copy traffic light timings
    for light in traci.trafficlight.getIDList():
        traci.trafficlight.setPhase(light, lightStates[light][0])
        traci.trafficlight.setPhaseDuration(light, lightStates[light][1])
    with open("savestates/lightstate_"+prevedge+"_remainingDuration.pickle", 'rb') as handle:
        remainingDuration = pickle.load(handle)
    with open("savestates/lightstate_"+prevedge+"_lastSwitchTimes.pickle", 'rb') as handle:
        lastSwitchTimes = pickle.load(handle)
    with open("savestates/lightstate_"+prevedge+"_sumoPredClusters.pickle", 'rb') as handle:
        sumoPredClusters = pickle.load(handle)
    with open("savestates/lightstate_"+prevedge+"_lightphases.pickle", 'rb') as handle:
        lightphases = pickle.load(handle)
    return (remainingDuration, lastSwitchTimes, sumoPredClusters, lightphases, newEdgeDict, newLaneDict)

#Magically makes the vehicle lists stop deleting themselves somehow???
def dontBreakEverything():
    if not useLibsumo:
        traci.switch("test")
        traci.simulationStep()
        traci.switch("main")

def routeFromHere(vehicle):
    route = currentRoutes[vehicle]
    edge = edgeDict[vehicle]
    edgeind = route.index(edge)
    return route[edgeind:]

# this is the main entry point of this script
if __name__ == "__main__":
    if len(sys.argv) >= 3:
        pSmart = float(sys.argv[2])
    if len(sys.argv) >= 4:
        useLastRNGState = sys.argv[3]
    if len(sys.argv) >= 5:
        appendTrainingData = sys.argv[4]
    main(sys.argv[1], pSmart, True, useLastRNGState, appendTrainingData)
