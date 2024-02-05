#!/usr/bin/env python

import os
import sys
import optparse
import random
from numpy import inf
import numpy as np
#import time
import matplotlib.pyplot as plt
from datetime import datetime

import sumolib #To query node/edge stuff about the network
import pickle #To save/load training data

import torch
from torch import nn
from torch.utils.data import Dataset, DataLoader

#import runnerQueueSplit18NN
import runnerQueueSplit19SUMOEverywhereGC as runnerQueueSplit18NN
from importlib import reload
from Net import Net

import openpyxl #For writing training data to .xlsx files
import time

#In case we want to pause a run and continue later, set these to false
reset = False
resetNN = reset
resetTrainingData2 = reset
#Remember to set appendTrainingData = True, resetTrainingData = False in runnerQueueSplitWhatever
#Also set testDumbtrac (there), testSurtrac (below) and FTP (there) appropriately, and surtracFreq = 1ish

testSurtrac = True

nEpochs = 10
nDaggers = 100

agents = dict()
optimizers = dict()

actions = [0, 1]
learning_rate = 0.0005
batch_size = 1

nLossesBeforeReset = 1000
losses = dict()
epochlosses = dict()    
daggertimes = dict()

loss_fn = torch.nn.MSELoss() #torch.nn.CrossEntropyLoss(weight=torch.Tensor([1, 100]))


#Neural net things
LOG_FILE = 'imitate.log' # Logs will be appended every time the code is run.
MODEL_FILES = dict()
def log(*args, **kwargs):
    if LOG_FILE:
        with open(LOG_FILE, 'a+') as f:
            print(*args, file=f, **kwargs)
    print(*args, **kwargs)

# class TrafficLoader(Dataset):
#     def __init__(self, datafile, light):
#         """
#         Arguments:
#             csv_file (string): Path to the csv file with annotations.
#             root_dir (string): Directory with all the images.
#             transform (callable, optional): Optional transform to be applied
#                 on a sample.
#         """
#         self.datafile = datafile
#         self.light = light
#         with open(datafile, 'rb') as handle:
#             temp = pickle.load(handle) #TODO: This is going to reload the pickle file for each light - this is slow
#         self.dataset = temp[light]

#     def __len__(self):
#         return len(self.dataset)

#     def __getitem__(self, idx):
#         if torch.is_tensor(idx):
#             idx = idx.tolist()

#         datapt = self.dataset[idx]
#         print(datapt)
#         sample = {'input': datapt[0], 'output': datapt[1]}

#         return sample

def readSumoCfg(sumocfg):
    netfile = ""
    roufile = ""
    with open(sumocfg, "r") as cfgfile:
        lines = cfgfile.readlines()
        for line in lines:
            if "net-file" in line:
                data = line.split('"')
                netfile = data[1]
            if "route-files" in line: #This is scary - probably breaks if there's many of them
                data = line.split('"')
                roufile = data[1]
    return (netfile, roufile)

def mainold(sumoconfig):
    global daggertimes
    if resetNN:
        print("Archiving old models.")
        lights = []
        (netfile, routefile) = readSumoCfg(sumoconfig)
        for node in sumolib.xml.parse(netfile, ['junction']):
            if node.type == "traffic_light":
                lights.append(node.id)
        for light in lights:
            try:
                os.rename("models/imitate_" + light + ".model", "models/Archive/imitate_" + light + str(datetime.now()) + ".model")
            except FileNotFoundError:
                print("No model found for light " + light + ", this is fine")

    if resetTrainingData2:
        try:
            print("Archiving old training data.")
            os.rename("trainingdata/trainingdata_" + sys.argv[1] + ".pickle", "trainingdata/Archive/trainingdata_" + sys.argv[1] + str(datetime.now()) + ".pickle")
        except FileNotFoundError:
            print("Nothing to archive, this is fine")
            pass

    firstIter = True
    #DAgger loop
    while True: #for daggernum in range(nDaggers):

        #Get new training data
        #IMPORTANT: Make sure runnerQueueSplit18NN is set to testNN=True, testDumbtrac= not testSurtrac, resetTrainingData=False, appendTrainingData=True
        if not(firstIter and not resetTrainingData2): #If first iteration and we already have training data, start by training on what we have
            print("Generating new training data")
            reload(runnerQueueSplit18NN)
            runnerQueueSplit18NN.main(sys.argv[1], 0, False, False, True)

        #Load current dataset
        print("Loading training data")
        with open("trainingdata/trainingdata_" + sys.argv[1] + ".pickle", 'rb') as handle:
            trainingdata = pickle.load(handle) #List of 2-elt tuples (in, out) = ([in1, in2, ...], out) indexed by light

        dumpTrainingData(trainingdata)
        
        #Set up and train initial neural nets on first iteration
        if firstIter:
            firstIter = False
            #Do NN setup
            for light in trainingdata:
                if testSurtrac:
                    agents[light] = Net(182, 1, 64)
                else:
                    agents[light] = Net(26, 1, 32)
                optimizers[light] = torch.optim.Adam(agents[light].parameters(), lr=learning_rate)
                MODEL_FILES[light] = 'models/imitate_'+light+'.model' # Once your program successfully trains a network, this file will be written
                if not resetNN:
                    try:
                        agents[light].load(MODEL_FILES[light])
                    except:
                        print("Warning: Model " + light + " not found, starting fresh")
                losses[light] = []
                epochlosses[light] = []
                daggertimes[light] = []
        
        #Train everything once - note that all losses are likely to spike after new training data comes in
        for light in trainingdata:
            trainLight(light, trainingdata)

        #New plan: Instead of training for a set number of epochs, we'll train the worst light until it stops improving, then run DAgger again
        while True:
            worstlight = None
            worstloss = -inf
            for light in trainingdata:
                if worstlight == None or epochlosses[light][-1] > worstloss:
                    worstlight = light
                    worstloss = epochlosses[light][-1]
            trainLight(worstlight, trainingdata)
            #If worst light didn't improve, break and get more data
            if epochlosses[worstlight][-1] >= epochlosses[worstlight][-2]:
                break #And get new data

        #Loop back to start the next DAgger loop, store when we did this for plotting purposes
        for light in trainingdata:
            daggertimes[light].append(len(losses[light]))

#Train the given light for a single epoch
def trainLight(light, trainingdata):
    global losses
    print(light)
    avgloss = 0
    nlosses = 0
    random.shuffle(trainingdata[light]) #Pretty sure this is working as intended (and not shuffling data and target independently - that'd be bad)
    totalloss = 0

    for data in trainingdata[light]:
        
        outputNN = agents[light](data[0]) # Output from NN
        target = torch.tensor([[float(data[1])]]) #torch.tensor([float(data[1])]) # Target from expert
        loss = loss_fn(outputNN, target) # calculate loss between network action and expert action (Surtrac action)

        avgloss += float(loss.item()) # record loss for printing
        totalloss += float(loss.item())
        nlosses += 1

        if nlosses % nLossesBeforeReset == 0:
            avgloss /= nLossesBeforeReset
            losses[light].append(avgloss)
            log('Iteration: {}   Loss: {}'.format(nlosses,avgloss))
            avgloss = 0
        loss.backward() # perform backprop to collect gradients
        optimizers[light].step() # perform one step of update with optimizer
        optimizers[light].zero_grad() # reset accumulated gradients to 0
    
    epochlosses[light].append(totalloss/len(trainingdata[light]))
    agents[light].save(MODEL_FILES[light])
    plt.figure()
    plt.plot(losses[light])
    for daggertime in daggertimes[light]:
        plt.axvline(x=daggertime, color='k', linestyle='--')
    plt.xlabel("Sets of " + str(nLossesBeforeReset) + " Points")
    plt.ylabel("Average Loss")
    plt.title("Losses, light=" + str(light))
    #plt.show() #NOTE: Blocks code execution until you close the plot
    plt.savefig("Plots/Losses, light=" + str(light)+".png")
    plt.close()

#Dumps training data to an Excel file for human readability
def dumpTrainingData(trainingdata):
    timeout = 60
    starttime = time.time()
    print("Writing training data to spreadsheet")
    try:
        book = openpyxl.Workbook()
        sheets = dict()
        for light in trainingdata:
            sheets[light] = book.create_sheet(light, -1)
            sheets[light].cell(1, 1, "Input")
            row = 2
            for batch in trainingdata[light]:
                if time.time() - starttime > timeout:
                    assert(False) #This should trigger the try-catch and get us out of here
                for linenum in range(batch[0].size(0)):
                    col = 1
                    for entry in batch[0][linenum]:
                        sheets[light].cell(row, col, float(entry))
                        col += 1
                    if row == 2:
                        sheets[light].cell(1, col, "Expert Output")
                    sheets[light].cell(row, col, float(batch[1][linenum]))

                    col += 1
                    if row == 2:
                        sheets[light].cell(1, col, "NN Output")
                    if len(batch) > 2:
                        sheets[light].cell(row, col, float(batch[2][linenum]))
                    row += 1
    except Exception as e:
        print(e)
        print("Error dumping training data to Excel, ignoring and continuing")
    finally:
        book.save("trainingdata/trainingdata_"+sys.argv[1]+".xlsx")

# this is the main entry point of this script
if __name__ == "__main__":
    mainold(sys.argv[1])



# #NOT WORKING; not sure what's going wrong
# #Supposed to use the nice dataset interface, which would let me batch, etc.
# def main(sumoconfig):
#     trafficdataset = dict()
#     dataloader = dict()
#     with open("trainingdata/trainingdata_" + sys.argv[1] + ".pickle", 'rb') as handle:
#         trainingdata = pickle.load(handle)
#     for light in trainingdata:
#         trafficdataset[light] = TrafficLoader("trainingdata/trainingdata_" + sys.argv[1] + ".pickle", light)
#         dataloader[light] = DataLoader(trafficdataset, batch_size = batch_size, shuffle=True, num_workers=0)

#         agents[light] = Net(26, 2, 512) #Net(26, 2, 512)
#         optimizers[light] = torch.optim.Adam(agents[light].parameters(), lr=learning_rate)
#         MODEL_FILES[light] = 'models/imitate_'+light+'.model' # Once your program successfully trains a network, this file will be written
#         if not resetNN:
#             try:
#                 agents[light].load(MODEL_FILES[light])
#             except:
#                 print("Warning: Model not found, starting fresh")
#         losses[light] = []

#     for epochnum in range(nEpochs):
#         print("Going through training data, epoch " + str(epochnum))

#         for light in trainingdata:
#             print(light)
#             avgloss = 0
#             nlosses = 0
            
#             for databatch in dataloader[light]:
                
#                 outputNN = agents[light](databatch['input']) # Output from NN
#                 # Find the best action to take from actions based on the output of the NN
#                 #actionNumber = outputNN.argmax(1).item() # Get best action number
#                 actionNumber = outputNN.argmax().item() # Get best action number
#                 actionNN = actions[actionNumber] # Get best action

#                 target = torch.tensor(databatch['target']) # Target from expert
#                 loss = loss_fn(outputNN, target) # calculate loss between network action and expert action (Surtrac action)

#                 avgloss += float(loss.item()) # record loss for printing
#                 nlosses += 1

#                 if nlosses % nLossesBeforeReset == 0:
#                     avgloss /= nLossesBeforeReset
#                     log('Iteration: {}   Loss: {}'.format(nlosses,avgloss))
#                     avgloss = 0
#                 loss.backward() # perform backprop to collect gradients
#                 optimizers[light].step() # perform one step of update with optimizer
#                 optimizers[light].zero_grad() # reset accumulated gradients to 0
            
#             agents[light].save(MODEL_FILES[light])