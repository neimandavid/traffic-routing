#!/usr/bin/env python

import os
import sys
import optparse
import random
from numpy import inf
import numpy as np
import matplotlib.pyplot as plt
from datetime import datetime

import sumolib #To query node/edge stuff about the network
import pickle #To save/load training data

import torch
from torch import nn
from torch.utils.data import Dataset, DataLoader
import torch.multiprocessing
torch.multiprocessing.set_sharing_strategy("file_system")

#import runnerQueueSplit
import runnerQueueSplit27 as runnerQueueSplit #KEEP THIS UP TO DATE!!! (If training from a network, not just IG)
import intersectionGeneratorBlocks15Pushforward as intersectionGenerator
from importlib import reload
from Net import Net

import openpyxl #For writing training data to .xlsx files
import time

#From: https://www.geeksforgeeks.org/how-to-use-gpu-acceleration-in-pytorch/
if torch.cuda.is_available():
    print(f"GPU: {torch.cuda.get_device_name(0)} is available.")
else:
    print("No GPU available. Training will run on CPU.")

device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
print(device)

#In case we want to pause a run and continue later, set these to false
reset = True
reuseOldData = False
resetNN = reset
resetTrainingData2 = True
superResetTrainingData = True
if superResetTrainingData:
    resetTrainingData2 = True
#UPDATE: Turns out appendTrainingData (there) gets updated automatically, as does noNNInMain
#Also, Surtrac network architecture works for FTPs as well
#So just make sure resetTrainingData=False, testDumbtrac and FTP are correct, and surtracFreq = 1ish (all in runnerQueueSplitWhatever)

crossEntropyLoss = True

agents = dict()
optimizers = dict()
dataloader = dict()

actions = [0, 1]
learning_rate = 0.00005
batch_size = 100

nLossesBeforeReset = 1000/batch_size
losses = dict()
epochlosses = dict()    
daggertimes = dict()

if crossEntropyLoss:
    loss_fn = torch.nn.CrossEntropyLoss()#(weight=torch.Tensor([1, 1.5])) #If training on IG, there'll be a reasonable number of "switch" scenarios
else:
    loss_fn = torch.nn.L1Loss()


#Neural net things
LOG_FILE = 'imitate.log' # Logs will be appended every time the code is run.
MODEL_FILES = dict()
def log(*args, **kwargs):
    if LOG_FILE:
        with open(LOG_FILE, 'a+') as f:
            print(*args, file=f, **kwargs)
    print(*args, **kwargs)


#Based on https://pytorch.org/tutorials/beginner/data_loading_tutorial.html
class TrafficDataset(Dataset):
    def __init__(self, datafile):
        with open(datafile, 'rb') as handle:
            temp = pickle.load(handle)
        self.dataset = temp["light"]

        nstick = 0
        ntotal = 0
        #for item in self.dataset:
        for itemnum in range(len(self.dataset)):
            item = self.dataset[itemnum]
            #We're going to assume all training data is MSE data (in particular, this means I can switch loss fns without retraining)
            #So if we're using cross-entropy loss, we need to convert time-to-switch into should-I-switch-now?
            if crossEntropyLoss:
                if item[1] > 0:
                    self.dataset[itemnum] = (item[0], torch.LongTensor([1]))
                else:
                    self.dataset[itemnum] = (item[0], torch.LongTensor([0]))

            nstick += item[1]
            ntotal += 1
        print("Stick fraction: " + str(nstick/ntotal))

    def __len__(self):
        return len(self.dataset)

    def __getitem__(self, idx):
        if torch.is_tensor(idx):
            idx = idx.tolist()

        datapt = self.dataset[idx]
        sample = {'input': datapt[0], 'target': datapt[1]}
        return sample

def readSumoCfg(sumocfg):
    netfile = ""
    roufile = ""
    with open(sumocfg, "r") as cfgfile:
        lines = cfgfile.readlines()
        for line in lines:
            if "net-file" in line:
                data = line.split('"')
                netfile = data[1]
            if "route-files" in line: #This is scary - probably breaks if there's many of them
                data = line.split('"')
                roufile = data[1]
    return (netfile, roufile)

def main(sumoconfigs):
    global daggertimes
    if resetNN:
        print("Archiving old models.")
        # lights = []
        # for sumoconfig in sumoconfigs:
        #     (netfile, routefile) = readSumoCfg(sumoconfig)
        #     for node in sumolib.xml.parse(netfile, ['junction']):
        #         if node.type == "traffic_light":
        #             lights.append(node.id)
        for light in ["light"]:# + lights:
            try:
                print("models/imitate_" + light + ".model")
                now = datetime.now()

                os.rename("models/imitate_" + light + ".model", "models/Archive/imitate_" + light + datetime.strftime(now, '%m-%d-%Y-%H-%M-%S') + ".model")
            except FileNotFoundError:
                print("No model found for light " + light + ", this is fine")
        
    if resetTrainingData2:
        try:
            print("Archiving old training data.")
            now = datetime.now()
            os.rename("trainingdata/trainingdata_" + sys.argv[1] + ".pickle", "trainingdata/Archive/trainingdata_" + sys.argv[1] + datetime.strftime(now, '%m-%d-%Y-%H-%M-%S') + ".pickle")
        except FileNotFoundError:
            print("Nothing to archive, this is fine")
            pass

    #Do NN setup
    for light in ["light"]:#trainingdata:
        maxnlanes = 3 #Going to assume we have at most 3 lanes per road, and that the biggest number lane is left-turn only
        maxnroads = 4 #And assume 4-way intersections for now
        maxnclusters = 5 #And assume at most 10 clusters per lane
        ndatapercluster = 3 #Arrival, departure, weight
        maxnphases = 12 #Should be enough to handle both leading and lagging lefts
        
        nextra = 1 #Remaining time in current lane, current time (garbage, set to 0)
        ninputs = maxnlanes*maxnroads*maxnclusters*ndatapercluster + maxnlanes*maxnroads*maxnphases + maxnphases + nextra #180+144+12+1=337

        if crossEntropyLoss:
            agents[light] = Net(ninputs, 2, 8192)
        else:
            agents[light] = Net(ninputs, 1, 8192)

        try:
            agents[light] = agents[light].to(device)
        except:
            print("Error sending model to " + device)
            pass
        
        optimizers[light] = torch.optim.Adam(agents[light].parameters(), lr=learning_rate)
        MODEL_FILES[light] = 'models/imitate_'+light+'.model' # Once your program successfully trains a network, this file will be written
        if not resetNN:
            # try:
            #     agents[light].load(MODEL_FILES[light])
            # except FileNotFoundError as e:
            #     print(e)
            #     print("Warning: Model " + light + " not found? Starting fresh")

            #Adapted from: https://pytorch.org/tutorials/beginner/saving_loading_models.html#saving-loading-a-general-checkpoint-for-inference-and-or-resuming-training
            
            try:
                checkpoint = torch.load(MODEL_FILES[light], weights_only=True)
                agents[light].load_state_dict(checkpoint['model_state_dict'])
                optimizers[light].load_state_dict(checkpoint['optimizer_state_dict'])
            except FileNotFoundError as e:
                    print(e)
                    print("Warning: Model " + light + " not found? Starting fresh")

        losses[light] = []
        epochlosses[light] = []
        daggertimes[light] = []

    #Read old data, because we can!
    if reuseOldData:
        directory = "trainingdata/Archive"
            
        for _ in range(1):
            count = 0
            for file in os.listdir(directory):
                filename = os.fsdecode(file)
                if filename.endswith(".pickle"): 
                    print(directory + "/" + filename)#os.path.join(directory, filename))
                    print("Loading training data")
                    #try:
                    trafficdataset = TrafficDataset(directory + "/" + filename)
                    #Train everything once - note that all losses are likely to spike after new training data comes in
                    for light in ["light"]:#trainingdata:
                        count += 1
                        trainLight(light, trafficdataset, count%10 == 0)

                    # except FileNotFoundError as e:
                    #     pass
            

    firstIter = True
    #DAgger loop
    while True:

        if superResetTrainingData:
            # try:
            #     os.remove("trainingdata/trainingdata_" + sys.argv[1] + ".pickle")
            # except FileNotFoundError:
            #     print("Super reset failed, this is probably fine")
            try:
                print("Archiving old training data.")
                now = datetime.now()
                os.rename("trainingdata/trainingdata_" + sys.argv[1] + ".pickle", "trainingdata/Archive/trainingdata_" + sys.argv[1] + datetime.strftime(now, '%m-%d-%Y-%H-%M-%S') + ".pickle")
            except FileNotFoundError:
                print("Nothing to archive, this is fine")
                pass

        #Get new training data
        if not(firstIter and not resetTrainingData2): #If first iteration and we already have training data, start by training on what we have
            print("Generating new training data")
            for sumoconfig in sumoconfigs:
                if sumoconfig == "IG":
                    #reload(intersectionGenerator)
                    intersectionGenerator.main()
                else:
                    reload(runnerQueueSplit)
                    runnerQueueSplit.main(sumoconfig, 0, False, False, True)
                with open("trainingdata/trainingdata_" + sys.argv[1] + ".pickle", 'rb') as handle:
                    trainingdata = pickle.load(handle)
            dumpTrainingData(trainingdata)

        #Load current dataset
        print("Loading training data")
        try:
            trafficdataset = TrafficDataset("trainingdata/trainingdata_" + sys.argv[1] + ".pickle")

        except FileNotFoundError as e:
            #No data, so generate some, then loop back
            firstIter = False
            continue
        
        #Train everything once - note that all losses are likely to spike after new training data comes in
        for light in ["light"]:#trainingdata:
            trainLight(light, trafficdataset)

        #New plan: Instead of training for a set number of epochs, we'll train the worst light until it stops improving, then run DAgger again
        while True and not superResetTrainingData:
            worstlight = None
            worstloss = -inf
            for light in ["light"]:#trainingdata:
                if worstlight == None or epochlosses[light][-1] > worstloss:
                    worstlight = light
                    worstloss = epochlosses[light][-1]
            trainLight(worstlight, trafficdataset)
            #If worst light didn't improve, break and get more data
            if epochlosses[worstlight][-1] >= epochlosses[worstlight][-2]:
                break #And get new data
            if epochlosses[worstlight][-1] < 1e-5:
                break #Probably good enough, get new data rather than continuously overfitting

        #Loop back to start the next DAgger loop, store when we did this for plotting purposes
        for light in ["light"]:#trainingdata:
            daggertimes[light].append(len(losses[light]))

#Train the given light for a single epoch
def trainLight(light, dataset, saveModel = True):
    #global losses
    print(light)
    avgloss = 0
    nlosses = 0
    totalloss = 0
    dataloader = DataLoader(dataset, batch_size = batch_size, shuffle=True, num_workers=0)
    print("End load")
    for i, databatch in enumerate(dataloader):
        if torch.cuda.is_available():
            databatch['input'] = databatch['input'].to('cuda')
            databatch['target'] = databatch['target'].to('cuda')
        outputNN = agents[light](databatch['input']).flatten(1) # Output from NN
        target = databatch['target'].clone().detach() # Target from expert
        if crossEntropyLoss:
            target = target.flatten()
        loss = loss_fn(outputNN, target) # calculate loss between network action and expert action (Surtrac action)

        avgloss += float(loss.item()) # record loss for printing. Pretty sure this is an average over the minibatch
        totalloss += float(loss.item())
        nlosses += 1

        if nlosses % nLossesBeforeReset == 0:# <= 1:#(nlosses-batch_size) % nLossesBeforeReset:
            avgloss /= (nLossesBeforeReset)
            losses[light].append(avgloss)
            log('Iteration: {}   Loss: {}'.format(nlosses*batch_size,avgloss))
            avgloss = 0
        loss.backward() # perform backprop to collect gradients
        optimizers[light].step() # perform one step of update with optimizer
        optimizers[light].zero_grad() # reset accumulated gradients to 0
    
    epochlosses[light].append(totalloss/dataset.__len__())
    #agents[light].save(MODEL_FILES[light]) #OLD

    if saveModel:
        print("Saving updated model")
        #Apparently saving is somewhat slow and mid-save the file gets messed up; to fix this, save to a temp file, then move the temp file when done
        torch.save({
                'model_state_dict': agents[light].state_dict(),
                'optimizer_state_dict': optimizers[light].state_dict()
                }, MODEL_FILES[light]+"saving")
        #os.remove(MODEL_FILES[light]) #Because apparently just saving over this moves the old file to trash, and that's annoying and eats space?
        os.rename(MODEL_FILES[light]+"saving", MODEL_FILES[light])

        plt.figure()
        plt.plot(losses[light])
        #Vertical lines each time we get new data - annoying since we don't reuse data now
        # for daggertime in daggertimes[light]:
        #     plt.axvline(x=daggertime, color='k', linestyle='--')
        plt.xlabel("Sets of " + str(nLossesBeforeReset*batch_size) + " Points")
        if crossEntropyLoss:
            plt.ylabel("Average Loss (CrossEntropy)")
        else:
            plt.ylabel("Average Loss (MeanSquaredError)")
        plt.title("Losses, light=" + str(light))
        #plt.show() #NOTE: Blocks code execution until you close the plot
        plt.savefig("Plots/Losses, light=" + str(light)+".png")
        plt.close()

#Dumps training data to an Excel file for human readability
def dumpTrainingData(trainingdata):
    timeout = 5#60
    starttime = time.time()
    print("Writing training data to spreadsheet")
    try:
        book = openpyxl.Workbook()
        sheets = dict()
        for light in trainingdata:
            sheets[light] = book.create_sheet(light, -1)
            sheets[light].cell(1, 1, "Input")
            row = 2
            for batch in trainingdata[light]:
                if time.time() - starttime > timeout:
                    print("Timed out while writing data to Excel")
                    assert(False) #This should trigger the try-catch and get us out of here
                for linenum in range(batch[0].size(0)):
                    col = 1
                    for entry in batch[0][linenum]:
                        sheets[light].cell(row, col, float(entry))
                        col += 1
                    if row == 2:
                        sheets[light].cell(1, col, "Expert Output")
                    sheets[light].cell(row, col, float(batch[1][linenum]))

                    col += 1
                    if row == 2:
                        sheets[light].cell(1, col, "NN Output")
                    if len(batch) > 2:
                        sheets[light].cell(row, col, float(batch[2][linenum])) #TODO this is wrong for cross-entropy loss
                    row += 1
    except Exception as e:
        print(e)
        print("Error dumping training data to Excel, ignoring and continuing")
    finally:
        book.save("trainingdata/trainingdata_"+sys.argv[1]+".xlsx")



# this is the main entry point of this script
if __name__ == "__main__":
    main(sys.argv[1:])